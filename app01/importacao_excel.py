# -*- coding: utf-8 -*-

from django.http import HttpResponseRedirect
import openpyxl
import os
import sys
import datetime
import unicodedata
from .models import Funcionario,Provento,Complemento,Ampliacao_ch,Suporte,SemCadastro,Information,Tabela,Secretaria,Setor,Funcao,Evento,Vinculo,CargaHoraria
from . import listagens,dicionarios


def import_excel_gravarFuncao_passo1(excel,idmunicipio):

    lst_secs=[sec.secretaria for sec in Secretaria.objects.filter(id_municipio=idmunicipio)]
    lst_sets=[sec.setor for sec in Setor.objects.filter(id_municipio=idmunicipio)]
    lst_funcs=[sec.funcao for sec in Funcao.objects.filter(id_municipio=idmunicipio)]
    lst_evts=[sec.evento for sec in Evento.objects.filter(id_municipio=idmunicipio)]
    lst_vinculos=[vic.vinculo for vic in Vinculo.objects.filter(id_municipio=idmunicipio)]
    lst_cargah=[ch.cargah_origem for ch in CargaHoraria.objects.filter(id_municipio=idmunicipio)]

    wb = openpyxl.load_workbook(excel)
    sheets = wb.sheetnames

    #sheet = sheets[0]
    sheet = wb.active
    nrows = get_maximum_rows(sheet_object=sheet)
    max_col = sheet.max_column
    lista_new_information=[]

    lista_sec_descricao=[]
    lista_set_descricao=[]
    lista_fun_descricao=[]
    lista_evt_descricao=[]
    lista_vic_descricao=[]
    lista_ch_descricao=[]

    ls_set_descricao=[]
    ls_fun_descricao=[]
    ls_evt_descricao=[]
    ls_vic_descricao=[]
    ls_ch_descricao=[]


    lista_sec_descricao_s=[]
    lista_set_descricao_s=[]
    lista_fun_descricao_s=[]
    lista_evt_descricao_s=[]

    lista_secs_12=[]
    lista_sets_12=[]
    lista_funcs_12=[]
    lista_evts_12=[]


    lista_objeto=[]
    id_cidade=0

    row=1
    limite=500
    maximo=0
    while row<limite and row<nrows+1:
        if row==1:
            for i in range(1, max_col + 1):
                maximo+=1
                cell_obj = sheet.cell(row = 1, column = i)
                titulo_coluna=cell_obj.value
                if titulo_coluna is not None and 1==1:
                    titulo_coluna=remove_combining_fluent(titulo_coluna)
                    titulo_coluna=titulo_coluna.strip()
                    titulo_coluna=titulo_coluna.upper()
                    if titulo_coluna=='MUNICIPIO':
                        n_municipio=letra_da_coluna(i)
                    elif titulo_coluna=='TABELA':
                        n_tabela=letra_da_coluna(i)
                    elif titulo_coluna=='DESCRICAO_QCHEGA':
                        n_descricao_e=letra_da_coluna(i)
                    elif titulo_coluna=='DESCRICAO_QSAI':
                        n_descricao_s=letra_da_coluna(i)
                    elif titulo_coluna=='COMPLEMENTO':
                        n_complemento=letra_da_coluna(i)
                    elif titulo_coluna=='CARATER-SALARIO':
                        n_caraterSalario=letra_da_coluna(i)
                    elif titulo_coluna=='POS-GRADUACAO':
                        n_posGraduacao=letra_da_coluna(i)

                if maximo>10:
                    break

            row+=1
            continue

        if sheet['B'+str(row)].value is None:
            break
        if sheet['B'+str(row)].value=='':
            break

        municipio = sheet[n_municipio + str(row)].value
        tabela = sheet[n_tabela + str(row)].value

        descricao_e = sheet[n_descricao_e + str(row)].value
        descricao_s = sheet[n_descricao_s + str(row)].value
        complemento = sheet[n_complemento + str(row)].value
        complemento2 = sheet[n_caraterSalario + str(row)].value
        complemento3 = sheet[n_posGraduacao + str(row)].value

        if descricao_e is None or  municipio is None:
            row+=1
            continue

        if descricao_s is None:
            descricao_s=''

        if complemento is None:
            complemento=''

        if complemento2 is None:
            complemento2=''

        if complemento3 is None:
            complemento3=''


        municipio=remove_combining_fluent(municipio)
        tabela=remove_combining_fluent(tabela)

        tabela=tabela.strip()
        tabela=tabela.lower()

        municipio=municipio.strip()
        municipio=municipio.upper()


        id_municipio=fun_id_municipio(municipio)
        if id_municipio!=idmunicipio:
            row+=1
            continue

        if tabela not in ['funcao','evento','setor','secretaria','vinculo','ch']:
            row+=1
            continue

        if tabela in ['funcao','evento','setor','secretaria','vinculo']:
            descricao_e=remove_combining_fluent(descricao_e)
            descricao_s=remove_combining_fluent(descricao_s)
            descricao_e = descricao_e.strip()
            descricao_s = descricao_s.strip()
            descricao_e =  descricao_e.upper()
            descricao_s =  descricao_s.upper()
            complemento = complemento.strip()
            complemento = complemento.upper()
            complemento2 = complemento2.strip()
            complemento2 = complemento2.upper()
            complemento3 = complemento3.strip()
            complemento3 = complemento3.upper()


        if complemento=='EXCLUIR':
            row+=1
            continue


        ampliacao='N'
        suporte='N'
        professor='N'
        fundeb='N'
        grupo='N'
        caraterSalario='N'
        posGraduacao='N'

        if complemento!='':
            if tabela=='evento':
                if complemento=='A':
                    ampliacao='S'
                elif complemento=='S':
                    suporte='S'
                else:
                    if complemento=='AS' or complemento=='SA':
                        ampliacao='S'
                        suporte='S'
            elif tabela=='funcao':
                if complemento=='P':
                    professor='S'
            elif tabela=='setor':
                if complemento=='F':
                    fundeb='S'
            elif tabela=='vinculo':
                if complemento=='EFETIVO':
                    grupo='E'
                elif complemento=='TEMPORARIO':
                    grupo='T'
                elif complemento=='COMISSIONADO':
                    grupo='C'

        if tabela=='evento':
            if complemento2=='SIM' or complemento2=='S':
                caraterSalario='S'
            if complemento3=='SIM' or complemento3=='S':
                posGraduacao='S'



        id_cidade=id_municipio


        if descricao_e==descricao_s:
            descricao_s=''

        if tabela=='secretaria':
            if descricao_e not in lst_secs:
                if descricao_e not in lista_sec_descricao :
                    lista_sec_descricao.append(descricao_e)

            if descricao_s!='':
                if descricao_s not in lst_secs:
                    if descricao_s not in lista_sec_descricao:
                        lista_sec_descricao.append(descricao_s)
            if descricao_s!='':
                if descricao_e+':'+descricao_s not in lista_secs_12:
                    lista_secs_12.append(descricao_e+':'+descricao_s)

        if tabela=='setor':
            if descricao_e not in lst_sets:
                if descricao_e not in lista_set_descricao:
                    lista_set_descricao.append(descricao_e)
                    ls_set_descricao.append([descricao_e,fundeb])

            if descricao_s!='':
                if descricao_s not in lst_sets:
                    if descricao_s not in lista_set_descricao:
                        lista_set_descricao.append(descricao_s)
                        ls_set_descricao.append([descricao_s,fundeb])

            if descricao_s!='':
                if descricao_e+':'+descricao_s not in lista_sets_12:
                    lista_sets_12.append(descricao_e+':'+descricao_s)



        if tabela=='funcao':
            if descricao_e not in lst_funcs:
                if descricao_e not in lista_fun_descricao:
                    lista_fun_descricao.append(descricao_e)
                    ls_fun_descricao.append([descricao_e,professor])

            if descricao_s!='':
                if descricao_s not in lst_funcs:
                    if descricao_s not in lista_fun_descricao:
                        lista_fun_descricao.append(descricao_s)
                        ls_fun_descricao.append([descricao_s,professor])

            if descricao_s!='':
                if descricao_e+':'+descricao_s not in lista_funcs_12:
                    lista_funcs_12.append(descricao_e+':'+descricao_s)


        if tabela=='evento':
            if descricao_e not in lst_evts:
                if descricao_e not in lista_evt_descricao:
                    lista_evt_descricao.append(descricao_e)
                    ls_evt_descricao.append([descricao_e,ampliacao,suporte,caraterSalario,posGraduacao])

            if descricao_s!='':
                if descricao_s not in lst_evts:
                    if descricao_s not in lista_evt_descricao:
                        lista_evt_descricao.append(descricao_s)
                        ls_evt_descricao.append([descricao_s,ampliacao,suporte,caraterSalario,posGraduacao])

            if descricao_s!='':
                if descricao_e+':'+descricao_s not in lista_evts_12:
                    lista_evts_12.append(descricao_e+':'+descricao_s)

        if tabela=='vinculo':
            if descricao_e not in lst_vinculos:
                if descricao_e not in lista_vic_descricao:
                    lista_vic_descricao.append(descricao_e)
                    ls_vic_descricao .append([descricao_e,grupo])


        if tabela=='ch':
            if descricao_s=='':
                descricao_s=descricao_e
            if descricao_e not in lst_cargah:
                if descricao_e not in lista_ch_descricao:
                    lista_ch_descricao.append(descricao_e)
                    ls_ch_descricao .append([descricao_e,descricao_s])


        row+=1

    lista_objeto=[]
    for descricao in lista_sec_descricao:
        new_objeto=Secretaria(
            id_municipio = id_cidade,
            secretaria = descricao,
            secretaria_out = descricao
        )
        lista_objeto.append(new_objeto)

    if len(lista_objeto)>0:
        Secretaria.objects.bulk_create(lista_objeto)

    lista_objeto=[]
    for kls in ls_set_descricao:
        descricao=kls[0]
        fundeb=kls[1]
        new_objeto=Setor(
            id_municipio = id_cidade,
            setor = descricao,
            setor_out = descricao,
            fundeb=fundeb
        )
        lista_objeto.append(new_objeto)

    if len(lista_objeto)>0:
        Setor.objects.bulk_create(lista_objeto)

    lista_objeto=[]
    if len(lista_fun_descricao)>0:
        for k in range(len(ls_fun_descricao)):
            descricao=ls_fun_descricao[k][0]
            professor=ls_fun_descricao[k][1]
            new_objeto= Funcao (
                id_municipio = id_cidade,
                funcao = descricao,
                funcao_out = descricao,
                professor=professor
            )
            lista_objeto.append(new_objeto)


        if len(lista_objeto)>0:
            Funcao.objects.bulk_create(lista_objeto)

    lista_objeto=[]
    for kls in ls_evt_descricao:
        descricao=kls[0]
        ampliacao=kls[1]
        suporte=kls[2]
        caraterSalario=kls[3]
        posGraduacao=kls[4]
        new_objeto= Evento (
            id_municipio = id_cidade,
            evento = descricao,
            evento_out = descricao,
            ampliacao_ch=ampliacao,
            suporte=suporte,
            carater_salario=caraterSalario,
            posgraduacao=posGraduacao
        )

        lista_objeto.append(new_objeto)

    if len(lista_objeto)>0:
        Evento.objects.bulk_create(lista_objeto)

    lista_objeto=[]
    if len(lista_vic_descricao)>0:
        for k in range(len(ls_vic_descricao)):
            descricao=ls_vic_descricao[k][0]
            grupo=ls_vic_descricao[k][1]
            new_objeto= Vinculo (
                id_municipio = id_cidade,
                vinculo = descricao,
                grupo = grupo
            )
            lista_objeto.append(new_objeto)

        if len(lista_objeto)>0:
            Vinculo.objects.bulk_create(lista_objeto)


    lista_objeto=[]
    for kls in range(len(ls_ch_descricao)):
        ch_origem=kls[0]
        ch_convertida=kls[1]
        new_objeto= CargaHoraria (
            id_municipio = id_cidade,
            cargah_origem = ch_origem,
            cargah_convertida = ch_convertida
        )
        lista_objeto.append(new_objeto)

    if len(lista_objeto)>0:
        CargaHoraria.objects.bulk_create(lista_objeto)



    if len(lista_secs_12)>0:
        retorno = ajustes1_tabela(id_cidade,lista_secs_12,'secretaria')
    if len(lista_sets_12)>0:
        retorno = ajustes1_tabela(id_cidade,lista_sets_12,'setor')
    if len(lista_funcs_12)>0:
        retorno = ajustes1_tabela(id_cidade,lista_funcs_12,'funcao')
    if len(lista_evts_12)>0:
        retorno = ajustes1_tabela(id_cidade,lista_evts_12,'evento')


    if len(lista_sec_descricao)>0:
        retorno = ajustes2_tabela(id_cidade,'secretaria')
    if len(lista_set_descricao)>0:
        retorno = ajustes2_tabela(id_cidade,'setor')
    if len(lista_fun_descricao)>0:
        retorno = ajustes2_tabela(id_cidade,'funcao')
    if len(lista_evt_descricao)>0:
        retorno = ajustes2_tabela(id_cidade,'evento')

    '''
    if len(lista_secs_12)>0:
        retorno = ajustes3_tabela(id_cidade,'secretaria',lista_secs_12)
    '''


    return 0


def fun_id_municipio(pmunicipio):
    if pmunicipio=='CANINDE':
        return 15
    if pmunicipio=='TIANGUA':
        return 16
    if pmunicipio=='SAO GONCALO':
        return 38
    if pmunicipio=='PEDRA BRANCA':
        return 44
    if pmunicipio=='CARIDADE':
        return 86
    if pmunicipio=='ITATIRA':
        return 92
    if pmunicipio=='QUIXELO':
        return 124
    if pmunicipio=='PALHANO':
        return 163
    if pmunicipio=='MOMBACA':
        return 42
    if pmunicipio=='GUAIUBA':
        return 74
    if pmunicipio=='COREAU':
        return 85
    if pmunicipio=='SOLONOPOLE':
        return 109
    if pmunicipio=='GRACA':
        return 134
    if pmunicipio=='OCARA':
        return 182
    if pmunicipio=='ALTO SANTO':
        return 125
    if pmunicipio=='ARATUBA':
        return 152
    if pmunicipio=='INDEPENDENCIA':
        return 76
    if pmunicipio=='IRAPUAN PINHEIRO':
        return 162
    if pmunicipio=='ITAPIUNA':
        return 98
    if pmunicipio=='JAGUARIBE':
        return 57





def remove_combining_fluent(string: str) -> str:
    normalized = unicodedata.normalize('NFD', string)
    return ''.join(
        [l for l in normalized if not unicodedata.combining(l)]
    )


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def fun_salario_cadastrado(salario,lista_de_salarios):
    for k in range(0,8):
        s0=str(round(salario+(k/100),2))
        (p1,p2)=s0.split('.')
        if len(p2)==1:
            s0=s0+'0'
        if s0 in lista_de_salarios:
            return s0
    for k in range(0,8):
        s0=str(round(salario-(k/100),2))
        (p1,p2)=s0.split('.')
        if len(p2)==1:
            s0=s0+'0'
        if s0 in lista_de_salarios:
            return s0
    return 0


def fun_pesquisa_salario(salario,lista_de_salarios_int,lista_de_salarios_dec):
    valor = int(salario)
    retorno = 0
    if valor in lista_de_salarios_int:
        return lista_de_salarios_dec[lista_de_salarios_int.index(valor)]

    valor=int(salario)-1;
    if valor in lista_de_salarios_int:
        return lista_de_salarios_dec[lista_de_salarios_int.index(valor)]

    valor=int(salario)+1;
    if valor in lista_de_salarios_int:
        return lista_de_salarios_dec[lista_de_salarios_int.index(valor)]
    return salario



def eventos_tabela_complementos(id_municipio):
    if id_municipio==38:
        return [1,9,227]
    else:
        return [1]

def fun_calcula_carga_horaria(pcarga_horaria):
    if pcarga_horaria==20:
        return 100
    elif pcarga_horaria==30:
        return 150
    elif pcarga_horaria==40:
        return 200
    elif pcarga_horaria==50:
        return 100
    elif pcarga_horaria==220:
        return 200
    return pcarga_horaria



def codigoPrevidencia(texto):
    if texto=='PREVIDENCIA MUNICIPAL':
        previdencia='M'
    elif texto=='INSS':
        previdencia='I'
    elif texto=='NÃO PAGA':
        previdencia='N'
    else:
        previdencia=''

    return previdencia


def eventos_tabela_complementos(id_municipio):
    if id_municipio==38:
        return [1,9,227]
    else:
        return [1]


def fun_salario100h(psalario_100h,plista_sal100,plista_sal200):
    sal100_0 = int(psalario_100h)-1
    sal100_1 = int(psalario_100h)
    sal100_2 = int(psalario_100h)+1

    for kls in plista_sal100:
        if kls==sal100_0:
            return kls

    for kls in plista_sal100:
        if kls == sal100_1:
            return kls

    for kls in plista_sal100:
        if kls == sal100_2:
            return kls

    return psalario_100h


def letra_da_coluna(pi):
    lista1=['A','B','C']
    lista2=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    lista3=[]
    for alfa in lista1:
        for beta in lista2:
            lista3.append(alfa+beta)
    listaColunas=lista2+lista3
    return listaColunas[pi-1]



def ajustes1_tabela(pid_municipio,plista_descricao_out,tabela):
    lista1=[]
    lista2=[]
    lista3=[]

    for k in range(len(plista_descricao_out)):
        (descricao1,descricao_out1) = plista_descricao_out[k].split(':')
        lista1.append(descricao_out1)
        lista3=[]
        for kk in range(len(plista_descricao_out)):
            (descricao2,descricao_out2) = plista_descricao_out[kk].split(':')
            if descricao_out1==descricao_out2:
                lista3.append(descricao2)
        lista2.append(lista3)

    for k in range(len(lista1)):
        descricao=lista1[k]

        if tabela=='secretaria':
            obj=Secretaria.objects.filter(id_municipio=pid_municipio,secretaria=descricao).first()
            id_apelido=obj.id_secretaria
            Secretaria.objects.filter(id_municipio=pid_municipio,secretaria__in=lista2[k]).update(id_secretaria_out=id_apelido,secretaria_out=descricao)


        if tabela=='setor':
            obj=Setor.objects.filter(id_municipio=pid_municipio,setor=descricao).first()
            id_apelido=obj.id_setor
            Setor.objects.filter(id_municipio=pid_municipio,setor__in=lista2[k]).update(id_setor_out=id_apelido,setor_out=descricao)


        if tabela=='funcao':
            obj=Funcao.objects.filter(id_municipio=pid_municipio,funcao=descricao).first()
            id_apelido=obj.id_funcao
            Funcao.objects.filter(id_municipio=pid_municipio,funcao__in=lista2[k]).update(id_funcao_out=id_apelido,funcao_out=descricao)

        if tabela=='evento':
            obj=Evento.objects.filter(id_municipio=pid_municipio,evento=descricao).first()
            id_apelido=obj.id_evento
            Evento.objects.filter(id_municipio=pid_municipio,evento__in=lista2[k]).update(id_evento_out=id_apelido,evento_out=descricao)


    return 1


def ajustes2_tabela(pid_municipio,tabela):
    lista1=[]
    lista2=[]
    lista3=[]

    if tabela=='secretaria':
        query = Secretaria.objects.filter(id_municipio=pid_municipio,id_secretaria_out=0)
        for  qy in query:
            id_obj=qy.id_secretaria
            qy.id_secretaria_out=id_obj
            qy.save()

    if tabela=='setor':
        query = Setor.objects.filter(id_municipio=pid_municipio,id_setor_out=0)
        for  qy in query:
            id_obj=qy.id_setor
            qy.id_setor_out=id_obj
            qy.save()

    if tabela=='funcao':
        query = Funcao.objects.filter(id_municipio=pid_municipio,id_funcao_out=0)
        for  qy in query:
            id_obj=qy.id_funcao
            qy.id_funcao_out=id_obj
            qy.save()



    if tabela=='evento':
        query = Evento.objects.filter(id_municipio=pid_municipio,id_evento_out=0)
        for  qy in query:
            id_obj=qy.id_evento
            qy.id_evento_out=id_obj
            qy.save()


def ajustes3_tabela(id_cidade,tabela,lista_secs_12):
    dic_sec12=dicionarios.de_secretarias_12(id_cidade)
    for k in range(len(lista_secs_12)):
        (s1,s2) = lista_secs_12[k].split(":")
        id_1 = dic_sec12.get(s1,0)
        id_2 = dic_sec12.get(s2,0)
        if tabela=='secretaria' and id_1>0 and id_2>0:
            sec1 = Secretaria.objects.get(id_municipio=id_cidade,id_secretaria=id_1)
            sec1.id_secretaria_out = id_2
            sec1.secretaria_out = s1
            sec1.save()
    return 1



def import_excel_gravarFuncao_passo2(excel,idmunicipio):

    lst_secs=[sec.secretaria for sec in Secretaria.objects.filter(id_municipio=idmunicipio)]
    lst_sets=[sec.setor for sec in Setor.objects.filter(id_municipio=idmunicipio)]
    lst_funcs=[sec.funcao for sec in Funcao.objects.filter(id_municipio=idmunicipio)]
    lst_evts=[sec.evento for sec in Evento.objects.filter(id_municipio=idmunicipio)]
    lst_cargahs=[sec.cargah_origem for sec in CargaHoraria.objects.filter(id_municipio=idmunicipio)]

    dict_secretariasXid=dicionarios.dict_secretariasXid(idmunicipio)
    dict_setoresXid=dicionarios.dict_setoresXid(idmunicipio)
    dict_funcoesXid=dicionarios.dict_funcoesXid(idmunicipio)
    dict_eventosXid=dicionarios.dict_eventosXid(idmunicipio)
    dict_cargahXid=dicionarios.dict_chorigemXid(idmunicipio)

    wb = openpyxl.load_workbook(excel)
    sheets = wb.sheetnames

    #sheet = sheets[0]
    sheet = wb.active
    nrows = get_maximum_rows(sheet_object=sheet)
    max_col = sheet.max_column
    lista_new_information=[]

    lista_sec_descricao=[]
    lista_set_descricao=[]
    lista_fun_descricao=[]
    lista_evt_descricao=[]
    lista_ch_descricao=[]

    ls_ch_descricao1=[]

    ls_sec_descricao1=[]
    ls_sec_descricao2=[]

    ls_set_descricao1=[]
    ls_set_descricao2=[]

    ls_fun_descricao1=[]
    ls_fun_descricao2=[]

    ls_evt_descricao1=[]
    ls_evt_descricao2=[]

    lista_excluir_funcao=[]
    lista_excluir_evento=[]

    lista_excluir_id_funcao=[]
    lista_excluir_id_evento=[]

    lista_objeto=[]
    id_cidade=0

    row=1
    limite=500
    maximo=0
    while row<limite and row<nrows+1:
        if row==1:
            for i in range(1, max_col + 1):
                maximo+=1
                cell_obj = sheet.cell(row = 1, column = i)
                titulo_coluna=cell_obj.value
                if titulo_coluna is not None and 1==1:
                    titulo_coluna=remove_combining_fluent(titulo_coluna)
                    titulo_coluna=titulo_coluna.strip()
                    titulo_coluna=titulo_coluna.upper()
                    if titulo_coluna=='MUNICIPIO':
                        n_municipio=letra_da_coluna(i)
                    elif titulo_coluna=='TABELA':
                        n_tabela=letra_da_coluna(i)
                    elif titulo_coluna=='DESCRICAO_QCHEGA':
                        n_descricao_e=letra_da_coluna(i)
                    elif titulo_coluna=='DESCRICAO_QSAI':
                        n_descricao_s=letra_da_coluna(i)
                    elif titulo_coluna=='COMPLEMENTO':
                        n_complemento=letra_da_coluna(i)
                    elif titulo_coluna=='CARATER-SALARIO':
                        n_complemento2=letra_da_coluna(i)
                    elif titulo_coluna=='POS-GRADUACAO':
                        n_complemento3=letra_da_coluna(i)

                if maximo>10:
                    break

            row+=1
            continue

        if sheet['B'+str(row)].value is None:
            break
        if sheet['B'+str(row)].value=='':
            break

        municipio = sheet[n_municipio + str(row)].value
        tabela = sheet[n_tabela + str(row)].value

        municipio=remove_combining_fluent(municipio)
        tabela=remove_combining_fluent(tabela)
        municipio=municipio.strip()
        municipio=municipio.upper()
        tabela=tabela.strip()
        tabela=tabela.lower()


        if tabela not in ['funcao','evento','setor','secretaria','ch']:
            row+=1
            continue


        descricao_e = sheet[n_descricao_e + str(row)].value
        descricao_s = sheet[n_descricao_s + str(row)].value
        complemento = sheet[n_complemento + str(row)].value
        complemento2 = sheet[n_complemento2 + str(row)].value
        complemento3 = sheet[n_complemento3+ str(row)].value


        if descricao_e is None or  municipio is None:
            row+=1
            continue

        if descricao_s is None:
            descricao_s=''

        if complemento is None:
            complemento=''

        if complemento2 is None:
            complemento2=''

        if complemento3 is None:
            complemento3=''



        if tabela in ['funcao','evento','setor','secretaria']:

            descricao_e=remove_combining_fluent(descricao_e)
            descricao_s=remove_combining_fluent(descricao_s)
            descricao_e = descricao_e.strip()
            descricao_s = descricao_s.strip()
            descricao_e =  descricao_e.upper()
            descricao_s =  descricao_s.upper()
            complemento = complemento.strip()
            complemento = complemento.upper()
            complemento2 = complemento2.strip()
            complemento2 = complemento2.upper()
            complemento3 = complemento3.strip()
            complemento3 = complemento3.upper()



        if complemento=='EXCLUIR':
            if tabela=='funcao':
                lista_excluir_funcao.append(descricao_e)
            elif tabela=='evento':
                lista_excluir_evento.append(descricao_e)
            row+=1
            continue

        ampliacao='N'
        suporte='N'
        professor='N'
        fundeb='N'
        caraterSalario='N'
        posGraduacao='N'
        if complemento!='':
            if tabela=='evento':
                if complemento=='A':
                    ampliacao='S'
                elif complemento=='S':
                    suporte='S'
                else:
                    if complemento=='AS' or complemento=='SA':
                        ampliacao='S'
                        suporte='S'
            elif tabela=='funcao':
                if complemento=='P':
                    professor='S'
            elif tabela=='setor':
                if complemento=='F':
                    fundeb='S'

        if tabela=='evento':
            if complemento2=='SIM' or complemento2=='S':
                caraterSalario='S'
            if complemento3=='SIM' or complemento3=='S':
                posGraduacao='S'


        id_municipio=fun_id_municipio(municipio)

        if id_municipio!=idmunicipio:
            row+=1
            continue

        id_cidade=id_municipio


        if descricao_e==descricao_s:
            descricao_s=''

        if tabela=='secretaria':
            if descricao_e in lst_secs:
                if descricao_e not in lista_sec_descricao:
                    lista_sec_descricao.append(descricao_e)
                    if descricao_s=='':
                        ls_sec_descricao1.append([descricao_e])
                    else:
                        ls_sec_descricao2.append([descricao_e,descricao_s])

        if tabela=='setor':
            if descricao_e in lst_sets:
                if descricao_e not in lista_set_descricao:
                    lista_set_descricao.append(descricao_e)
                    if descricao_s=='':
                        ls_set_descricao1.append([descricao_e,fundeb])
                    else:
                        ls_set_descricao2.append([descricao_e,descricao_s,fundeb])

        if tabela=='funcao':
            if descricao_e in lst_funcs:
                if descricao_e not in lista_fun_descricao:
                    lista_fun_descricao.append(descricao_e)
                    if descricao_s=='':
                        ls_fun_descricao1.append([descricao_e,professor])
                    else:
                        ls_fun_descricao2.append([descricao_e,descricao_s,professor])

        if tabela=='evento':
            if descricao_e in lst_evts:
                if descricao_e not in lista_evt_descricao:
                    lista_evt_descricao.append(descricao_e)
                    if descricao_s=='':
                        ls_evt_descricao1.append([descricao_e,ampliacao,suporte,caraterSalario,posGraduacao])
                    else:
                        ls_evt_descricao2.append([descricao_e,descricao_s,ampliacao,suporte,caraterSalario,posGraduacao])

        if tabela=='ch':
            if descricao_e in lst_cargahs:
                if descricao_e not in lista_ch_descricao:
                    lista_ch_descricao.append(descricao_e)
                    if descricao_s=='':
                        descricao_s=descricao_e
                    ls_ch_descricao1.append([descricao_e,descricao_s])


        row+=1

    for kls in ls_sec_descricao1:
        secret=kls[0]
        id_sec=dict_secretariasXid.get(secret,0)
        if id_sec>0:
            obj=Secretaria.objects.get(id_secretaria=id_sec)
            if obj is not None:
                obj.id_secretaria_out=obj.id_secretaria
                obj.secretaria_out=obj.secretaria
                obj.save()


    for kls in ls_sec_descricao2:
        secretaria=kls[0]
        secretaria_out=kls[1]
        id_sec=dict_secretariasXid.get(secretaria,0)
        id_sec_out=dict_secretariasXid.get(secretaria_out,0)
        if id_sec>0 and id_sec_out>0:
            obj1=Secretaria.objects.get(id_secretaria=id_sec)
            obj2=Secretaria.objects.get(id_secretaria=id_sec_out)
            if obj1 is not None and obj2 is not None:
                obj1.id_secretaria_out=obj2.id_secretaria
                obj1.secretaria_out=obj2.secretaria
                obj1.save()


    for kls in ls_set_descricao1:
        setor=kls[0]
        fundeb=kls[1]
        id_setor=dict_setoresXid.get(setor,0)
        if id_setor>0:
            obj=Setor.objects.get(id_setor=id_setor)
            if obj is not None:
                obj.id_setor_out=obj.id_setor
                obj.setor_out=obj.setor
                obj.fundeb=fundeb
                obj.save()


    for kls in ls_set_descricao2:
        setor=kls[0]
        setor_out=kls[1]
        fundeb=kls[2]
        id_setor=dict_setoresXid.get(setor,0)
        id_setor_out=dict_setoresXid.get(setor_out,0)
        if id_setor>0 and id_setor_out>0:
            obj1=Setor.objects.get(id_setor=id_setor)
            obj2=Setor.objects.get(id_setor=id_setor_out)
            if obj1 is not None and obj2 is not None:
                obj1.id_setor_out=obj2.id_setor
                obj1.setor_out=obj2.setor
                obj1.fundeb=fundeb
                obj2.fundeb=fundeb
                obj1.save()
                obj2.save()


    for kls in ls_fun_descricao1:
        func=kls[0]
        professor=kls[1]
        id_func=dict_funcoesXid.get(func,0)
        if id_func>0:
            obj=Funcao.objects.get(id_funcao=id_func)
            if obj is not None:
                obj.id_funcao_out=obj.id_funcao
                obj.funcao_out=obj.funcao
                obj.professor=professor
                obj.save()

    for kls in ls_fun_descricao2:
        func=kls[0]
        func_out=kls[1]
        professor=kls[2]
        id_func=dict_funcoesXid.get(func,0)
        id_func_out=dict_funcoesXid.get(func_out,0)
        if id_func>0 and id_func_out>0:
            obj1=Funcao.objects.get(id_funcao=id_func)
            obj2=Funcao.objects.get(id_funcao=id_func_out)
            if obj1 is not None and obj2 is not None:
                obj1.id_funcao_out=obj2.id_funcao
                obj1.funcao_out=obj2.funcao
                obj1.professor=professor
                obj1.save()
                obj2.professor=professor
                obj2.save()



    for kls in ls_evt_descricao1:
        evento=kls[0]
        ampliacao=kls[1]
        suporte=kls[2]
        caraterSalario=kls[3]
        posGraduacao=kls[4]

        id_evento=dict_eventosXid.get(evento,0)
        if id_evento>0:
            obj=Evento.objects.get(id_evento=id_evento)
            if obj is not None:
                obj.id_evento_out=obj.id_evento
                obj.evento_out=obj.evento
                obj.ampliacao_ch=ampliacao
                obj.suporte=suporte
                obj.carater_salario=caraterSalario
                obj.posgraduacao=posGraduacao
                obj.save()

    for kls in ls_evt_descricao2:
        evento=kls[0]
        evento_out=kls[1]
        ampliacao=kls[2]
        suporte=kls[3]
        caraterSalario=kls[4]
        posGraduacao=kls[5]
        id_evento=dict_eventosXid.get(evento,0)
        id_evento_out=dict_eventosXid.get(evento_out,0)
        if id_evento>0 and id_evento_out>0:
            obj1=Evento.objects.get(id_evento=id_evento)
            obj2=Evento.objects.get(id_evento=id_evento_out)
            if obj1 is not None and obj2 is not None:
                obj1.id_evento_out=obj2.id_evento
                obj1.evento_out=obj2.evento
                obj1.ampliacao_ch=ampliacao
                obj1.suporte=suporte
                obj1.carater_salario=caraterSalario
                obj1.posgraduacao=posGraduacao
                obj1.save()
                obj2.ampliacao_ch=ampliacao
                obj2.suporte=suporte
                obj2.carater_salario=caraterSalario
                obj2.posgraduacao=posGraduacao
                obj2.save()


    for kls in ls_ch_descricao1:
        ch_origem=kls[0]
        ch_convertida=kls[1]
        id_ch=dict_cargahXid.get(ch_origem,0)
        if id_ch>0:
            obj=CargaHoraria.objects.get(id_cargahoraria=id_ch)
            if obj is not None:
                obj.cargah_convertida=ch_convertida
                obj.save()


    for funcao in lista_excluir_funcao:
        id_funcao=dict_funcoesXid.get(funcao,0)
        if id_funcao>0:
            lista_excluir_id_funcao.append(id_funcao)

    if len(lista_excluir_id_funcao)>0:
        Funcao.objects.filter(id_municipio=idmunicipio,id_funcao__in=lista_excluir_id_funcao).update(cancelado='S')



    for evento in lista_excluir_evento:
        id_evento=dict_eventosXid.get(evento,0)
        if id_evento>0:
            lista_excluir_id_evento.append(id_evento)

    if len(lista_excluir_id_evento)>0:
        Evento.objects.filter(id_municipio=idmunicipio,id_evento__in=lista_excluir_id_evento).update(cancelado='S')


    return 0


