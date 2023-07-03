# -*- coding: utf-8 -*-
from django.http import HttpResponseRedirect
import openpyxl
import os
import sys
import datetime
import unicodedata
from .models import Funcionario,Provento,Complemento,Ampliacao_ch,Suporte,SemCadastro,Information,Tabela,Secretaria,Setor,Funcao,Evento,CaraterSalario,Posgraduation
from . import listagens,dicionarios

def importacao_excel_folha(excel,ano,num_mes,anomes,id_municipio,current_user):
    roda=0
    (dict_sec1,dict_sec2)=dicionarios.de_secretarias(id_municipio)
    (dict_set1,dict_set2,lista_setores_fundeb)=dicionarios.de_setores(id_municipio)
    (dict_evento1,dict_evento2,lista_eventos_suporte,lista_eventos_ampliacaoch,lista_pos,lista_csalario)=dicionarios.de_eventos(id_municipio)
    (dict_func1,dict_func2,lista_funcoes_prof)=dicionarios.de_funcoes(id_municipio)
    (dict_salario_int,dict_salario_dec)=dicionarios.de_tabela_salarios(id_municipio)
    dict_carga_horaria=dicionarios.de_tab_salariosch(id_municipio)
    dict_vinculos = dicionarios.de_vinculos(id_municipio)
    dict_id_vinculoXgrupo = dicionarios.dict_id_vinculoXgrupo(id_municipio)
    lista_de_vantagens=listagens.lista_de_vantagens(id_municipio)
    dict_cargah_origemXcargah_convertida=dicionarios.cargah_origemXcargah_convertida(id_municipio)
    lista_eventos=[k.evento for k in Evento.objects.filter(id_municipio=id_municipio,cancelado='N')]


    lista_eventos_complementos=[] #eventos_tabela_complementos(id_municipio)
    lista_sal100=[]
    lista_sal200=[]

    wb = openpyxl.load_workbook(excel)
    sheets = wb.sheetnames

    #sheet = sheets[0]
    sheet = wb.active
    nrows = get_maximum_rows(sheet_object=sheet)
    max_col = sheet.max_column

    erro=0
    erros=0
    qtde_itens=0
    retorno = True

    erro=0
    lista=[]
    codigos_ja_incluidos=[]
    lista_ch=[]
    funcionarios_ja_incluidos=[]
    lista_funcionario_suporte=[]
    lista_retorno=[]

    lista_new_suporte=[]
    lista_new_complemento=[]
    lista_new_ampliacao_ch=[]
    lista_new_funcionario=[]
    lista_new_provento=[]

    lista_eventos_nao_cadastrados=[]
    lista_funcoes_nao_cadastrados=[]
    lista_setores_nao_cadastrados=[]
    lista_secretarias_nao_cadastrados=[]
    lista_sem_cadastro=[]

    lista_funcao_sal100=[]
    lista_valor_int_sal100=[]
    lista_valor_dec_sal100=[]

    lista_pos_funcionario=[]
    lista_csalario_funcionario=[]
    lista_new_pos=[]
    lista_new_csalario=[]
    processo=True

    lista_mes=['1','2','3','4','5','6','7','8','9','10','11','12']
    lixo=0
    lista_salario_not_found=[]

    flag=0

    row=1
    limite=5000
    maximo=0

    lista_nome_evento=[]
    lista_letra_coluna=[]
    planilha_completa=False
    lista_coluna=[]
    lista_relacao_eventos=[]
    lista_nao_cadastrados=[]
    while row<limite and row<nrows+1:
        lista_valores=[]
        lista_relacao_eventos=[]
        if row==1:
            for i in range(1, max_col + 1):
                maximo+=1
                cell_obj = sheet.cell(row = 1, column = i)
                titulo_coluna=cell_obj.value
                if titulo_coluna is not None:
                    titulo_coluna=fun_tratamentoString(titulo_coluna)

                    if titulo_coluna=='MATRICULA':
                        n_codigo_servidor=letra_da_coluna(i)
                    elif titulo_coluna=='NOME':
                        n_nome_funcionario=letra_da_coluna(i)
                    elif titulo_coluna=='ORGAO':
                        n_secretaria=letra_da_coluna(i)
                    elif titulo_coluna=='SETOR':
                        n_setor=letra_da_coluna(i)
                    elif titulo_coluna=='VINCULO':
                        n_natureza=letra_da_coluna(i)
                    elif titulo_coluna=='DTA. ADMISSAO':
                        n_data_admissao=letra_da_coluna(i)
                    elif titulo_coluna=='CARGO':
                        n_funcao=letra_da_coluna(i)
                    elif titulo_coluna=='CH':
                        n_carga_horaria=letra_da_coluna(i)
                    elif titulo_coluna=='CODIGO FOLHA':
                        n_mes_da_folha=letra_da_coluna(i)
                    elif titulo_coluna=='REFERENCIA':
                        n_ref_evento=letra_da_coluna(i)
                    else:
                        coluna = letra_da_coluna(i)
                        nome_evento = titulo_coluna
                        if nome_evento in lista_de_vantagens:
                            lista_nome_evento.append(nome_evento)
                            lista_letra_coluna.append(coluna)
                    if maximo>85:
                        break

                    if titulo_coluna in ['MATRICULA','NOME','ORGAO','SETOR','VINCULO','DTA. ADMISSAO','CARGO','CH','CODIGO FOLHA','REFERENCIA']:
                        lista_coluna.append(titulo_coluna)
            dict_letraXevento=dict(zip(lista_letra_coluna,lista_nome_evento))
            if row==1:
                row+=1
                continue

        if sheet['A'+str(row)].value is None:
            break
        if str(sheet['A'+str(row)].value)=='':
            break

        codigo_funcionario = sheet[n_codigo_servidor + str(row)].value
        mes_da_folha = num_mes # sheet[n_mes_da_folha + str(row)].value
        nome_funcionario = (sheet[n_nome_funcionario + str(row)].value)
        carga_horaria = sheet[n_carga_horaria + str(row)].value
        data_admissao    = sheet[n_data_admissao + str(row)].value

        secretaria  = sheet[n_secretaria + str(row)].value
        funcao = sheet[n_funcao + str(row)].value
        setor = sheet[n_setor + str(row)].value
        previdencia='INSS'
        natureza = sheet[n_natureza + str(row)].value
        ref_evento = '30d' #sheet[n_ref_evento + str(row)].value

        if mes_da_folha is None:
            row+=1
            continue

        if int(mes_da_folha)!=int(num_mes):
            row+=1
            continue

        tem_ampliacao_ch='N'
        tem_suporte='N'
        tem_posgraduacao='N'
        tem_caraterSalario='N'
        salario_base=0
        valor_evento=0

        secretaria = fun_tratamentoString(secretaria)
        setor = fun_tratamentoString(setor)
        funcao = fun_tratamentoString(funcao)
        nome_funcionario = fun_tratamentoString(nome_funcionario)
        natureza = fun_tratamentoString(natureza)

        id_natureza=dict_vinculos.get(natureza,0)
        id_secretaria_origem=dict_sec1.get(secretaria,0)
        id_setor_origem=dict_set1.get(setor,0)
        id_funcao_origem=dict_func1.get(funcao,0)
        vinculo_grupo=dict_id_vinculoXgrupo.get(id_natureza,'')

        if id_secretaria_origem==0:
            if ['secretaria',secretaria] not in lista_sem_cadastro:
                lista_sem_cadastro.append(['secretaria',secretaria])
                lista_nao_cadastrados.append({'tabela':'secretaria','descricao':secretaria})

        if id_setor_origem==0:
            if ['setor',setor] not in lista_sem_cadastro:
                lista_sem_cadastro.append(['setor',setor])
                lista_nao_cadastrados.append({'tabela':'setor','descricao':setor})

        if id_funcao_origem==0:
            if ['funcao',funcao] not in lista_sem_cadastro:
                lista_sem_cadastro.append(['funcao',funcao])
                lista_nao_cadastrados.append({'tabela':'funcao','descricao':funcao})

        if len(lista_nao_cadastrados)>0:
            row+=1
            continue

        for key,nome_do_evento in dict_letraXevento.items():
            vl_evento = sheet[key + str(row)].value
            if vl_evento is not None:
                lista_valores.append([nome_do_evento,vl_evento])
                lista_relacao_eventos.append(nome_do_evento)

                idevento=dict_evento1.get(nome_do_evento,0)
                if idevento in lista_eventos_suporte:
                    tem_suporte='S'

                if idevento in lista_eventos_ampliacaoch:
                    tem_ampliacao_ch='S'

                if (nome_do_evento=='SALARIO BASE' or nome_do_evento=='VENCIMENTO'):
                    vencimento_base = vl_evento

        id_secretaria=dict_sec2.get(id_secretaria_origem,0)
        id_setor=dict_set2.get(id_setor_origem,0)
        id_funcao=dict_func2.get(id_funcao_origem,0)

        grupamento='S'
        classificacao='O'

        fundeb='N'
        if id_setor in lista_setores_fundeb:
            fundeb='S'
        elif id_setor_origem in lista_setores_fundeb:
            fundeb='S'

        data_admissao=str(data_admissao)[0:10]

        classificacao='O'
        participacao=0
        carga_horaria_origem=carga_horaria

        carga_horaria = dict_cargah_origemXcargah_convertida.get(carga_horaria_origem,carga_horaria_origem)

        num_dias=fun_num_dias(ref_evento)
        dados_salario = fun_calcula_salario(carga_horaria,vencimento_base,num_dias)

        if codigo_funcionario not in funcionarios_ja_incluidos:
            funcionarios_ja_incluidos.append(codigo_funcionario)

            new_complemento=Complemento(
                    id_municipio=id_municipio,
                    anomes=anomes,
                    codigo=codigo_funcionario,
                    salario=dados_salario['salario'],
                    salario_100H=dados_salario['salario_100H'],
                    num_dias=num_dias,
                    participacao=dados_salario['participacao'],
                    vencimento_base=vencimento_base,
                    fundeb=fundeb
            )
            lista_new_complemento.append(new_complemento)

            new_funcionario = Funcionario(
                id_municipio=id_municipio,
                anomes=anomes,
                codigo=codigo_funcionario,
                nome_servidor=nome_funcionario,
                carga_horaria=dados_salario['carga_horaria'],
                tipo_admissao=natureza,
                data_admissao=data_admissao,
                previdencia=previdencia,
                id_secretaria=id_secretaria,
                id_setor=id_setor,
                id_funcao=id_funcao,
                id_secretaria_origem = id_secretaria_origem,
                id_setor_origem = id_setor_origem,
                id_funcao_origem = id_funcao_origem,
                fundeb=fundeb,
                carga_horaria_origem=carga_horaria_origem
            )
            lista_new_funcionario.append(new_funcionario)

            for k in range(len(lista_valores)):
                evento=lista_valores[k][0]
                id_evento_origem=dict_evento1.get(evento,0)
                id_evento=dict_evento2.get(id_evento_origem,0)
                valor_evento=lista_valores[k][1]
                new_provento = Provento(
                    id_municipio=id_municipio,
                    anomes=anomes,
                    codigo=codigo_funcionario,
                    previdencia='I',
                    tipo=1,
                    id_evento=id_evento,
                    id_evento_origem=id_evento_origem,
                    valor_evento=valor_evento,
                    classificacao='O',
                    grupamento='S',
                    lixo=k
                )
                lista_new_provento.append(new_provento)

            if tem_ampliacao_ch=='S': # and id_funcao_origem in lista_funcoes_prof:
                new_ampliacao_ch = Ampliacao_ch(
                    id_municipio=id_municipio,
                    anomes=anomes,
                    cod_servidor=codigo_funcionario
                )
                lista_new_ampliacao_ch.append(new_ampliacao_ch)

            if (tem_suporte=='S' or vinculo_grupo=='C'):
                new_suporte = Suporte(
                    id_municipio=id_municipio,
                    anomes=anomes,
                    cod_servidor=codigo_funcionario
                )
                lista_new_suporte.append(new_suporte)

        if len(lista_new_complemento)>150:
            Complemento.objects.bulk_create(lista_new_complemento)
            lista_new_complemento=[]

        if len(lista_new_ampliacao_ch)>150:
            Ampliacao_ch.objects.bulk_create(lista_new_ampliacao_ch)
            lista_new_ampliacao_ch=[]

        if len(lista_new_suporte)>150:
            Suporte.objects.bulk_create(lista_new_suporte)
            lista_new_suporte=[]

        if len(lista_new_funcionario)>150:
            Funcionario.objects.bulk_create(lista_new_funcionario)
            lista_new_funcionario=[]

        if len(lista_new_provento)>150:
            Provento.objects.bulk_create(lista_new_provento)
            lista_new_provento=[]

        row+=1

    if len(lista_nao_cadastrados)>0:
        return lista_nao_cadastrados

    if len(lista_new_complemento)>0:
        Complemento.objects.bulk_create(lista_new_complemento)
        lista_new_complemento=[]

    if len(lista_new_ampliacao_ch)>0:
        Ampliacao_ch.objects.bulk_create(lista_new_ampliacao_ch)
        lista_new_ampliacao_ch=[]

    if len(lista_new_suporte)>0:
        Suporte.objects.bulk_create(lista_new_suporte)
        lista_new_suporte=[]

    if len(lista_new_funcionario)>0:
        Funcionario.objects.bulk_create(lista_new_funcionario)
        lista_new_funcionario=[]

    if len(lista_new_provento)>0:
        Provento.objects.bulk_create(lista_new_provento)
        lista_new_provento=[]

    return []

def letra_da_coluna(pi):
    lista1=['A','B','C']
    lista2=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    lista3=[]
    for k1 in range(len(lista1)):
        alfa=lista1[k1]
        for k2 in range(len(lista2)):
            lista3.append(alfa+lista2[k2])
    listaColunas=lista2+lista3
    return listaColunas[pi-1]

def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows

def remover_acentuacao(string: str) -> str:
    normalized = unicodedata.normalize('NFD', string)
    return ''.join(
        [l for l in normalized if not unicodedata.combining(l)]
    )

def fun_calcula_carga_horaria(pcarga_horaria,id_municipio):
    if id_municipio==42:
       return pcarga_horaria*5

    if pcarga_horaria==4:
        return 100
    elif pcarga_horaria==8:
        return 200
    elif pcarga_horaria==20:
        return 100
    elif pcarga_horaria==40:
        return 200
    elif pcarga_horaria==80:
        return 200
    else:
        return pcarga_horaria

def eventos_tabela_complementos(id_municipio):
    #Coluna da tabela Excel que vem do municipio "Cod. Evento"
    if id_municipio==38:
        return [1,9,227]
    else:
        return [1]

def fun_calcula_salario(carga_horaria,vencimento_base,num_dias):
    salario = (vencimento_base/num_dias)*30
    if carga_horaria<=150:
        salario_100H=round((salario/carga_horaria)*100,3)
        carga_horaria=100
        sal_participacao=salario_100H
    else:        
        salario_200H=round((salario/carga_horaria)*200,3)
        salario_100H=(round((salario/carga_horaria)*200,3))/2
        sal_participacao=salario_200H
        carga_horaria=200

    if sal_participacao>0:        
        participacao=vencimento_base/sal_participacao
    else:
        participacao=0
    dados = {
        'carga_horaria':carga_horaria,
        'salario':salario,
        'salario_100H':salario_100H,
        'participacao':participacao
    }
    return dados

def fun_tratamentoString(string):
    retorno=remover_acentuacao(string)
    retorno=retorno.strip()
    retorno=retorno.upper()
    if retorno[0] in ['1','2','3','4','5','6','7','8','9','0']:
        if '-' in retorno[0:7]: 
            retorno=fun_titulo_coluna(retorno)
    return retorno

def fun_titulo_coluna(titulo_coluna):
    if '-' in titulo_coluna[0:7]:
        array = titulo_coluna.split("-",1)
        return array[1].strip()
    else:
        return titulo_coluna

def fun_nao_cadastrados(lista,id_secretaria_origem,id_setor_origem,id_funcao_origem,secretaria,setor,funcao):
    if id_secretaria_origem==0:
        if secretaria not in lista:
            lista.append({'tabela':'secretaria','descricao':secretaria})

    if id_setor_origem==0:
        if setor not in lista:
            lista.append({'tabela':'setor','descricao':setor})

    if id_funcao_origem==0:
        if funcao not in lista:
            lista.append({'tabela':'funcao','descricao':funcao})
    return lista

def fun_num_dias(ref_evento):
    ref_evento=ref_evento.replace('d','')
    ref_evento=ref_evento.strip()
    for k in range(1,31):
        if ref_evento==str(k):
            return int(ref_evento)
    return 30

def fun_eventos_nao_cadastrados(lista,lista_eventos,evento):
    if evento not in lista_eventos:
        lista.append({'tabela':'evento','descricao':evento})
    return lista











