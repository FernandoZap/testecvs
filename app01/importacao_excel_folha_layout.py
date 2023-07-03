# -*- coding: utf-8 -*-
from django.http import HttpResponseRedirect
import openpyxl
import os
import sys
import datetime
import unicodedata
from .models import Funcionario,Provento,Complemento,Ampliacao_ch,Suporte,SemCadastro,Information,Tabela,Secretaria,Setor,Funcao,Evento,CaraterSalario,Posgraduation
from . import listagens,dicionarios

def importacao_excel_folha(excel,ano,num_mes,anomes,id_municipio,current_user):
    roda=0
    (dict_sec1,dict_sec2)=dicionarios.de_secretarias(id_municipio)
    (dict_set1,dict_set2,lista_setores_fundeb)=dicionarios.de_setores(id_municipio)
    (dict_evento1,dict_evento2,lista_eventos_suporte,lista_eventos_ampliacaoch,lista_pos,lista_csalario)=dicionarios.de_eventos(id_municipio)
    (dict_func1,dict_func2,lista_funcoes_prof)=dicionarios.de_funcoes(id_municipio)
    (dict_salario_int,dict_salario_dec)=dicionarios.de_tabela_salarios(id_municipio)
    dict_carga_horaria=dicionarios.de_tab_salariosch(id_municipio)
    dict_vinculos = dicionarios.de_vinculos(id_municipio)
    dict_id_vinculoXgrupo = dicionarios.dict_id_vinculoXgrupo(id_municipio)
    lista_de_vantagens=listagens.lista_de_vantagens(id_municipio)
    dict_cargah_origemXcargah_convertida=dicionarios.cargah_origemXcargah_convertida(id_municipio)


    lista_eventos_complementos=[] #eventos_tabela_complementos(id_municipio)
    lista_sal100=[]
    lista_sal200=[]

    wb = openpyxl.load_workbook(excel)
    sheets = wb.sheetnames

    #sheet = sheets[0]
    sheet = wb.active
    nrows = get_maximum_rows(sheet_object=sheet)
    max_col = sheet.max_column


    erro=0
    erros=0
    qtde_itens=0
    retorno = True

    erro=0
    lista=[]
    codigos_ja_incluidos=[]
    lista_ch=[]
    ref_eventos_ja_incluidos=[]
    lista_funcionario_suporte=[]
    lista_retorno=[]

    lista_new_suporte=[]
    lista_new_complemento=[]
    lista_new_ampliacao_ch=[]
    lista_new_funcionario=[]
    lista_new_provento=[]

    lista_eventos_nao_cadastrados=[]
    lista_funcoes_nao_cadastrados=[]
    lista_setores_nao_cadastrados=[]
    lista_secretarias_nao_cadastrados=[]
    lista_sem_cadastro=[]

    lista_funcao_sal100=[]
    lista_valor_int_sal100=[]
    lista_valor_dec_sal100=[]


    lista_pos_funcionario=[]
    lista_csalario_funcionario=[]
    lista_new_pos=[]
    lista_new_csalario=[]
    processo=True


    lista_mes=['1','2','3','4','5','6','7','8','9','10','11','12']
    lixo=0
    lista_salario_not_found=[]

    flag=0

    row=1
    limite=15000
    maximo=0

    lista_nome_evento=[]
    lista_letra_coluna=[]

    planilha_completa=False
    lista_coluna=[]
    lista_relacao_eventos=[]
    while row<limite and row<nrows+1:
        lista_valores=[]
        lista_relacao_eventos=[]
        if row==1:
            for i in range(1, max_col + 1):
                maximo+=1
                cell_obj = sheet.cell(row = 1, column = i)
                titulo_coluna=cell_obj.value
                if titulo_coluna is not None:
                    titulo_coluna=remover_acentuacao(titulo_coluna)
                    titulo_coluna=titulo_coluna.strip()
                    titulo_coluna=titulo_coluna.upper()

                    if titulo_coluna=='MATRICULA':
                        n_codigo_servidor=letra_da_coluna(i)
                    elif titulo_coluna=='NOME':
                        n_nome_funcionario=letra_da_coluna(i)
                    elif titulo_coluna=='SECRETARIA':
                        n_secretaria=letra_da_coluna(i)
                    elif titulo_coluna=='SETOR':
                        n_setor=letra_da_coluna(i)
                    elif titulo_coluna=='SISTEMA':
                        n_natureza=letra_da_coluna(i)
                    elif titulo_coluna=='DTADMISSAO':
                        n_data_admissao=letra_da_coluna(i)
                    elif titulo_coluna=='CARGO':
                        n_funcao=letra_da_coluna(i)
                    elif titulo_coluna=='HORTRAB':
                        n_carga_horaria=letra_da_coluna(i)
                    elif titulo_coluna=='CODIGO FOLHA':
                        n_mes_da_folha=letra_da_coluna(i)
                    elif titulo_coluna=='DIAS':
                        n_ref_evento=letra_da_coluna(i)
                    else:
                        coluna = letra_da_coluna(i)
                        nome_evento = fun_titulo_coluna(titulo_coluna)
                        if nome_evento in lista_de_vantagens:
                            lista_nome_evento.append(nome_evento)
                            lista_letra_coluna.append(coluna)
                    if maximo>85:
                        break

            dict_letraXevento=dict(zip(lista_letra_coluna,lista_nome_evento))
            if row==1:
                row+=1
                continue

        if sheet['A'+str(row)].value is None:
            break
        if str(sheet['A'+str(row)].value)=='':
            break


        codigo_funcionario = sheet[n_codigo_servidor + str(row)].value
        codigo_funcionario=int(codigo_funcionario.replace('-',''))


        mes_da_folha = sheet[n_mes_da_folha + str(row)].value
        nome_funcionario = (sheet[n_nome_funcionario + str(row)].value)
        carga_horaria = sheet[n_carga_horaria + str(row)].value

        if carga_horaria is None:
            carga_horaria='100'
        carga_horaria=carga_horaria.strip()
        if carga_horaria=='':
            carga_horaria =  '100'
        carga_horaria = int(carga_horaria)


        data_admissao    = sheet[n_data_admissao + str(row)].value
        data_admissao = '2020-5-5'

        secretaria  = sheet[n_secretaria + str(row)].value
        funcao = sheet[n_funcao + str(row)].value
        setor = sheet[n_setor + str(row)].value
        #previdencia = sheet[n_previdencia + str(row)].value
        previdencia='INSS'
        natureza = sheet[n_natureza + str(row)].value

        ref_evento = sheet[n_ref_evento + str(row)].value

        if mes_da_folha is None:
            row+=1
            continue

        if int(mes_da_folha)!=int(num_mes):
            row+=1
            continue

        tem_ampliacao_ch='N'
        tem_suporte='N'
        tem_posgraduacao='N'
        tem_caraterSalario='N'
        salario_base=0
        valor_evento=0

        for key,nome_do_evento in dict_letraXevento.items():
            vl_evento = sheet[key + str(row)].value
            if vl_evento=='0E-18':
                break
            if vl_evento is not None:
                lista_valores.append([nome_do_evento,vl_evento])
                lista_relacao_eventos.append(nome_do_evento)

                idevento=dict_evento1.get(nome_do_evento,0)
                if idevento in lista_eventos_suporte:
                    tem_suporte='S'
                if idevento in lista_eventos_ampliacaoch:
                    tem_ampliacao_ch='S'
                if (nome_do_evento=='SALARIO'):
                    salario_base = vl_evento
                    valor_evento =  vl_evento
        #print ('196. cod_funcionario: '+str(codigo_funcionario))
        if vl_evento=='0E-18':
            row+=1
            continue
        if len(lista_valores)<2:
            row+=1
            continue
        if salario_base==0:
            row+=1
            continue

        secretaria=remover_acentuacao(secretaria)
        setor=remover_acentuacao(setor)
        funcao=remover_acentuacao(funcao)
        nome_funcionario=remover_acentuacao(nome_funcionario)
        natureza=remover_acentuacao(natureza)

        secretaria=secretaria.strip()
        setor=setor.strip()
        funcao=funcao.strip()
        nome_funcionario=nome_funcionario.strip()
        natureza=natureza.strip()

        secretaria=secretaria.upper()
        setor=setor.upper()
        funcao=funcao.upper()
        natureza=natureza.upper()
        nome_funcionario=nome_funcionario.upper()

        funcao=fun_titulo_coluna(funcao)
        setor=fun_titulo_coluna(setor)
        secretaria=fun_titulo_coluna(secretaria)
        natureza=fun_titulo_coluna(natureza)



        id_natureza=dict_vinculos.get(natureza,0)
        id_secretaria_origem=dict_sec1.get(secretaria,0)
        id_setor_origem=dict_set1.get(setor,0)
        id_funcao_origem=dict_func1.get(funcao,0)

        vinculo_grupo=dict_id_vinculoXgrupo.get(id_natureza,'')

        if id_secretaria_origem==0:
            if secretaria not in lista_secretarias_nao_cadastrados:
                lista_secretarias_nao_cadastrados.append(secretaria)
                processo=False

        if id_setor_origem==0:
            if setor not in lista_setores_nao_cadastrados:
                lista_setores_nao_cadastrados.append(setor)
                processo=False

        if id_funcao_origem==0:
            if funcao not in lista_funcoes_nao_cadastrados:
                lista_funcoes_nao_cadastrados.append(funcao)
                processo=False


        if processo==False:
            row+=1
            continue


        id_secretaria=dict_sec2.get(id_secretaria_origem,0)
        id_setor=dict_set2.get(id_setor_origem,0)
        id_funcao=dict_func2.get(id_funcao_origem,0)

        grupamento='S'
        classificacao='O'

        fundeb='N'
        if id_setor in lista_setores_fundeb:
            fundeb='S'
        elif id_setor_origem in lista_setores_fundeb:
            fundeb='S'

        data_admissao=str(data_admissao)[0:10]

        classificacao='O'
        participacao=0
        carga_horaria_origem=carga_horaria

        #carga_horaria = dict_cargah_origemXcargah_convertida.get(carga_horaria_origem,carga_horaria_origem)
        carga_horaria = fun_calcula_carga_horaria(carga_horaria_origem)


        if fundeb =='S':
            str_ch_errada=str(id_funcao_origem)+'-'+str(int(salario_base))+'-'+str(carga_horaria)
            ch_certa=dict_carga_horaria.get(str_ch_errada,'000')
            if ch_certa!='000':
                carga_horaria=int(ch_certa)

        if ('SALARIO BASE' in lista_relacao_eventos or 'VENCIMENTO' in lista_relacao_eventos) and fundeb=='S':
            ref_eventos_ja_incluidos.append(codigo_funcionario)
            num_dias=ref_evento
            if num_dias==0:
                num_dias=30
            elif num_dias==99:
                num_dias=30

            if carga_horaria>=50 and carga_horaria<=150:
                if carga_horaria==100 and num_dias==30:
                    salario=salario_base
                    salario_100H=salario_base
                else:
                    salario=round((salario_base/carga_horaria)*100,3)
                    salario_100H=salario
                if salario==0:
                    participacao=1
                else:
                    participacao=round(valor_evento/salario,6)
                carga_horaria=100
            elif carga_horaria>150:
                if carga_horaria==200:
                    salario = salario_base
                    salario_200H = salario_base
                else:
                    salario=round((salario_base/carga_horaria)*200,3)
                    salario_200H=salario

                salario_100H=round(salario_200H/2,3)
                if salario_200H==0:
                    participacao=1
                else:
                    participacao=round(valor_evento/salario_200H,6)
                carga_horaria=200

            lista_sal100=dict_salario_int.get(id_funcao_origem,[])
            lista_sal200=dict_salario_dec.get(id_funcao_origem,[])

            if len(lista_sal100)>0:
                salario_100H = fun_salario100h(salario_100H,lista_sal100,lista_sal200)

            new_complemento=Complemento(
                    id_municipio=id_municipio,
                    anomes=anomes,
                    codigo=codigo_funcionario,
                    salario=salario,
                    salario_100H=salario_100H,
                    num_dias=num_dias,
                    participacao=participacao,
                    vencimento_base=valor_evento,
                    fundeb=fundeb
            )
            lista_new_complemento.append(new_complemento)


        if fundeb=='S' and tem_ampliacao_ch=='S': # and id_funcao_origem in lista_funcoes_prof:
            new_ampliacao_ch = Ampliacao_ch(
                id_municipio=id_municipio,
                anomes=anomes,
                cod_servidor=codigo_funcionario
            )
            lista_new_ampliacao_ch.append(new_ampliacao_ch)


        if fundeb=='S' and (tem_suporte=='S' or vinculo_grupo=='C'):
            new_suporte = Suporte(
                id_municipio=id_municipio,
                anomes=anomes,
                cod_servidor=codigo_funcionario
            )
            lista_new_suporte.append(new_suporte)

        #if fundeb=='S' and id_evento_origem in lista_pos and codigo_funcionario not in lista_pos_funcionario:
        if 1==2:
            new_pos = Posgraduation(
                id_municipio=id_municipio,
                anomes=anomes,
                cod_servidor=codigo_funcionario
            )
            lista_new_pos.append(new_pos)
            lista_pos_funcionario.append(codigo_funcionario)


        #if fundeb=='S' and id_evento_origem in lista_csalario and codigo_funcionario not in lista_csalario_funcionario:
        if 1==2:
            new_csalario = CaraterSalario(
                id_municipio=id_municipio,
                anomes=anomes,
                cod_servidor=codigo_funcionario
            )
            lista_new_csalario.append(new_csalario)
            lista_csalario_funcionario.append(codigo_funcionario)

        #print ('codigo do funcionario:' +str(codigo_funcionario))
        if codigo_funcionario not in codigos_ja_incluidos:
            new_funcionario = Funcionario(
                id_municipio=id_municipio,
                anomes=anomes,
                codigo=codigo_funcionario,
                nome_servidor=nome_funcionario,
                carga_horaria=carga_horaria,
                tipo_admissao=natureza,
                data_admissao=data_admissao,
                previdencia=previdencia,
                id_secretaria=id_secretaria,
                id_setor=id_setor,
                id_funcao=id_funcao,
                id_secretaria_origem = id_secretaria_origem,
                id_setor_origem = id_setor_origem,
                id_funcao_origem = id_funcao_origem,
                fundeb=fundeb,
                carga_horaria_origem=carga_horaria_origem
            )
            lista_new_funcionario.append(new_funcionario)
            codigos_ja_incluidos.append(codigo_funcionario)


        for kl in lista_valores:
            evento=kl[0]
            id_evento_origem=dict_evento1.get(evento,0)
            id_evento=dict_evento2.get(id_evento_origem,0)
            valor_evento=kl[1]
            new_provento = Provento(
                id_municipio=id_municipio,
                anomes=anomes,
                codigo=codigo_funcionario,
                previdencia='I',
                tipo=1,
                id_evento=id_evento,
                id_evento_origem=id_evento_origem,
                valor_evento=valor_evento,
                classificacao='O',
                grupamento='S',
                lixo=k
            )
            lista_new_provento.append(new_provento)


        if len(lista_new_complemento)>150:
            Complemento.objects.bulk_create(lista_new_complemento)
            lista_new_complemento=[]

        if len(lista_new_ampliacao_ch)>150:
            Ampliacao_ch.objects.bulk_create(lista_new_ampliacao_ch)
            lista_new_ampliacao_ch=[]

        if len(lista_new_suporte)>150:
            Suporte.objects.bulk_create(lista_new_suporte)
            lista_new_suporte=[]


        if len(lista_new_pos)>150:
            Posgraduation.objects.bulk_create(lista_new_pos)
            lista_new_pos=[]

        if len(lista_new_csalario)>150:
            CaraterSalario.objects.bulk_create(lista_new_csalario)
            lista_new_csalario=[]

        if len(lista_new_funcionario)>150:
            Funcionario.objects.bulk_create(lista_new_funcionario)
            lista_new_funcionario=[]

        if len(lista_new_provento)>150:
            Provento.objects.bulk_create(lista_new_provento)
            lista_new_provento=[]


        row+=1

    for kl in lista_eventos_nao_cadastrados:
        objeto = SemCadastro(
            id_municipio=id_municipio,
            tabela='evento',
            descricao=kl
        )
        lista_sem_cadastro.append(objeto)


    for kl in lista_funcoes_nao_cadastrados:
        objeto = SemCadastro(
            id_municipio=id_municipio,
            tabela='funcao',
            descricao=kl
        )
        lista_sem_cadastro.append(objeto)


    for kl in lista_setores_nao_cadastrados:
        objeto = SemCadastro(
            id_municipio=id_municipio,
            tabela='setor',
            descricao=kl
        )
        lista_sem_cadastro.append(objeto)

    for kl in lista_secretarias_nao_cadastrados:
        objeto = SemCadastro(
            id_municipio=id_municipio,
            tabela='secretaria',
            descricao=kl
        )
        lista_sem_cadastro.append(objeto)


    if len(lista_sem_cadastro)>0:
        SemCadastro.objects.bulk_create(lista_sem_cadastro)

    if processo==False:
        return str(id_municipio)

    if len(lista_new_complemento)>0:
        Complemento.objects.bulk_create(lista_new_complemento)
        lista_new_complemento=[]

    if len(lista_new_ampliacao_ch)>0:
        Ampliacao_ch.objects.bulk_create(lista_new_ampliacao_ch)
        lista_new_ampliacao_ch=[]

    if len(lista_new_suporte)>0:
        Suporte.objects.bulk_create(lista_new_suporte)
        lista_new_suporte=[]

    if len(lista_new_funcionario)>0:
        Funcionario.objects.bulk_create(lista_new_funcionario)
        lista_new_funcionario=[]

    if len(lista_new_provento)>0:
        Provento.objects.bulk_create(lista_new_provento)
        lista_new_provento=[]
    if len(lista_new_pos)>0:
        Posgraduation.objects.bulk_create(lista_new_pos)
        lista_new_pos=[]

    if len(lista_new_csalario)>0:
        CaraterSalario.objects.bulk_create(lista_new_csalario)
        lista_new_csalario=[]

    return '0'



def letra_da_coluna(pi):
    lista1=['A','B','C']
    lista2=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    lista3=[]
    for k1 in lista1:
        alfa=k1
        for k2 in lista2:
            lista3.append(alfa+k2)
    listaColunas=lista2+lista3
    return listaColunas[pi-1]


def fun_titulo_coluna(titulo_coluna):
    titulo_coluna=titulo_coluna.upper()
    titulo_coluna=titulo_coluna.strip()
    array = titulo_coluna.split("-",1)

    if len(array)>1:
        titulo = array[1].strip()
        return titulo
    else:
        return titulo_coluna


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows

def remover_acentuacao(string: str) -> str:
    normalized = unicodedata.normalize('NFD', string)
    return ''.join(
        [l for l in normalized if not unicodedata.combining(l)]
    )



def fun_calcula_carga_horaria(pcarga_horaria):
    if pcarga_horaria==10:
        return 100
    elif pcarga_horaria==20:
        return 100
    elif pcarga_horaria==30:
        return 150
    elif pcarga_horaria==40:
        return 200
    else:
        return pcarga_horaria


def eventos_tabela_complementos(id_municipio):
    #Coluna da tabela Excel que vem do municipio "Cod. Evento"
    if id_municipio==38:
        return [1,9,227]
    else:
        return [1]


