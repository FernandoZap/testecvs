# -*- coding: utf-8 -*-

#import pyodbc as p
import openpyxl
import os
import sys
import datetime
import unicodedata
from .models import Salario,Situacao
from . import listagens



def importacao(planilha,ano,mes,id_municipio,current_user):
    lote = str(datetime.datetime.now().today())[0:19]


    (dict_secretarias1,dict_secretarias2)=listagens.dict_secretarias(id_municipio)
    (dict_setores1,dict_setores2,lista_setores_fundeb)=listagens.dict_setores(id_municipio)
    (dict_funcoes1,dict_funcoes2)=listagens.dict_funcoes(id_municipio)
    (dict_eventos1,dict_eventos2,lista_eventos_suporte,lista_eventos_ampliacao_ch)=listagens.dict_eventos(id_municipio)
    dict_natureza = listagens.dict_natureza(id_municipio)
    dict_tab_salarios = listagens.dict_tab_salarios(id_municipio)


    idop = current_user

    wb = openpyxl.load_workbook(planilha)
    sheets = wb.sheetnames

    sheet0 = sheets[0]

    sheet = wb.get_sheet_by_name(sheet0)

    row=2
    lista_situacao=[]
    lista_salario=[]
    lista_d1=[]
    lista_d2=[]

    erro=0
    while row<sheet.max_row+1 and row<31000:

        cod_servidor = sheet['A' + str(row)].value
        codigo_folha = sheet['C' + str(row)].value
        nome_servidor = sheet['I' + str(row)].value
        carga_horaria = sheet['L' + str(row)].value
        data_admissao = sheet['M' + str(row)].value
        secretaria = sheet['W' + str(row)].value
        setor = sheet['X' + str(row)].value
        funcao = sheet['AB' + str(row)].value
        natureza = sheet['AF' + str(row)].value
        tipo_evento = sheet['AI' + str(row)].value
        evento = sheet['AJ' + str(row)].value
        ref_evento = sheet['AK' + str(row)].value
        valor_evento = sheet['AL' + str(row)].value
        codigo_evento = sheet['AV' + str(row)].value

        if tipo_evento=='4':
            row+=1
            continue
        if codigo_folha>mes:
            row+=1
            continue
        if codigo_folha<mes:
            row+=1
            continue

        nome_servidor = remover_acentuacao(nome_servidor)
        secretaria = remover_acentuacao(secretaria)
        setor = remover_acentuacao(setor)
        funcao = remover_acentuacao(funcao)
        evento = remover_acentuacao(evento)
        natureza = remover_acentuacao(natureza)

        nome_servidor=nome_servidor.strip()
        secretaria=secretaria.strip()
        setor=setor.strip()
        funcao=funcao.strip()
        evento=evento.strip()
        natureza=natureza.strip()

        id_secretaria=dict_secretarias1.get(secretaria,0)
        id_setor=dict_setores1.get(setor,0)
        id_funcao=dict_funcoes1.get(funcao,0)
        id_evento=dict_eventos1.get(evento,0)
        id_natureza=dict_natureza.get(natureza,0)

        id_secretaria_out=dict_secretarias2.get(id_secretaria,0)
        id_setor_out=dict_setores2.get(id_setor,0)
        id_funcao_out=dict_funcoes2.get(id_funcao,0)
        id_evento_out=dict_eventos2.get(id_evento,0)

        if id_setor in lista_setores_fundeb:
            fundeb='S'
        else:
            fundeb='N'

        if id_evento in lista_eventos_suporte:
            suporte='S'
        else:
            suporte='N'

        if id_evento in lista_eventos_ampliacao_ch:
            ampliacao_ch='S'
        else:
           ampliacao_ch='N'

        anomes=str(ano)+str(mes+100)[1:]

        if int(codigo_evento)==1 and fundeb=='S':
            carga_100h=100
            carga_200h=200
            v_sal_100 = dict_tab_salarios.get(str(id_funcao)+'-'+str(carga_100h),0)
            v_sal_200 = dict_tab_salarios.get(str(id_funcao)+'-'+str(carga_200h),0)
            carga_horaria_origem = carga_horaria
            carga_horaria = get_carga_horaria(id_municipio,carga_horaria)
            num_dias=int(ref_evento.replace('/30',''))
            salario_30d = get_salario_30dias(valor_evento,num_dias)
            if v_sal_100==0:
                salario_100h = get_salario_100horas(salario_30d,carga_horaria)
            else:
                salario_100h =  v_sal_100
            participacao =  get_participacao(salario_30d,carga_horaria,valor_evento)
            if carga_horaria <=150:
                carga_horaria=100
            else:
                carga_horaria=200
                salario=valor_evento
            if cod_servidor not in lista_d2:
                objeto_carga = Salario (
                    id_municipio = id_municipio,
                    anomes = anomes,
                    cod_servidor = cod_servidor,
                    vencimento_base=valor_evento,
                    carga_horaria = carga_horaria,
                    num_dias = num_dias,
                    carga_horaria_origem = carga_horaria,
                    num_dias_origem = ref_evento,
                    salario_100h = salario_100h,
                    participacao=participacao
                )
                lista_salario.append(objeto_carga)
                lista_d2.append(cod_servidor)
        if len(lista_salario)>250:
            Salario.objects.bulk_create(lista_salario)
            lista_salario=[]

        if cod_servidor not in lista_d1:
            objeto_carga = Situacao(
                id_municipio=id_municipio,
                anomes=anomes,
                nome=nome_servidor,
                cod_servidor=cod_servidor,
                data_admissao = data_admissao,
                natureza = id_natureza,
                id_secretaria_origem=id_secretaria,
                id_setor_origem=id_setor,
                id_funcao_origem=id_funcao,
                id_evento_origem=id_evento,
                id_secretaria=id_secretaria_out,
                id_setor=id_setor_out,
                id_funcao=id_funcao_out,
                id_evento=id_evento_out,
                suporte=suporte
            )
            lista_situacao.append(objeto_carga)
            lista_d1.append(cod_servidor)
        if len(lista_situacao)>250:
            Situacao.objects.bulk_create(lista_situacao)
            lista_situacao=[]

        row+=1

    if len(lista_situacao)>0:
        Situacao.objects.bulk_create(lista_situacao)
    if len(lista_salario)>0:
        Salario.objects.bulk_create(lista_salario)

    return None


def remover_acentuacao(string: str) -> str:
    normalized = unicodedata.normalize('NFD', string)
    return ''.join(
        [l for l in normalized if not unicodedata.combining(l)]
    )


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows

def get_carga_horaria(vid_municipio,vcarga_horaria):
    if vcarga_horaria==20:
        retorno=100
    elif vcarga_horaria==30:
        retorno=150
    elif vcarga_horaria==40:
        retorno=200
    elif vcarga_horaria==50:
        retorno=100
    elif vcarga_horaria==220:
        retorno=200
    else:
        retorno=vcarga_horaria
    return retorno

def get_salario_30dias(vvalor_evento,vnum_dias):
    salario_30=round((vvalor_evento/vnum_dias)*30,2)
    return salario_30

def get_salario_100horas(vsalario_30d,vcarga_horaria):
    salario=round((vsalario_30d/vcarga_horaria)*100,2)
    return salario

def get_participacao(salario_30d,carga_horaria,valor_evento):
    if carga_horaria<150:
        fator=100
    else:
        fator=200
    divisor = (salario_30d/carga_horaria)*fator
    retorno = valor_evento/divisor

    return round(retorno,2)
