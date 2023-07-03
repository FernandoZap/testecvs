# -*- coding: utf-8 -*-

from django.http import HttpResponseRedirect
import openpyxl
import os
import sys
import datetime
import unicodedata
from .models import Tabela_salario
from . import listagens,dicionarios


def tabela_salario(excel,ano,num_mes,anomes,id_municipio):
    (dict_set1,dict_set2,lista_setores_fundeb)=dicionarios.de_setores(id_municipio)
    (dict_evento1,dict_evento2,lista_eventos_suporte,lista_eventos_ampliacaoch)=dicionarios.de_eventos(id_municipio)
    (dict_func1,dict_func2,lista_funcoes_prof)=dicionarios.de_funcoes(id_municipio)
    dict_carga_horaria=dicionarios.de_tab_salariosch(id_municipio)

    lista_eventos_complementos=eventos_tabela_complementos(id_municipio)


    num_mes=9
    wb = openpyxl.load_workbook(excel)
    sheets = wb.sheetnames
    Tabela_salario.truncate()

    #sheet = sheets[0]
    sheet = wb.active
    nrows = get_maximum_rows(sheet_object=sheet)


    row=2
    erro=0
    erros=0
    qtde_itens=0
    retorno = True


    erro=0
    lista=[]
    codigos_ja_incluidos=[]
    lista_ch=[]
    ref_eventos_ja_incluidos=[]
    lista_funcionario_suporte=[]
    lista_retorno=[]

    lista_new_suporte=[]
    lista_new_complemento=[]
    lista_new_ampliacao_ch=[]
    lista_new_funcionario=[]
    lista_new_provento=[]

    lista_eventos_nao_cadastrados=[]
    lista_funcoes_nao_cadastrados=[]
    lista_setores_nao_cadastrados=[]
    lista_secretarias_nao_cadastrados=[]
    lista_sem_cadastro=[]

    lista_funcao_sal100=[]
    lista_valor_int_sal100=[]
    lista_valor_dec_sal100=[]
    lista_new_salario=[]
    lista_salarios_gravados=[]
    processo=True


    lista_mes=['1','2','3','4','5','6','7','8','9','10','11','12']
    lixo=0
    lista_salario_not_found=[]

    flag=0
    maximo=0
    max_col = sheet.max_column
    row=1
    limite=30000

    while row<limite and row<nrows+1:

        if row==1:
            for i in range(1, max_col + 1):
                maximo+=1
                cell_obj = sheet.cell(row = 1, column = i)
                titulo_coluna=cell_obj.value
                if titulo_coluna is not None and 1==1:
                    titulo_coluna=remove_combining_fluent(titulo_coluna)
                    titulo_coluna=titulo_coluna.strip()
                    if titulo_coluna=='Codigo Servidor':
                        n_codigo_servidor=letra_da_coluna(i)
                    if titulo_coluna=='Codigo Folha':
                        n_codigo_folha=letra_da_coluna(i)
                    elif titulo_coluna=='Carga Horaria':
                        n_carga_horaria=letra_da_coluna(i)
                    elif titulo_coluna=='Setor':
                        n_setor=letra_da_coluna(i)
                    elif titulo_coluna=='Funcao':
                        n_funcao=letra_da_coluna(i)
                    elif titulo_coluna=='Tipo':
                        n_tipo=letra_da_coluna(i)
                    elif titulo_coluna=='Evento':
                        n_evento=letra_da_coluna(i)
                    elif titulo_coluna=='Cod. Evento':
                        n_cod_evento=letra_da_coluna(i)
                    elif titulo_coluna=='Ref. Evento':
                        n_ref_evento=letra_da_coluna(i)
                    elif titulo_coluna=='Valor Evento':
                        n_valor_evento=letra_da_coluna(i)
                if maximo>20:
                    break

            row+=1
            continue

        if sheet[n_codigo_folha + str(row)].value!=num_mes:
            row+=1
            continue

        codigo_funcionario = sheet[n_codigo_servidor + str(row)].value
        carga_horaria = sheet[n_carga_horaria + str(row)].value
        setor  = str((sheet[n_setor + str(row)].value)).strip()
        funcao = str((sheet[n_funcao + str(row)].value)).strip()
        tipo = str(sheet[n_tipo + str(row)].value)
        cod_evento = sheet[n_cod_evento + str(row)].value
        evento = sheet[n_evento + str(row)].value

        if tipo=='4':
            row+=1
            continue

        if cod_evento>1:
            row+=1
            continue

        evento=remove_combining_fluent(evento)
        evento=evento.strip()

        id_evento_origem=dict_evento1.get(evento,0)

        id_evento=dict_evento2.get(id_evento_origem,0)

        ref_evento    = sheet[n_ref_evento + str(row)].value
        valor_evento = sheet[n_valor_evento + str(row)].value

        setor=remove_combining_fluent(setor)
        funcao=remove_combining_fluent(funcao)

        setor=setor.strip()
        funcao=funcao.strip()

        id_setor_origem=dict_set1.get(setor,0)
        id_setor=dict_set2.get(id_setor_origem,0)

        id_funcao_origem=dict_func1.get(funcao,0)

        if id_setor_origem in lista_setores_fundeb:
            fundeb='S'
        elif id_setor in lista_setores_fundeb:
            fundeb='S'
        else:
           fundeb='N'

        if fundeb=='N':
            row+=1
            continue

        refev= ref_evento.replace('/30','')
        if refev!='30':
            row+=1
            continue

        ref_evento=ref_evento.strip()
        carga_horaria_origem=carga_horaria

        if cod_evento in lista_eventos_complementos and fundeb=='S':
            if codigo_funcionario not in ref_eventos_ja_incluidos:
                ref_eventos_ja_incluidos.append(codigo_funcionario)
                num_dias=int(ref_evento.replace('/30',''))
                vencimento_base=valor_evento
                salario=valor_evento
                salario_100H=valor_evento

                if num_dias!=30:
                    salario_30_dias=round((valor_evento/num_dias)*30,2)
                else:
                    salario_30_dias=valor_evento


                str_ch_errada=str(id_funcao_origem)+str(int(salario_30_dias))+str(carga_horaria)

                ch_certa=dict_carga_horaria.get(str_ch_errada,'000')
                if ch_certa!='000':
                    carga_horaria=int(ch_certa)


                carga_horaria_origem=carga_horaria

                carga_horaria = fun_calcula_carga_horaria(carga_horaria_origem)

                #lista_de_salarios_dec=dict_salario_dec.get(str(id_funcao_origem),[])
                #lista_de_salarios_int=dict_salario_int.get(str(id_funcao_origem),[])

                if carga_horaria>=50 and carga_horaria<=150:
                    if carga_horaria==100 and num_dias==30:
                        salario=salario_30_dias
                        salario_100H=salario_30_dias
                    else:
                        salario=round((salario_30_dias/carga_horaria)*100,2)

                        salario_100H=salario
                    participacao=round(valor_evento/salario,2)
                    carga_horaria=100
                elif carga_horaria>150 and carga_horaria<=250:
                    if carga_horaria==200 and num_dias==30:
                        salario = salario_30_dias
                        salario_200H = salario_30_dias
                    else:
                        salario=round((salario_30_dias/carga_horaria)*200,2)
                        salario_200H=salario

                    salario_100H=round(salario_200H/2,2)

                    participacao=round(valor_evento/salario_200H,2)
                    carga_horaria=200


                chave = str(id_funcao_origem)+'-'+str(int(salario_100H))
                if chave not in lista_salarios_gravados:
                    lista_salarios_gravados.append(chave)
                    new_salario=Tabela_salario(
                        id_municipio=id_municipio,
                        id_funcao_origem=id_funcao_origem,
                        salario_100h_int=int(salario_100H),
                        salario_100h_dec=salario_100H
                        )
                    lista_new_salario.append(new_salario)
                if len(lista_new_salario)>200:
                    Tabela_salario.objects.bulk_create(lista_new_salario)
                    lista_new_salario=[]

        row+=1
    if len(lista_new_salario)>0:
        Tabela_salario.objects.bulk_create(lista_new_salario)

    return '0'

def remove_combining_fluent(string: str) -> str:
    normalized = unicodedata.normalize('NFD', string)
    return ''.join(
        [l for l in normalized if not unicodedata.combining(l)]
    )


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def eventos_tabela_complementos(id_municipio):
    if id_municipio==38:
        return [1,9,227]
    else:
        return [1]

def fun_calcula_carga_horaria(pcarga_horaria):
    if pcarga_horaria==20:
        return 100
    elif pcarga_horaria==30:
        return 150
    elif pcarga_horaria==40:
        return 200
    elif pcarga_horaria==50:
        return 100
    elif pcarga_horaria==220:
        return 200
    return pcarga_horaria


def eventos_tabela_complementos(id_municipio):
    if id_municipio==38:
        return [1,9,227]
    else:
        return [1]

def letra_da_coluna(pi):
    lista1=['A','B','C']
    lista2=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    lista3=[]
    for k1 in range(len(lista1)):
        alfa=lista1[k1]
        for k2 in range(len(lista2)):
            lista3.append(alfa+lista2[k2])
    listaColunas=lista2+lista3
    return listaColunas[pi-1]

