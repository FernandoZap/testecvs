# -*- coding: utf-8 -*-
from django.http import HttpResponseRedirect
import openpyxl
import os
import sys
import datetime
import unicodedata
from .models import Funcionario,Provento,Complemento,Ampliacao_ch,Suporte,SemCadastro,Information,Tabela,Secretaria,Setor,Funcao,Evento,LogErro
from . import listagens,dicionarios

from celery import shared_task



@shared_task
def saveFuncao(id_municipio):
    query = [func.funcao for func in Funcao.objects.filter(id_municipio=44)]
    for q in range(len(query)):
        if q<5:
            LogErro.objects.create(id_municipio=10,anomes=202209,numero_linha=15,codigo='celery',observacao=query[q])


'''
@shared_task
def import_excel(excel,ano,num_mes,anomes,id_municipio,current_user,truncate):
    if truncate:
        Funcionario.truncate()
        Provento.truncate()
        Complemento.truncate()
        Ampliacao_ch.truncate()
        Suporte.truncate()
        SemCadastro.truncate()

    (dict_sec1,dict_sec2)=dicionarios.de_secretarias(id_municipio)
    (dict_set1,dict_set2,lista_setores_fundeb)=dicionarios.de_setores(id_municipio)
    (dict_evento1,dict_evento2,lista_eventos_suporte,lista_eventos_ampliacaoch)=dicionarios.de_eventos(id_municipio)
    (dict_func1,dict_func2,lista_funcoes_prof)=dicionarios.de_funcoes(id_municipio)
    (dict_salario_int,dict_salario_dec)=dicionarios.de_tabela_salarios(id_municipio)
    dict_carga_horaria=dicionarios.de_tab_salariosch(id_municipio)
    dict_vinculos = dicionarios.de_vinculos(id_municipio)

    lista_eventos_complementos=eventos_tabela_complementos(id_municipio)
    lista_sal100=[]
    lista_sal200=[]

    wb = openpyxl.load_workbook(excel)
    sheets = wb.sheetnames

    #sheet = sheets[0]
    sheet = wb.active
    nrows = get_maximum_rows(sheet_object=sheet)
    max_col = sheet.max_column


    erro=0
    erros=0
    qtde_itens=0
    retorno = True

    erro=0
    lista=[]
    codigos_ja_incluidos=[]
    lista_ch=[]
    ref_eventos_ja_incluidos=[]
    lista_funcionario_suporte=[]
    lista_retorno=[]

    lista_new_suporte=[]
    lista_new_complemento=[]
    lista_new_ampliacao_ch=[]
    lista_new_funcionario=[]
    lista_new_provento=[]

    lista_eventos_nao_cadastrados=[]
    lista_funcoes_nao_cadastrados=[]
    lista_setores_nao_cadastrados=[]
    lista_secretarias_nao_cadastrados=[]
    lista_sem_cadastro=[]

    lista_funcao_sal100=[]
    lista_valor_int_sal100=[]
    lista_valor_dec_sal100=[]
    processo=True


    lista_mes=['1','2','3','4','5','6','7','8','9','10','11','12']
    lixo=0
    lista_salario_not_found=[]

    flag=0

    row=1
    limite=30000
    maximo=0

    planilha_completa=False
    while row<limite and row<nrows+1:

        if row==1:
            for i in range(1, max_col + 1):
                maximo+=1
                cell_obj = sheet.cell(row = 1, column = i)
                titulo_coluna=cell_obj.value
                if titulo_coluna is not None:
                    titulo_coluna=remover_acentuacao(titulo_coluna)
                    titulo_coluna=titulo_coluna.strip()
                    if titulo_coluna=='Codigo Servidor':
                        n_codigo_servidor=letra_da_coluna(i)
                    elif titulo_coluna=='Codigo Folha':
                        n_mes_folha=letra_da_coluna(i)
                    elif titulo_coluna=='Nome do Servidor':
                        n_nome_funcionario=letra_da_coluna(i)
                    elif titulo_coluna=='Secretaria':
                        n_secretaria=letra_da_coluna(i)
                    elif titulo_coluna=='Setor':
                        n_setor=letra_da_coluna(i)
                    elif titulo_coluna=='Previdencia':
                        n_previdencia=letra_da_coluna(i)
                    elif titulo_coluna=='Carga Horaria':
                        n_carga_horaria=letra_da_coluna(i)
                    elif titulo_coluna=='Natureza':
                        n_natureza=letra_da_coluna(i)
                    elif titulo_coluna=='Data de Admissao':
                        n_data_admissao=letra_da_coluna(i)
                    elif titulo_coluna=='Funcao':
                        n_funcao=letra_da_coluna(i)
                    elif titulo_coluna=='Evento':
                        n_evento=letra_da_coluna(i)
                    elif titulo_coluna=='Codigo Folha':
                        n_codigo_folha=letra_da_coluna(i)
                    elif titulo_coluna=='Ref. Evento':
                        n_ref_evento=letra_da_coluna(i)
                    elif titulo_coluna=='Valor Evento':
                        n_valor_evento=letra_da_coluna(i)
                    elif titulo_coluna=='Cod. Evento':
                        n_codigo_evento=letra_da_coluna(i)
                    elif titulo_coluna=='Salario Base':
                        n_salario_base=letra_da_coluna(i)
                    elif titulo_coluna=='Tipo':
                        n_tipo=letra_da_coluna(i)
                if maximo>64:
                    break

            row+=1
            continue

        if sheet['A'+str(row)].value is None:
            break
        if str(sheet['A'+str(row)].value)=='':
            break

        codigo_funcionario = sheet[n_codigo_servidor + str(row)].value
        mes_da_folha = sheet[n_mes_folha + str(row)].value
        nome_funcionario = (sheet[n_nome_funcionario + str(row)].value)
        carga_horaria = sheet[n_carga_horaria + str(row)].value
        data_admissao    = sheet[n_data_admissao + str(row)].value
        secretaria  = sheet[n_secretaria + str(row)].value
        funcao = sheet[n_funcao + str(row)].value
        setor = sheet[n_setor + str(row)].value
        previdencia = sheet[n_previdencia + str(row)].value
        natureza = sheet[n_natureza + str(row)].value
        evento = sheet[n_evento + str(row)].value
        tipo       = sheet[n_tipo + str(row)].value
        ref_evento = sheet[n_ref_evento + str(row)].value
        valor_evento = sheet[n_valor_evento + str(row)].value
        cod_evento = sheet[n_codigo_evento + str(row)].value
        salario_base = sheet[n_salario_base + str(row)].value



        if codigo_funcionario is None:
            row+=1
            break


        if id_municipio!=38:
            if mes_da_folha!=num_mes:
                row+=1
                continue

        if tipo=='4':
            row+=1
            continue

        evento=remover_acentuacao(evento)
        evento=evento.strip()

        id_evento_origem=dict_evento1.get(evento,0)
        if id_evento_origem==0:
            if evento not in lista_eventos_nao_cadastrados:
                lista_eventos_nao_cadastrados.append(evento)
                processo=False

        id_evento=dict_evento2.get(id_evento_origem,0)


        cod_previdencia=remover_acentuacao(previdencia)

        if cod_previdencia=='PREVIDENCIA MUNICIPAL':
            previdencia='M'
        elif cod_previdencia=='INSS':
            previdencia='I'
        elif cod_previdencia=='NÃO PAGA':
            previdencia='N'
        else:
            previdencia=''

        secretaria=remover_acentuacao(secretaria)
        setor=remover_acentuacao(setor)
        funcao=remover_acentuacao(funcao)

        nome_funcionario=remover_acentuacao(nome_funcionario)
        natureza=remover_acentuacao(natureza)
        secretaria=secretaria.strip()
        setor=setor.strip()
        funcao=funcao.strip()
        natureza=natureza.strip()
        nome_funcionario=nome_funcionario.strip()

        id_natureza=dict_vinculos.get(natureza,0)
        id_secretaria_origem=dict_sec1.get(secretaria,0)
        id_setor_origem=dict_set1.get(setor,0)
        id_funcao_origem=dict_func1.get(funcao,0)


        if id_secretaria_origem==0:
            if secretaria not in lista_secretarias_nao_cadastrados:
                lista_secretarias_nao_cadastrados.append(secretaria)
                processo=False

        if id_setor_origem==0:
            if setor not in lista_setores_nao_cadastrados:
                lista_setores_nao_cadastrados.append(setor)
                processo=False

        if id_funcao_origem==0:
            if funcao not in lista_funcoes_nao_cadastrados:
                lista_funcoes_nao_cadastrados.append(funcao)
                processo=False

        if processo==False:
            row+=1
            continue


        id_secretaria=dict_sec2.get(id_secretaria_origem,0)
        id_setor=dict_set2.get(id_setor_origem,0)
        id_funcao=dict_func2.get(id_funcao_origem,0)

        grupamento='S'
        classificacao='O'

        if id_setor in lista_setores_fundeb:
            fundeb='S'
        elif id_setor_origem in lista_setores_fundeb:
            fundeb='S'
        else:
            fundeb='N'

        tipo=str(tipo)

        data_admissao=str(data_admissao)[0:10]
        if ref_evento is None:
            ref_evento=''
        if classificacao is None:
            classificacao=''


        classificacao=classificacao.strip()
        ref_evento=ref_evento.strip()
        participacao=0
        carga_horaria_origem=carga_horaria

        if cod_evento in lista_eventos_complementos and fundeb=='S':
            if codigo_funcionario not in ref_eventos_ja_incluidos:
                ref_eventos_ja_incluidos.append(codigo_funcionario)
                num_dias=int(ref_evento.replace('/30',''))

                str_ch_errada=str(id_funcao_origem)+str(int(salario_base))+str(carga_horaria)

                ch_certa=dict_carga_horaria.get(str_ch_errada,'000')
                if ch_certa!='000':
                    carga_horaria=int(ch_certa)


                carga_horaria_origem=carga_horaria

                if carga_horaria_origem==20:
                    carga_horaria=100
                elif carga_horaria_origem==30:
                    carga_horaria=150
                elif carga_horaria_origem==40:
                    carga_horaria=200
                elif carga_horaria_origem==50:
                    carga_horaria=100
                elif carga_horaria_origem==220:
                    carga_horaria=200
                else:
                    carga_horaria=carga_horaria_origem

                if carga_horaria>=50 and carga_horaria<=150:
                    if carga_horaria==100 and num_dias==30:
                        salario=salario_base
                        salario_100H=salario_base
                    else:
                        salario=round((salario_base/carga_horaria)*100,3)

                        salario_100H=salario
                    participacao=round(valor_evento/salario,6)
                    carga_horaria=100
                elif carga_horaria>150:
                    if carga_horaria==200:
                        salario = salario_base
                        salario_200H = salario_base
                    else:
                        salario=round((salario_base/carga_horaria)*200,3)
                        salario_200H=salario


                    salario_100H=round(salario_200H/2,3)

                    participacao=round(valor_evento/salario_200H,6)
                    carga_horaria=200

                lista_sal100=dict_salario_int.get(id_funcao_origem,[])
                lista_sal200=dict_salario_dec.get(id_funcao_origem,[])

                if len(lista_sal100)>0:
                    salario_100H = fun_salario100h(salario_100H,lista_sal100,lista_sal200)

                new_complemento=Complemento(
                    id_municipio=id_municipio,
                    anomes=anomes,
                    codigo=codigo_funcionario,
                    salario=salario,
                    salario_100H=salario_100H,
                    num_dias=num_dias,
                    participacao=participacao,
                    vencimento_base=valor_evento,
                    fundeb=fundeb
                )
                lista_new_complemento.append(new_complemento)


        if fundeb=='S' and id_evento_origem in lista_eventos_ampliacaoch and codigo_funcionario not in lista_ch: # and id_funcao_origem in lista_funcoes_prof:
            new_ampliacao_ch = Ampliacao_ch(
                id_municipio=id_municipio,
                anomes=anomes,
                cod_servidor=codigo_funcionario
            )
            lista_new_ampliacao_ch.append(new_ampliacao_ch)
            lista_ch.append(codigo_funcionario)

        if fundeb=='S' and id_evento_origem in lista_eventos_suporte and codigo_funcionario not in lista_funcionario_suporte:
            new_suporte = Suporte(
                id_municipio=id_municipio,
                anomes=anomes,
                cod_servidor=codigo_funcionario
            )
            lista_new_suporte.append(new_suporte)
            lista_funcionario_suporte.append(codigo_funcionario)


        if codigo_funcionario not in codigos_ja_incluidos:
            new_funcionario = Funcionario(
                id_municipio=id_municipio,
                anomes=anomes,
                codigo=codigo_funcionario,
                nome_servidor=nome_funcionario,
                carga_horaria=carga_horaria,
                tipo_admissao=natureza,
                data_admissao=data_admissao,
                previdencia=previdencia,
                id_secretaria=id_secretaria,
                id_setor=id_setor,
                id_funcao=id_funcao,
                id_secretaria_origem = id_secretaria_origem,
                id_setor_origem = id_setor_origem,
                id_funcao_origem = id_funcao_origem,
                fundeb=fundeb,
            )
            lista_new_funcionario.append(new_funcionario)
            codigos_ja_incluidos.append(codigo_funcionario)

        if tipo in ['1','2','3']:
            lixo+=1
            tipo='1'
            new_provento = Provento(
                id_municipio=id_municipio,
                anomes=anomes,
                codigo=codigo_funcionario,
                previdencia=previdencia,
                tipo=tipo,
                id_evento=id_evento,
                id_evento_origem=id_evento_origem,
                valor_evento=valor_evento,
                classificacao=classificacao,
                grupamento=grupamento,
                lixo=lixo
            )
            lista_new_provento.append(new_provento)

        if len(lista_new_complemento)>250:
            Complemento.objects.bulk_create(lista_new_complemento)
            lista_new_complemento=[]

        if len(lista_new_ampliacao_ch)>150:
            Ampliacao_ch.objects.bulk_create(lista_new_ampliacao_ch)
            lista_new_ampliacao_ch=[]

        if len(lista_new_suporte)>150:
            Suporte.objects.bulk_create(lista_new_suporte)
            lista_new_suporte=[]

        if len(lista_new_funcionario)>250:
            Funcionario.objects.bulk_create(lista_new_funcionario)
            lista_new_funcionario=[]

        if len(lista_new_provento)>300:
            Provento.objects.bulk_create(lista_new_provento)
            lista_new_provento=[]

        row+=1

    if processo==False:
        if len(lista_eventos_nao_cadastrados)>0:
            for k in range(len(lista_eventos_nao_cadastrados)):
                objeto = SemCadastro(
                    id_municipio=id_municipio,
                    tabela='evento',
                    descricao=lista_eventos_nao_cadastrados[k]
                )
                lista_sem_cadastro.append(objeto)
        if len(lista_funcoes_nao_cadastrados)>0:
            for k in range(len(lista_funcoes_nao_cadastrados)):
                objeto = SemCadastro(
                    id_municipio=id_municipio,
                    tabela='funcao',
                    descricao=lista_funcoes_nao_cadastrados[k]
                )
                lista_sem_cadastro.append(objeto)
        if len(lista_setores_nao_cadastrados)>0:
            for k in range(len(lista_setores_nao_cadastrados)):
                objeto = SemCadastro(
                    id_municipio=id_municipio,
                    tabela='setor',
                    descricao=lista_setores_nao_cadastrados[k]
                )
                lista_sem_cadastro.append(objeto)
        if len(lista_secretarias_nao_cadastrados)>0:
            for k in range(len(lista_secretarias_nao_cadastrados)):
                objeto = SemCadastro(
                    id_municipio=id_municipio,
                    tabela='secretaria',
                    descricao=lista_secretarias_nao_cadastrados[k]
                )
                lista_sem_cadastro.append(objeto)

        if len(lista_sem_cadastro)>0:
            SemCadastro.objects.bulk_create(lista_sem_cadastro)

    if processo==False:
        return str(id_municipio)

    if len(lista_new_complemento)>0:
        Complemento.objects.bulk_create(lista_new_complemento)
        lista_new_complemento=[]

    if len(lista_new_ampliacao_ch)>0:
        Ampliacao_ch.objects.bulk_create(lista_new_ampliacao_ch)
        lista_new_ampliacao_ch=[]

    if len(lista_new_suporte)>0:
        Suporte.objects.bulk_create(lista_new_suporte)
        lista_new_suporte=[]

    if len(lista_new_funcionario)>0:
        Funcionario.objects.bulk_create(lista_new_funcionario)
        lista_new_funcionario=[]

    if len(lista_new_provento)>0:
        Provento.objects.bulk_create(lista_new_provento)
        lista_new_provento=[]

'''





def remover_acentuacao(string: str) -> str:
    normalized = unicodedata.normalize('NFD', string)
    return ''.join(
        [l for l in normalized if not unicodedata.combining(l)]
    )


def get_maximum_rows(*, sheet_object):
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def eventos_tabela_complementos(id_municipio):
    if id_municipio==38:
        return [1,9,227]
    else:
        return [1]


def fun_salario100h(psalario_100h,plista_sal100,plista_sal200):
    sal100_0 = int(psalario_100h)-1
    sal100_1 = int(psalario_100h)
    sal100_2 = int(psalario_100h)+1

    for k in range(len(plista_sal100)):
        if plista_sal100[k]==sal100_0:
            return plista_sal200[k]
    for k in range(len(plista_sal100)):
        if plista_sal100[k] == sal100_1:
            return plista_sal200[k]
    for k in range(len(plista_sal100)):
        if plista_sal100[k] == sal100_2:
            return plista_sal200[k]

    return psalario_100h



def letra_da_coluna(pi):
    lista1=['A','B','C']
    lista2=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    lista3=[]
    for k1 in range(len(lista1)):
        alfa=lista1[k1]
        for k2 in range(len(lista2)):
            lista3.append(alfa+lista2[k2])
    listaColunas=lista2+lista3
    return listaColunas[pi-1]
