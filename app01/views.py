from django.shortcuts import render,redirect
from django.views.generic import (ListView)
from django.http import HttpResponse,HttpResponseRedirect,Http404
from . import choices ,importar_planilha_folha,importacao_excel,processamentoFolha,listagens,funcoes,tabela_salario,importacao_excel_folha,dicionarios,importacao_excel_folha_Aspec,importacao_excel_folha_layout,funcoes_listar_models,funcoes_query
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from .models import Municipio,Folhames,Secretaria,Setor,Funcao,Evento,Folhaevento,Setor_fundeb,Tab_salarioch,Tab_salario,SemCadastro,Vinculo,CargaHoraria,Documento
from accounts.models import User
from django.db.models import Count,Sum,Min,Avg,Max
import csv
import datetime
import os
import json
import openpyxl
from django.core.files import File
from django.db import connection
import unicodedata

def mesReferencia(mes):
    lista_mes=['','JANEIRO','FEVEREIRO','MARÇO','ABRIL','MAIO','JUNHO','JULHO','AGOSTO','SETEMBRO','OUTUBRO','NOVEMBRO','DEZEMBRO']
    return lista_mes[int(mes)]


def cadastrarCargaHorariaErrada(request):
    municipios = Municipio.objects.filter(empresa__in=['SS','Layout','Aspec']).order_by('municipio')
    if request.method=="POST":
        lista_dados=[]
        id_municipio = int(request.POST['municipio'])
        id_funcao1 = int(request.POST['funcao1'])
        id_funcao2 = int(request.POST['funcao2'])
        id_funcao3 = int(request.POST['funcao3'])
        id_funcao4 = int(request.POST['funcao4'])

        valor1 = fun_valorch(request.POST['valor1'])
        valor2 = fun_valorch(request.POST['valor2'])
        valor3 = fun_valorch(request.POST['valor3'])
        valor4 = fun_valorch(request.POST['valor4'])

        cargah_errada1 =  fun_ch(request.POST['cargahoraria_e1'])
        cargah_errada2 =  fun_ch(request.POST['cargahoraria_e2'])
        cargah_errada3 =  fun_ch(request.POST['cargahoraria_e3'])
        cargah_errada4 =  fun_ch(request.POST['cargahoraria_e4'])

        cargah_correta1 =  fun_ch(request.POST['cargahoraria_c1'])
        cargah_correta2 =  fun_ch(request.POST['cargahoraria_c2'])
        cargah_correta3 =  fun_ch(request.POST['cargahoraria_c3'])
        cargah_correta4 =  fun_ch(request.POST['cargahoraria_c4'])

        lista_dados.append({'id_funcao':id_funcao1,'valor':valor1,'cargah_errada':cargah_errada1,'cargah_correta':cargah_correta1})
        lista_dados.append({'id_funcao':id_funcao2,'valor':valor2,'cargah_errada':cargah_errada2,'cargah_correta':cargah_correta2})
        lista_dados.append({'id_funcao':id_funcao3,'valor':valor3,'cargah_errada':cargah_errada3,'cargah_correta':cargah_correta3})
        lista_dados.append({'id_funcao':id_funcao4,'valor':valor4,'cargah_errada':cargah_errada4,'cargah_correta':cargah_correta4})


        for kitem in lista_dados:
            if kitem['id_funcao']>0:
                if kitem['valor']>0:
                    if kitem['cargah_errada']>0:
                        if kitem['cargah_correta']>0:
                            id_funcao=kitem['id_funcao']
                            valor=kitem['valor']
                            cargah_errada=kitem['cargah_errada']
                            cargah_correta=kitem['cargah_correta']


                            if not Tab_salarioch.objects.filter(id_municipio=id_municipio,id_funcao_origem=id_funcao,carga_horaria_errada=cargah_errada,valor=valor).exists():
                                Tab_salarioch.objects.create(id_municipio=id_municipio,id_funcao_origem=id_funcao,carga_horaria_errada=cargah_errada,valor=valor,carga_horaria_certa=cargah_correta)
                            else:
                                Tab_salarioch.objects.filter(id_municipio=id_municipio,id_funcao_origem=id_funcao,carga_horaria_errada=cargah_errada,valor=valor).update(carga_horaria_certa=cargah_correta)
        return HttpResponseRedirect('/app01/listar_ch')
    return render(request, 'app01/carga_horaria.html',{'titulo':'Carga Horaria','municipios':municipios})






def removerCargaHorariaErrada(request,id):
    obj = Tab_salarioch.objects.get(pk=id)
    obj.delete()
    return HttpResponseRedirect('/app01/listar_ch')



def listarCargaHorariaErrada(request):
    if request.method=="POST":

        id_municipio=request.POST['municipio']

        cursor = connection.cursor()
        cursor.execute("select t.id,m.municipio,f.funcao,t.valor,t.carga_horaria_errada,t.carga_horaria_certa \
            from tab_salariosch t,municipios m,funcoes f \
            where m.id_municipio=t.id_municipio and t.id_funcao_origem=f.id_funcao and m.id_municipio=%s",[id_municipio])
        query = dictfetchall(cursor)

        lista=[]
        for qp in query:
            id = qp['id']
            municipio = qp['municipio']
            funcao = qp['funcao']
            valor = qp['valor']
            ch_errada = qp['carga_horaria_errada']
            ch_certa = qp['carga_horaria_certa']

            lista.append(
                {
                 'id':id,'municipio':municipio,'funcao':funcao,'valor':valor,'ch_errada':ch_errada,'ch_certa':ch_certa
                }
            )
        cursor.close()
        del cursor

        return render(request, 'app01/listar_carga_horaria.html',{'lista':lista})
    else:
        municipios=Municipio.objects.filter(empresa__in=['Aspec','Layout','SS']).order_by('municipio')

    return render(request, 'app01/listar_ch_errada.html',{'municipios':municipios})






def tabela_salarios(request):
    municipios = Municipio.objects.filter(empresa__in=['SS','Layout','Aspec']).order_by('municipio')
    funcoes = Funcao.objects.all().order_by('funcao')
    if request.method=="POST":
        id_municipio = int(request.POST['municipio'])
        id_funcao = int(request.POST['funcao'])
        carga_horaria = int(request.POST['carga_horaria'])
        valor = (request.POST['valor_salario']).replace('R$','')
        valor = valor.replace('.','')
        valor = valor.replace(',','.')
        valor =  float(valor)
        if not Tab_salario.objects.filter(id_municipio=id_municipio,id_funcao=id_funcao,carga_horaria=carga_horaria).exists():
            Tab_salario.objects.create(id_municipio=id_municipio,id_funcao=id_funcao,carga_horaria=carga_horaria,valor=valor)
        return HttpResponseRedirect('/app01/tab_salarios')
    return render(request, 'app01/tabela_salarios.html',{'titulo':'Cadastrar Salario','municipios':municipios,'funcoes':funcoes})





def dictfetchall(cursor):
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]

def formatMilhar(valor):
    if valor=='':
        return valor
    elif valor is None:
        return valor
    if valor<0:
        valor=valor*(-1)
        sinal='-'
    else:
        sinal=''

    vd = f"{valor:,.2f}"
    vd = vd.replace('.','-')
    vd = vd.replace(',','.')
    vd = vd.replace('-',',')
    return sinal+vd



def importar_excel_folha(request):

    #sessao(request)
    current_user=166
    if (request.method == "POST" and request.FILES['filename']):

        id_municipio=request.POST['municipio']
        ano=request.POST['ano']
        mes=request.POST['mes']
        planilha=request.FILES['filename']
        anomes=str(ano)+ str(int(mes)+100)[1:]
        mesano=str(int(mes)+100)[1:]+str(ano)
        obj_municipio=Municipio.objects.get(id_municipio=id_municipio)
        empresa=obj_municipio.empresa

        atualizar=funcoes.atualizarTables()

        lista_retorno = funcoes.funFolhames(id_municipio,int(anomes))

        if funcoes.fun_planilha(planilha.name,id_municipio,anomes,mesano)==0:
            return HttpResponse("<h1>Nome da planilha não confere!</h1>")


        if len(lista_retorno)>0:
            return redirect ('app01:testeparam2', id_municipio=lista_retorno[0],anomes=lista_retorno[1])

        funcoes.funTruncateTables(id_municipio,anomes)

        if empresa=='SS':
            lista=importacao_excel_folha.importacao_excel_folha(planilha,ano,int(mes),anomes,id_municipio,current_user)
        elif empresa=='Aspec':
            lista=importacao_excel_folha_Aspec.importacao_excel_folha(planilha,ano,int(mes),anomes,id_municipio,current_user)
        elif empresa=='Layout':
            lista=importacao_excel_folha_layout.importacao_excel_folha(planilha,ano,int(mes),anomes,id_municipio,current_user)


        if len(lista)>0:
            return render(request, 'app01/dados_nao_cadastrados.html',{'lista':lista,'municipios':obj_municipio})


        return HttpResponseRedirect('/app01/importar_excel_folha')
    else:

        titulo = 'Importacao Planilha Folha'
        municipios = Municipio.objects.filter(empresa__in=['Layout','Aspec','SS']).order_by('municipio')
    return render(request, 'app01/importar_excel_folha.html',
            {
                'titulo_pagina': titulo,
                'municipios': municipios,
                'usuario':'fernando.paz'
            }
          )



def processarFolha(request):



    titulo_html = 'Processar Folha'

    mensagem=''
    municipios=Municipio.objects.filter(empresa__in=['SS','Layout','Aspec']).order_by('municipio')

    if (request.method == "POST"):

        current_user = 0  #request.user.iduser
        id_municipio=int(request.POST['municipio'])
        ano=request.POST['ano']
        mes=request.POST['mes']
        tabela=request.POST['tabela']

        mesf = str(int(mes)+100)[-2:]
        anomes=int(ano+mesf)


        obj=Municipio.objects.filter(id_municipio=id_municipio).first()
        municipio=obj.municipio
        empresa=obj.empresa
        entidade=obj.entidade


        obj = Folhames.objects.filter(anomes=anomes,id_municipio=id_municipio).first()
        if obj is not None:
            mensagem='Esta folha já está processada!'
            return redirect('app01:mensagemDelFolha', id_municipio=id_municipio, anomes =  anomes, mensagem = mensagem)

        mes_ref = mesReferencia(mes)

        if tabela=='Secretaria':
            retorno = processamentoFolha.importarSecretaria(id_municipio,anomes,empresa)
        elif tabela=='Funcao':
            retorno = processamentoFolha.importarFuncao(id_municipio,anomes,empresa)
        elif tabela=='Evento':
            retorno = processamentoFolha.importarEventos(id_municipio,anomes,empresa)
        elif tabela=='Setor':
            retorno = processamentoFolha.importarSetores(id_municipio,anomes,empresa)
        elif tabela=='Vinculos':
            retorno = processamentoFolha.importarVinculos(id_municipio,anomes,empresa)
        elif tabela=='Servidor':
            retorno = processamentoFolha.importarServidores(id_municipio,anomes,empresa)
        elif tabela=='Folha':
            retorno = processamentoFolha.importarFolhaPasso1(id_municipio,anomes,empresa)
            retorno = processamentoFolha.importarFolhaPasso2(id_municipio,anomes,empresa)
        elif tabela == 'Geral':
            #retorno = processamentoFolha.importarSecretaria(id_municipio,anomes,empresa)
            if 1==1:
                retorno = processamentoFolha.importarServidores(id_municipio,anomes,empresa)
                retorno = processamentoFolha.importarFolhaPasso1(id_municipio,anomes,empresa)
                retorno = processamentoFolha.importarFolhaPasso3(id_municipio,anomes,empresa)

            else:
                return render(request, 'app01/planilhaErrada.html',
                        {
                            'titulo': 'Processamento da Folha',
                            'municipio':municipio,
                            'anomes':str(mes)+'/'+str(ano),
                            'mensagem':'Nao existe nenhum registro desse municipio e desse mes para ser processado!'

                        }
                    )

        return HttpResponseRedirect(reverse('app01:processarFolha'))


    return render(request, 'app01/processarFolha.html',
            {
                'titulo': titulo_html,
                'mensagem':mensagem,
                'municipios':municipios
            }
          )


def planilhaErrada(request):
    return render(request, 'app01/planilhaErrada.html')


def listSomaEventosExcel(request):

    opcao=''
    query1=None
    query2=None
    titulo='Listar Soma por Eventos'
    id_municipio=0
    anomes=''
    municipio=''
    referencia=''
    rs=0
    lista_eventos=[]
    lista2=[]

    total=0
    total_v=0
    total_d=0
    total_r=0
    qT=0
    municipios = Municipio.objects.filter(empresa__in=['SS','Layout','Aspec']).order_by('municipio')

    if (request.method == "POST"):
        id_municipio=request.POST['municipio']
        ano=request.POST['ano']
        mes=request.POST['mes']
        obj=Municipio.objects.get(id_municipio=id_municipio)
        municipio=obj.municipio

        mesx=str(int(mes)+100)[-2:]
        anomes = int(ano+mesx)
        referencia = mes+"/"+ano

        query=Folhaevento.objects.filter(id_municipio=id_municipio,anomes=anomes).values('id_evento').annotate(soma=Sum('valor'))
        qtdeFuncionarios = Folhames.objects.filter(id_municipio=id_municipio,anomes=anomes).count()
        dicionarioEventos = listagens.criarDictIdEventosVantagens(id_municipio)

        lista1=[]
        lista2=[]
        for q in query:
            lista1.append(q['id_evento'])
            lista2.append(q['soma'])
        dicionarioValores = dict(zip(lista1,lista2))

        lista_eventos=[]
        lista_valores=[]

        for id_evento in dicionarioEventos:
            total_v+=dicionarioValores.get(id_evento,0)
            if dicionarioValores.get(id_evento,0)==0:
                continue

            total+=dicionarioValores[id_evento]

            lista_eventos.append(
                {
                    'evento':dicionarioEventos[id_evento],
                    'tipo':'(V)',
                    'valor':formatMilhar(dicionarioValores[id_evento])
                }
                )



        munic = Municipio.objects.get(id_municipio=id_municipio)
        municipio=munic.municipio
        empresa = munic.empresa
        entidade = munic.entidade

        label_arquivo='Resumo_'+entidade+'_'+mesPorExtenso(mes,1)+'_'+str(ano)+'.csv'


        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename='+label_arquivo

        cabecalho = ['Evento','Valor']
        writer = csv.writer(response, delimiter=';')
        response.write(u'\ufeff'.encode('utf8'))
        writer.writerow(cabecalho)

        row=1

        for ki in lista_eventos:
            writer.writerow([ki['evento'],ki['valor']])
            row+=1

        ci="B2"
        cf='B'+str(row)
        formula="=soma("+ci+":"+cf+")"

        writer.writerow(['T o t a l',formula])

        return response
    else:

        return render(request, 'app01/listSomaPorEventosExcel.html',
                {
                    'titulo': titulo,
                    'eventos':lista_eventos,
                    'municipios':municipios,
                    'id_municipio':id_municipio,
                    'anomes':anomes,
                    'municipio':municipio,
                    'referencia':referencia,
                    'qtde_funcionario':qT,
                    'total_v':total_v,
                    'total_d':total_d,
                    'total_r':total_r,
                    'qT':qT

                }
              )

def listSomaEventos(request):

    opcao=''
    query1=None
    query2=None
    titulo='Listar Soma por Eventos'
    id_municipio=0
    anomes=''
    municipio=''
    referencia=''
    rs=0
    lista_eventos=[]
    lista2=[]
    lista=[]

    total=0
    total_v=0
    total_d=0
    total_r=0
    qT=0
    municipios = Municipio.objects.filter(empresa__in=['SS','Layout','Aspec']).order_by('municipio')

    if (request.method == "POST"):
        id_municipio=request.POST['municipio']
        ano=request.POST['ano']
        mes=request.POST['mes']
        modo_form=request.POST.getlist('modo')
        obj=Municipio.objects.get(id_municipio=id_municipio)
        municipio=obj.municipio

        mesx=str(int(mes)+100)[-2:]
        anomes = int(ano+mesx)
        referencia = mesx+"/"+ano

        query=Folhaevento.objects.filter(id_municipio=id_municipio,anomes=anomes).values('id_evento').annotate(soma=Sum('valor'))
        qtdeFuncionarios = Folhames.objects.filter(id_municipio=id_municipio,anomes=anomes).count()
        dicionarioEventos = listagens.criarDictIdEventosVantagens(id_municipio)

        lista1=[]
        lista2=[]
        for q in query:
            lista1.append(q['id_evento'])
            lista2.append(q['soma'])
        dicionarioValores = dict(zip(lista1,lista2))

        lista_eventos=[]
        lista_valores=[]

        for id_evento in dicionarioEventos:
            total_v+=dicionarioValores.get(id_evento,0)
            if dicionarioValores.get(id_evento,0)==0:
                continue

            total+=dicionarioValores[id_evento]

            lista_eventos.append(
                {
                    'evento':dicionarioEventos[id_evento],
                    'tipo':'(V)',
                    'valor':formatMilhar(dicionarioValores[id_evento])
                }
                )

        munic = Municipio.objects.filter(id_municipio=id_municipio).first()
        municipio=munic.municipio
        empresa = munic.empresa
        entidade = munic.entidade
        label_arquivo='Resumo_'+entidade+'_'+mesPorExtenso(mes,1)+'_'+str(ano)+'.csv'

        row=1

        if 'Excel' in modo_form:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename='+label_arquivo

            cabecalho = ['Evento','Valor']
            writer = csv.writer(response, delimiter=';')
            response.write(u'\ufeff'.encode('utf8'))
            writer.writerow(cabecalho)
            row=1
            for ki in lista_eventos:
                writer.writerow([ki['evento'],ki['valor']])
                row+=1


            ci="B2"
            cf='B'+str(row)
            formula="=soma("+ci+":"+cf+")"


            writer.writerow(['T o t a l', formula])

            return response
        else:
            lista=lista_eventos
            total_v=formatMilhar(total_v)
            qT=qtdeFuncionarios

    return render(request, 'app01/listSomaPorEventos.html',
            {
                'titulo': titulo,
                'eventos':lista,
                'municipios':municipios,
                'id_municipio':id_municipio,
                'anomes':anomes,
                'municipio':municipio,
                'referencia':referencia,
                'qtde_funcionario':qT,
                'total_v':total_v,
                'total_d':total_d,
                'total_r':total_r,
                'qT':qT

            }
          )



def mesPorExtenso(mes,modelo):

    lista_mes=['','JANEIRO','FEVEREIRO','MARÇO','ABRIL','MAIO','JUNHO','JULHO','AGOSTO','SETEMBRO','OUTUBRO','NOVEMBRO','DEZEMBRO']
    if modelo==1:
        return lista_mes[int(mes)]
    elif modelo==2:
        return (lista_mes[int(mes)])[0:3]

def folhasProcessadas(request):

    titulo='Folhas Processadas'
    municipios=Municipio.objects.filter(empresa__in=['SS','Layout','Aspec']).order_by('municipio')
    sel_empresa=''


    municipio=''
    contador=0
    nome_da_empresa=''
    lista_folha=[]
    if request.method=='POST':
        id_municipio = int(request.POST['municipio'])
        sel_empresa = request.POST['empresa']

        if id_municipio==0:
            municipio='Todos os municípios'
            nome_da_empresa=''
        else:
            munic = Municipio.objects.filter(id_municipio=id_municipio)
            municipio = 'Municipio: '+ munic.municipio
            nome_da_empresa = munic.empresa

        lista_municipios=[k.id_municipio for k in Municipio.objects.filter(empresa=sel_empresa)]

        dictMunicipios=listagens.criarDictMunicipios()
        if id_municipio>0:
            query=Folhames.objects.filter(id_municipio=id_municipio).values('id_municipio','anomes','data_criacao').annotate(qtde=Count("id_folha")).order_by('id_municipio','anomes')
        elif sel_empresa!='Todas':
            query=Folhames.objects.filter(id_municipio__in=lista_municipios).values('id_municipio','anomes','data_criacao').annotate(qtde=Count("id_folha")).order_by('id_municipio','anomes')
        else:
            query=Folhames.objects.values('id_municipio','anomes','data_criacao').annotate(qtde=Count("id_folha")).order_by('id_municipio','anomes')

        lista1=[]
        lista2=[]
        lista3=[]
        for kk in query:
            lista1.append(str(kk['id_municipio'])+':'+str(kk['anomes']))
            lista2.append(kk['qtde'])
            lista3.append(kk['data_criacao'])

        dicionario=dict(zip(lista1,lista2))
        dicionario2=dict(zip(lista1,lista3))

        if id_municipio>0:
            resumo=Folhaevento.objects.filter(id_municipio=id_municipio).values('id_municipio','anomes').annotate(valor_total=Sum("valor")).order_by('id_municipio','anomes')
        elif sel_empresa!='Todas':
            resumo=Folhaevento.objects.filter(id_municipio__in=lista_municipios).values('id_municipio','anomes').annotate(valor_total=Sum("valor")).order_by('id_municipio','anomes')
        else:
            resumo=Folhaevento.objects.values('id_municipio','anomes').annotate(valor_total=Sum("valor")).order_by('id_municipio','anomes')

            #resumo
            #<QuerySet [{'id_municipio': 76, 'anomes': 202201, 'total': Decimal('2119447.46')}, {'id_municipio': 76, 'anomes': 202202, 'total': Decimal('2168637.07')}]>

        for res in resumo:
            contador+=1
            id_municipio=res['id_municipio']
            anomes=res['anomes']
            valor=res['valor_total']
            quantidade=dicionario[str(res['id_municipio'])+':'+str(res['anomes'])]
            data_criacao=dicionario2[str(res['id_municipio'])+':'+str(res['anomes'])]
            if data_criacao is not None:
                data_criacao=data_criacao.strftime("%d/%m/%Y %H:%M:%S")
            lista_folha.append(
                {
                    'item':contador,
                    'municipio':dictMunicipios[id_municipio],
                    'mesref':str(anomes)[-2:]+'/'+str(anomes)[0:4],
                    'quantidade':quantidade,
                    'data_criacao':data_criacao,
                    'valor':formatMilhar(valor)
                }
            )
        if nome_da_empresa!='':
            sel_empresa=nome_da_empresa

    return render(request, 'app01/folhasProcessadas.html',
        {
            'titulo': titulo,
            'municipios':municipios,
            'mensagem':'',
            'municipio':municipio,
            'lista_folha':lista_folha,
            'sel_empresa':'Empresa(s): '+sel_empresa
        }
    )


def testeparam(request,id_municipio):
    query = SemCadastro.objects.filter(id_municipio=id_municipio).order_by('tabela')
    munic=Municipio.objects.filter(id_municipio=id_municipio).first()
    municipio=munic.municipio
    return render (request, 'app01/listar_sem_cadastro.html', {'municipio':municipio, 'query':query})


def testeparam2(request,id_municipio,anomes):
    #query = SemCadastro.objects.filter(id_municipio=id_municipio).order_by('tabela')
    munic=Municipio.objects.filter(id_municipio=id_municipio).first()
    municipio=munic.municipio
    sanomes = str(anomes)
    mesEano=sanomes[-2:]+'/'+sanomes[0:4]
    return render (request, 'app01/mensagem_folha_ja_processada_01.html', {'municipio':municipio, 'mesEano':mesEano})




def deletar_folha(request):

    titulo='Deletar Folha Processada'
    municipios=Municipio.objects.filter(empresa__in=['SS','Layout','Aspec']).order_by('municipio')
    sel_empresa=''

    municipio=''
    contador=0
    nome_da_empresa=''
    lista_folha=None
    if request.method=='POST':
        id_municipio = int(request.POST['municipio'])
        ano = request.POST['ano']
        mes = request.POST['mes']
        anomes = int(str(ano)+str(int(mes)+100)[1:])
        sanomes=str(anomes)
        mesEano=sanomes[-2:]+'/'+sanomes[0:4]

        Folhames.objects.filter(id_municipio=id_municipio,anomes=anomes).delete()
        Folhaevento.objects.filter(id_municipio=id_municipio,anomes=anomes).delete()
        qtde1 = Folhames.objects.filter(id_municipio=id_municipio,anomes=anomes).count()
        qtde2 = Folhaevento.objects.filter(id_municipio=id_municipio,anomes=anomes).count()
        if qtde1+qtde2>0:
               mensagem = 'Erro ao tentar deletar a folha! Contate o programador'
        else:
              mensagem = 'Folha excluida com sucesso!'

        return redirect('app01:mensagemDelFolha', id_municipio=id_municipio, anomes =  anomes, mensagem = mensagem)


    return render (request, 'app01/deletar_folha.html', {'municipios': municipios,'titulo': titulo})




def mensagem_deletar_folha(request,id_municipio,anomes,mensagem):
    #query = SemCadastro.objects.filter(id_municipio=id_municipio).order_by('tabela')
    munic=Municipio.objects.filter(id_municipio=id_municipio).first()
    municipio=munic.municipio
    mesEano=str(anomes)
    mesEano=mesEano[-2:]+'/'+mesEano[0:4]
    return render (request, 'app01/mensagem_folha_ja_processada_02.html', {'municipio':municipio, 'mesEano':mesEano, 'mensagem':mensagem})


def atualizar_tabela_salario(request):

    titulo='Atualizar Salario'
    municipios=Municipio.objects.filter(empresa__in=['SS','Layout','Aspec']).order_by('municipio')
    sel_empresa=''

    municipio=''
    contador=0
    nome_da_empresa=''
    lista_folha=None
    if request.method=='POST':
        id_municipio = int(request.POST['municipio'])
        planilha=request.FILES['filename']
        ano = request.POST['ano']
        mes = request.POST['mes']
        anomes = int(str(ano)+str(int(mes)+100)[1:])
        sanomes=str(anomes)
        mesEano=sanomes[-2:]+'/'+sanomes[0:4]
        mensagem='Atualização realizada com sucesso!'

        retorno = tabela_salario.tabela_salario(planilha,ano,mes,anomes,id_municipio)


        return redirect('app01:mensagemDelFolha', id_municipio=id_municipio, anomes =  anomes, mensagem = mensagem)


    return render (request, 'app01/atualizar_salario.html', {'municipios': municipios,'titulo': titulo})



def ler_excel(request):

    current_user=166
    if (request.method == "POST" and request.FILES['filename']):

        id_municipio=request.POST['municipio']
        tabela=request.POST['tabela'] 
        planilha=request.FILES['filename']


        retorno=importacao_excel.import_excel_teste(planilha,int(id_municipio),tabela)
        if int(retorno)>0:
            return redirect ('app01:testeparam', id_municipio=int(retorno))

        return HttpResponseRedirect('/app01/lerExcel')
    else:

        titulo = 'Importacao Planilha Folha'
        municipios = Municipio.objects.filter(empresa__in=['Layout','Aspec','SS']).order_by('municipio')
    return render(request, 'app01/ler_excel.html',
            {
                'titulo_pagina': titulo,
                'municipios': municipios,
                'usuario':'fernando.paz'
            }
          )



def gravarFuncao(request):

    #sessao(request)
    current_user=166
    if (request.method == "POST" and request.FILES['filename']):

        id_municipio=request.POST['municipio']
        #ano=request.POST['ano']
        #mes=request.POST['mes']
        planilha=request.FILES['filename']
        #anomes=str(ano)+ str(int(mes)+100)[1:]


        retorno=importacao_excel.import_excel_gravarFuncao_passo1(planilha,int(id_municipio))
        '''
        if retorno is None:
            retorno=38
        if retorno>0:            
            return redirect ('app01:testeparam', id_municipio=int(retorno))
        '''

        retorno=importacao_excel.import_excel_gravarFuncao_passo2(planilha,int(id_municipio))
        if retorno is None:
            retorno=38
        if retorno>0:
            return redirect ('app01:testeparam', id_municipio=int(retorno))

        return HttpResponseRedirect('/app01/listarModels_01')
    else:

        titulo = 'Importacao Planilha Folha'
        municipios = Municipio.objects.filter(empresa__in=['Layout','Aspec','SS']).order_by('municipio')
    return render(request, 'app01/ler_excel.html',
            {
                'titulo_pagina': titulo,
                'municipios': municipios,
                'usuario':'fernando.paz'
            }
          )


def fun_valorch(valorch):
    if valorch =='':
        return 0
    valor = valorch.replace('R$','')
    valor = valor.replace('.','')
    valor = valor.replace(',','.')
    return float(valor)

def fun_ch(valorch):
    if valorch=='':
        return 0
    return int(valorch)


def cargahoraria_ajax_01(request):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if is_ajax:
        id_municipio = request.GET.get('opcao')

        funcoes = Funcao.objects.filter(id_municipio=id_municipio)
        data = []

        for item in funcoes:
            key = item.id_funcao
            value = item.funcao
            data.append({'key':key,'value':value})
        data = json.dumps(data)
        return HttpResponse(data, content_type='application/json')
    else:
        raise Http404


def listarModels_01(request):
    municipios=Municipio.objects.filter(empresa__in=['SS','Layout','Aspec'])
    if (request.method == "POST"):
        #request.session['pasta']=request.POST['pasta']
        id_municipio=request.POST['municipio']
        tabela=request.POST['tabela']
        if 'ampliacao' in request.POST:
            ampliacao='A'
        else:
            ampliacao='X'

        if 'suporte' in request.POST:
            suporte='S'
        else:
            suporte='X'

        if 'cancelado' in request.POST:
            cancelado='C'
        else:
            cancelado='X'


        if 'carater_salario' in request.POST:
            carater_salario='S'
        else:
            carater_salario='X'

        if 'pos_graduacao' in request.POST:
            pos_graduacao='S'
        else:
            pos_graduacao='X'


        if 'apelido' in request.POST:
            apelido='a'
        else:
            apelido=''

        if 'professor' in request.POST:
            professor='P'
        else:
            professor='X'

        if 'fundeb' in request.POST:
            fundeb='F'
        else:
            fundeb='X'

        complemento=cancelado+ampliacao+suporte+apelido+professor+fundeb

        return redirect('app01:listarModels_02', id_municipio=id_municipio,tabela=tabela,complemento=complemento,carater_salario=carater_salario,pos_graduacao=pos_graduacao)
    titulo='Listagem Tabelas'
    return render(request, 'app01/listarModels_01.html', {'municipios':municipios,'titulo':titulo})



def listarModels_02(request,id_municipio,tabela,complemento,carater_salario,pos_graduacao):
    obj=Municipio.objects.get(id_municipio=id_municipio)
    municipio=obj.municipio
    (lista, filtro, obs) = funcoes_listar_models.listar_models_02(id_municipio,tabela,complemento,carater_salario,pos_graduacao)
    return render(request, 'app01/listarModels_02.html',{'municipio':municipio,'lista':lista,'tabela':tabela,'filtro':filtro,'obs':obs})    


def consultarSuporte(request):
    #sessao(request)
    if (request.method == "POST"):

        id_municipio=request.POST['municipio']
        ano=request.POST['ano']
        mes=request.POST['mes']
        funcao=request.POST['funcao']

        mesx=str(int(mes)+100)[-2:]

        anomes=str(ano)+mesx
        return redirect ('app01:imprimeAnalise', id_municipio=int(id_municipio),anomes=int(anomes),codFuncao=funcao)
    else:

        titulo = 'Consultar Suporte'
        municipios = Municipio.objects.filter(empresa__in=['Layout','Aspec','SS']).order_by('municipio')
    return render(request, 'app01/consultarSuporte.html',
            {
                'titulo_pagina': titulo,
                'municipios': municipios,
            }
          )    


def listarModels_03B(request,id_municipio,tabela):
    obj=Municipio.objects.get(id_municipio=id_municipio)
    municipio=obj.entidade
    lista=[]

    if tabela=='Eventos':
        objeto='Evento'
        eventos = Evento.objects.filter(id_municipio=id_municipio,cancelado='N').order_by('id_evento')
        for obj in eventos:
            if obj.ampliacao_ch=='S':
                ampliacao='A'
            else:
                ampliacao=''
            if obj.suporte=='S':
                suporte='S'
            else:
                suporte=''
            complemento=ampliacao+suporte

            lista.append(
                {
                'descricao1':obj.evento,
                'descricao2':obj.evento_out,
                'complemento':complemento
                }
            ) 
    elif tabela=='Funcoes':
        objeto='Funcao'
        funcoes = Funcao.objects.filter(id_municipio=id_municipio,cancelado='N').order_by('id_funcao')
        for obj in funcoes:
            professor=''
            if obj.professor=='S':
                professor='P'
            lista.append(
                {
                'descricao1':obj.funcao,
                'descricao2':obj.funcao_out,
                'complemento':professor
                }
            ) 

    elif tabela=='Setores':
        objeto='Setor'
        setores = Setor.objects.filter(id_municipio=id_municipio).order_by('setor')
        for obj in setores:

            fundeb=''
            if obj.fundeb=='S':
                fundeb='F'
            lista.append(
                {
                'descricao1':obj.setor,
                'descricao2':obj.setor_out,
                'complemento':fundeb,
                }
            ) 
    elif tabela=='Secretarias':
        objeto='Secretaria'
        secretarias = Secretaria.objects.filter(id_municipio=id_municipio).order_by('secretaria')
        for obj in secretarias:
            lista.append(
                {
                'descricao1':obj.secretaria,
                'descricao2':obj.secretaria_out,
                'complemento':''
                }
            ) 
    elif tabela=='Vinculos':
        objeto='Vinculo'
        vinculos = Vinculo.objects.filter(id_municipio=id_municipio).order_by('vinculo')
        for obj in vinculos:

            lista.append(
                {
                'descricao1':obj.vinculo,
                'descricao2':'',
                'complemento':obj.grupo
                }
            ) 

    return render(request, 'app01/listarModels_03B.html',{'municipio':municipio,'lista':lista,'tabela':tabela,'tabela':objeto})    

'''
def abrirExcel(request):
    workbook = openpyxl.load_workbook('/home/fernandopaz/documentos/funcoesCaninde.xlsx')
    response =  HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] =  'attachment; filename=funcoesCaninde.xlsx'
    workbook.save(response)
    return response
'''



def abrirExcel(request):
    if (request.method == "POST"):
        id_municipio=request.POST['municipio']
        ano=request.POST['ano']
        mes=request.POST['mes']
        anomes = str(ano)+str(int(mes)+100)[-2:]
        opcao = request.POST["radioFilter"]
        op_opcao=''
        if opcao=='regente':
            op_opcao='regente'
        elif opcao=='suporte':
            op_opcao='suporte'
        elif opcao=='matriz':
            op_matriz='matriz'

        doc=Documento.objects.filter(id_municipio=id_municipio,anomes=anomes,tipo=op_opcao).first()
        filename='/home/fernandopaz/documentos/'+doc.nome_do_arquivo
        workbook = openpyxl.load_workbook(filename)
        filename=doc.nome_do_arquivo

        response =  HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] =  'attachment; filename='+filename
        workbook.save(response)
        return response
    else:
        municipios=Municipio.objects.filter(empresa__in=['Aspec','Layout','SS']).order_by('municipio')
        return render(request, 'app01/abrirExcel.html',{'municipios':municipios,'titulo':'Abrir Arquivo Excel'})    




def imprimeAnalise(request,id_municipio,anomes,codFuncao):
    lista = funcoes_listar_models.imprime_analise(id_municipio,anomes,codFuncao)
    obj=Municipio.objects.get(id_municipio=id_municipio)
    municipio=obj.entidade

    label_arquivo='Analise_'+municipio+'_'+str(anomes)+'.csv'

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename='+label_arquivo
    cabecalho=['Funcao','Cod.Servidor','Suporte','Ampliacao_CH','Carga_Horaria_Origem','Carga_Horaria','Salario_100H','Venc.Base','Salario','Vinculo','Participacao', 'H * K','Diferenca' ]
    writer = csv.writer(response, delimiter=';')
    response.write(u'\ufeff'.encode('utf8'))
    if len(lista)>0: 
        writer.writerow(cabecalho)
        for k in lista:
            writer.writerow(k)
    if len(lista)>0:
        return response
    return HttpResponse ("<h1>Nenhum registro localizado!</h1>")

def listarParticipacaoAmpliacao(request):
    if (request.method == "POST"):
        id_municipio=request.POST['municipio']
        ano=request.POST['ano']
        mes=request.POST['mes']

        participacao=False
        ampliacao=False
        carga_horaria=False
        if request.POST["radioFilter"]=='1':
            participacao=True
        elif request.POST["radioFilter"]=='2':
            ampliacao=True
        elif request.POST["radioFilter"]=='3':
            carga_horaria=True

        obj=Municipio.objects.get(id_municipio=id_municipio)
        label_arquivo='Participacoes_'+obj.entidade+'_'+mesPorExtenso(mes,1)+'_'+str(ano)+'.csv'
        if participacao:
            lista =  funcoes_query.listagemParticipacao(id_municipio,ano,mes)
            label_arquivo='Participacoes_'+obj.entidade+'_'+mesPorExtenso(mes,1)+'_'+str(ano)+'.csv'
        elif ampliacao:
            lista =  funcoes_query.listagemAmpliacao(id_municipio,ano,mes)
            label_arquivo='Ampliacoes_'+obj.entidade+'_'+mesPorExtenso(mes,1)+'_'+str(ano)+'.csv'
        elif carga_horaria:
            lista =  funcoes_query.listagemCargaHoraria(id_municipio,ano,mes)
            label_arquivo='CargaHoraria_'+obj.entidade+'_'+mesPorExtenso(mes,1)+'_'+str(ano)+'.csv'


        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename='+label_arquivo
        if participacao:
            cabecalho = ['Cod-servidor','Funcao','Cod-funcao','Suporte','Vinculo','Carga-Horaria-Origem','Carga-Horaria','Salario','Salario-100H','Vencimento-Base','Participacao']
        elif ampliacao:
            cabecalho = ['Cod-servidor','Funcao','Cod-funcao','Suporte','Vinculo','Carga-Horaria-Origem','Carga-Horaria','Salario','Salario-100H','Vencimento-Base','Participacao','Valor-Ampliacao','Participacao-Ampliacao']
        elif carga_horaria:
            cabecalho = ['Carga-Horaria-Origem','Carga-Horaria','Funcao','Cod-funcao','Salario-100H']


        writer = csv.writer(response, delimiter=';')
        response.write(u'\ufeff'.encode('utf8'))
        writer.writerow(cabecalho)
        if lista:
            for qq in lista:
                if participacao:
                    writer.writerow(
                        [
                        qq['cod_servidor'],
                        qq['funcao'],
                        qq['id_funcao'],
                        qq['suporte'],
                        qq['vinculo'],
                        qq['carga_horaria_origem'],
                        qq['carga_horaria'],
                        funcoes.formatMilhar(qq['salario']),
                        funcoes.formatMilhar(qq['salario_100H']),
                        funcoes.formatMilhar(qq['vencimento_base']),
                        funcoes.formatMilhar(qq['participacao'])
                        ]
                    )
                elif ampliacao:
                    writer.writerow(
                        [
                        qq['cod_servidor'],
                        qq['funcao'],
                        qq['id_funcao'],
                        qq['suporte'],
                        qq['vinculo'],
                        qq['carga_horaria_origem'],
                        qq['carga_horaria'],
                        funcoes.formatMilhar(qq['salario']),
                        funcoes.formatMilhar(qq['salario_100H']),
                        funcoes.formatMilhar(qq['vencimento_base']),
                        funcoes.formatMilhar(qq['participacao']),
                        funcoes.formatMilhar(qq['valor']),
                        funcoes.formatMilhar(qq['ampliacao'])
                        ]
                    )
                elif carga_horaria:
                    writer.writerow(
                        [
                        qq['carga_horaria_origem'],
                        qq['carga_horaria'],
                        qq['funcao'],
                        qq['id_funcao'],
                        funcoes.formatMilhar(qq['salario_100H'])
                        ]
                    )


        return response
    municipios = Municipio.objects.filter(empresa__in=['SS','Layout','Aspec']).order_by('municipio')
    titulo='Listar Participação/Ampliação'
    return render(request, 'app01/listarParticipacaoAmpliacao.html', {'municipios':municipios,'titulo':titulo})


def abrirRegenteSuporteMatriz(request,id_arquivo):
    doc=Documento.objects.get(id_documento=id_arquivo)
    filename='/home/civitas/documentos/'+doc.nome_do_arquivo
    workbook = openpyxl.load_workbook(filename)
    response =  HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] =  'attachment; filename='+filename
    workbook.save(response)
    return response


def listarExcel(request):
    if (request.method == "POST"):
        arquivo=request.POST['arquivo']
        doc=Documento.objects.get(id_documento=arquivo)
        #filename='/home/fernandopaz/documentos/'+doc.nome_do_arquivo
        filename='/home/fernandopaz/documentos/'+doc.nome_do_arquivo
        #filename='funcoesCaninde.xlsx'
        #filename='regente_'+str(anomes)+'.xlsx'
        #workbook = openpyxl.load_workbook('funcoesCaninde.xlsx')
        workbook = openpyxl.load_workbook(filename)
        response =  HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        #response['Content-Disposition'] =  'attachment; filename=funcoesCaninde.xlsx'
        response['Content-Disposition'] =  'attachment; filename='+filename
        workbook.save(response)
        return response

    else:
        municipios=Municipio.objects.filter(empresa__in=['Aspec','Layout','SS']).order_by('municipio')
        docs = Documento.objects.filter(id_municipio=0)
        lista=[]
        for k in docs:
            lista.append(k.nome_do_arquivo)
        return render(request, 'app01/listarExcel.html',{'municipios':municipios,'arquivos':lista,'titulo':'Abrir Arquivo Excel'})    


def listarExcel_ajax_01(request):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if is_ajax:
        id_municipio = request.GET.get('param1')
        ano = request.GET.get('param2')
        ano=int(ano)
        id_municipio=int(id_municipio)
        print ('novo: ' +str(id_municipio)+'-'+str(ano))
        if ano>2000:
            docs = Documento.objects.filter(id_municipio=id_municipio,ano=ano)
        else:
            docs = Documento.objects.filter(id_municipio=id_municipio)            
        data = []

        for item in docs:
            key = item.id_documento
            value = item.nome_do_arquivo
            data.append({'key':key,'value':value})
        data = json.dumps(data)
        return HttpResponse(data, content_type='application/json')
    else:
        raise Http404


