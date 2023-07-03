from django.shortcuts import render,redirect
from django.http import HttpResponse,HttpResponseRedirect,Http404
from . import listagens,funcoes
from .. import dicionarios,funcoes as f_funcoes
from decimal import *

from ..models import Municipio,Folhames,Folhaevento,Evento,Vinculo,Funcao,Setor,XEvento,Documento
from django.db.models import Count,Sum,Min,Avg,Max,Q
from accounts.models import User
import csv
import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Border, Side
from openpyxl.styles import PatternFill, Border, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting import Rule
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter


from django.core.files import File
from django.db import connection
import unicodedata


def sessao(request):
    if not request.session.get('username'):
        request.session['username'] = request.user.username
    return


def imprimirFolha(request):
    titulo = 'Impressao do Excel'
    municipios=Municipio.objects.filter(empresa__in=['SS','Layout','Aspec']).order_by('municipio')


    if request.method=='POST':
        id_municipio = request.POST['municipio']
        ano=request.POST['ano']
        mes=request.POST['mes']

        mesx=str(int(mes)+100)[-2:]
        anomes=int(ano+mesx)
        lista=[]

        query=Folhaevento.objects.filter(id_municipio=id_municipio,anomes=anomes).values('id_evento').annotate(soma=Sum('valor'))
        lista_id_evento = [e['id_evento'] for e in query]
        eventos = [ev.evento_out for ev in Evento.objects.filter(id_municipio=id_municipio,id_evento__in=lista_id_evento).order_by('evento')]

        ls_municipio = funcoes.entidade(id_municipio)
        if len(ls_municipio)>0:
            municipio=ls_municipio[0]
            empresa = ls_municipio[1]
            entidade = ls_municipio[2]
        label_arquivo=entidade+'_'+funcoes.mesPorExtenso(mes,1)+'_'+str(ano)+'.csv'


        colunas=listagens.colunasValores()

        ultima_coluna = colunas[len(eventos)+8]

        query=Folhames.objects.filter(id_municipio=id_municipio,anomes=anomes).values('cod_servidor','id_secretaria','id_setor','id_funcao','id_vinculo','previdencia','carga_horaria','num_dias').order_by('cod_servidor')

        dicNomeDoServidor=listagens.criarDictNomeServidor(id_municipio)
        dicNomeDaSecretaria=listagens.criarDictIdSecretarias(id_municipio)
        dicNomeDoSetor=listagens.criarDictIdSetores(id_municipio)
        dicNomeDaFuncao=listagens.criarDictIdFuncoes(id_municipio)
        dicNomeDoVinculo=listagens.criarDictIdVinculos(id_municipio)
        contador=2

        dictEventos=funcoes.eventosMesDoServidor(id_municipio,anomes)
        lista=[]

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename='+label_arquivo

        cabecalho = funcoes.cabecalhoFolha(id_municipio,eventos)
        writer = csv.writer(response, delimiter=';')
        response.write(u'\ufeff'.encode('utf8'))
        writer.writerow(cabecalho)

        for qy in query:
            cod_servidor=qy['cod_servidor']


            lista.append(dicNomeDaSecretaria[qy['id_secretaria']])
            lista.append(dicNomeDoSetor.get(qy['id_setor'],'ERRO: '+str(qy['id_setor'])))
            lista.append(cod_servidor)
            lista.append(dicNomeDoServidor[cod_servidor]['nome'])
            lista.append(dicNomeDaFuncao.get(qy['id_funcao'],'ERRO: '+str(qy['id_funcao'])))
            lista.append(dicNomeDoVinculo[qy['id_vinculo']])
            lista.append(dicNomeDoServidor[cod_servidor]['data'])
            lista.append('CH '+str(qy['carga_horaria']))
            lista.append(qy['num_dias'])

            soma = 0

            eventosDoServidor=dictEventos.get(cod_servidor)
            if eventosDoServidor is None:
                lista=[]
                continue

            #[{'evento': 'ADC PTEMPSERV', 'valor': Decimal('381.28')}, {'evento': 'SALARIO BASE', 'valor': Decimal('3177.33')}]
            dicionario=funcoes.montarDicionarioEventoDoServidor(eventosDoServidor)
            #{'ADC PTEMPSERV': Decimal('381.28'), 'SALARIO BASE': Decimal('3177.33')}
            listaEventosDoServidor=funcoes.montaListaEventoDoServidor(eventosDoServidor)

            #['ADC PTEMPSERV', 'SALARIO BASE', 'SALARIO FAMILIA']
            for qq in range(len(eventos)):
                if eventos[qq] in listaEventosDoServidor:
                    valor=dicionario[eventos[qq]]
                    valor_str=str(valor)
                    valor_str = valor_str.replace('.',',')
                else:
                    valor_str='0'
                lista.append(valor_str)


            ci="J"+str(contador)

            cf=ultima_coluna+str(contador)
            formula="=soma("+ci+":"+cf+")"

            contador+=1
            lista.append(formula)

            writer.writerow(lista)
            lista=[]
        return response

    return render(request, 'app01/imprimirFolhaExcel.html',
        {
            'titulo': titulo,
            'municipios':municipios,
            'mensagem':''

        }
    )



def fun_nova_listaEventos(id_municipio,anomes):
    cursor = connection.cursor()
    cursor.execute("Select fm.suporte,case when fm.vinculo in ('T','C') then 'T' \
    else 'E' end as vinculo,fe.id_evento,sum(valor) as soma \
    from folhames fm,folhaeventos fe \
    where fm.id_municipio=fe.id_municipio and fm.anomes=fe.anomes and fm.cod_servidor=fe.cod_servidor\
    and fm.id_municipio=%s and fm.anomes=%s \
    and fm.fundeb='S' and fm.vinculo in ('E','T','C')\
    group by fm.suporte,fm.vinculo,fe.id_evento",[id_municipio,anomes])
    query = dictfetchall(cursor)
    return query



def dictfetchall(cursor):
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]


def lista_Soma_Eventos(id_municipio,anomes,tabela):
    lista_regente_efetivo=[]
    lista_regente_temporario=[]
    lista_suporte_efetivo=[]
    lista_suporte_temporario=[]
    dicionario=listagens.criarDictEventos(id_municipio)
    #lista_eventos_carater_salario=[ev.id_evento for ev in Evento.objects.filter(id_municipio=id_municipio,carater_salario='S')]
    lista_eventos_pos=[ev.id_evento for ev in Evento.objects.filter(id_municipio=id_municipio,posgraduacao='S')]
    lista_eventos_suporte=[ev.id_evento for ev in Evento.objects.filter(id_municipio=id_municipio,suporte='S')]
    lista_eventos_salBase=[ev.id_evento for ev in Evento.objects.filter(id_municipio=id_municipio,evento__in=['VENCIMENTO BASE','SALARIO BASE'])]
    lista_eventos_ampliacao=[ev.id_evento for ev in Evento.objects.filter(id_municipio=id_municipio,ampliacao_ch='S')]
    lista_eventos_caraterSalario=[ev.id_evento for ev in Evento.objects.filter(id_municipio=id_municipio,carater_salario='S')]

    re_soma_aux=Decimal(0)
    rt_soma_aux=Decimal(0)
    se_soma_aux=Decimal(0)
    st_soma_aux=Decimal(0)

    lista_soma=fun_nova_listaEventos(id_municipio,anomes) 
    for k in lista_soma:
        ordenacao=8
        tipo_evento='adicional'
        id_evento=k['id_evento']
        desc_evento=dicionario.get(id_evento,'')
        vecnto_base_e_ampliacao=0
        soma=k['soma']


        if id_evento in lista_eventos_salBase or id_evento in lista_eventos_ampliacao:
            if k['suporte']=='S':
                if k['vinculo']=='E':
                    se_soma_aux+=k['soma']
                    vecnto_base_e_ampliacao=1
                else:
                    st_soma_aux+=k['soma']
                    vecnto_base_e_ampliacao=1
            else:
                if k['vinculo']=='E':
                    re_soma_aux+=k['soma']
                    vecnto_base_e_ampliacao=1
                else:
                    rt_soma_aux+=k['soma']
                    vecnto_base_e_ampliacao=1

        elif id_evento in lista_eventos_suporte:
            tipo_evento='suporte'
            ordenacao=4
        elif id_evento in lista_eventos_caraterSalario:
            tipo_evento='carater-salario'
            ordenacao=2
        else:
            tipo_evento='adicional'
            ordenacao=8

        if vecnto_base_e_ampliacao==0: 
            if k['suporte']=='S':
                if k['vinculo']=='E':
                    lista_suporte_efetivo.append([ordenacao,tipo_evento,desc_evento,soma])
                else:
                    lista_suporte_temporario.append([ordenacao,tipo_evento,desc_evento,soma])
            else:
                if k['vinculo']=='E':
                    lista_regente_efetivo.append([ordenacao,tipo_evento,desc_evento,soma])
                else:
                    lista_regente_temporario.append([ordenacao,tipo_evento,desc_evento,soma])
    if re_soma_aux>0:
        lista_regente_efetivo.append([1,'salario-base','VENCIMENTO BASE',re_soma_aux])
    if rt_soma_aux>0:
        lista_regente_temporario.append([1,'salario-base','VENCIMENTO BASE',rt_soma_aux])

    if se_soma_aux>0:
        lista_suporte_efetivo.append([1,'salario-base','VENCIMENTO BASE',se_soma_aux])
    if st_soma_aux>0:
        lista_suporte_temporario.append([1,'salario-base','VENCIMENTO BASE',st_soma_aux])



    lista_regente_efetivo.sort()
    lista_regente_temporario.sort()
    lista_suporte_efetivo.sort()
    lista_suporte_temporario.sort()
    if tabela=='REGENTE':
        return [lista_regente_efetivo,lista_regente_temporario]
    elif tabela=='SUPORTE':
        return [lista_suporte_efetivo,lista_suporte_temporario]
    else:
        return [lista_regente_efetivo,lista_suporte_efetivo,lista_regente_temporario,lista_suporte_temporario]


def fun_lista_participacao(id_municipio,anomes):
    cursor = connection.cursor()

    cursor.execute("Select fm.posgraduacao,fu.id_funcao,CAST(salario_100H AS DECIMAL(0)) as salario_100H_int,carga_horaria,fm.suporte, \
    fm.grupo, \
    MIN(salario_100H) AS salario_100H,sum(participacao) as somaParticipacao,sum(vencimento_base) as somaVencimento \
    from folhames fm,funcoes fu where fm.id_municipio=%s and anomes=%s \
    and fm.fundeb='S' and fm.vinculo in ('E','T','C') \
    and fu.id_funcao=fm.id_funcao and fu.id_municipio=fm.id_municipio\
    and salario>0 group by fm.posgraduacao,fu.id_funcao,salario_100H_int,carga_horaria,fm.suporte,fm.grupo order by fm.id_funcao,salario_100H_int",[id_municipio,anomes])
    query = dictfetchall(cursor)
    return query


def fun_dictAmpliacaoCh_01(id_municipio,anomes,tupla_eventos_ampliacao):
    cursor = connection.cursor()
    cursor.execute("select fm.posgraduacao,fm.suporte, \
    fm.grupo, \
    fm.id_funcao,CAST(fm.salario_100H AS DECIMAL(0)) as salario_100H_int,sum(fe.valor/fm.salario_100H) as participacao \
    from folhaeventos fe,folhames fm \
    where fe.cod_servidor=fm.cod_servidor and fe.id_municipio=fm.id_municipio and fe.anomes=fm.anomes \
    and fe.id_municipio=%s and fe.anomes=%s \
    and fm.vinculo in ('E','T','C') \
    and fm.fundeb='S' \
    and fe.id_evento in %s \
    and fm.salario_100H IS NOT NULL and fm.ampliacao_ch='S' \
    group by fm.posgraduacao,fm.suporte,fm.grupo,fm.id_funcao,CAST(fm.salario_100H AS DECIMAL(0)) order by fm.posgraduacao,fm.suporte",[id_municipio,anomes,tupla_eventos_ampliacao])
    query = dictfetchall(cursor)

    lisSalario=[]
    lisParticipacao=[]
    for qp in query:
        chave = qp['posgraduacao']+qp['suporte']+qp['grupo']+str(qp['id_funcao'])+str(qp['salario_100H_int'])
        valor = round(qp['participacao'],4)
        lisSalario.append(chave)
        lisParticipacao.append(valor)
    return dict(zip(lisSalario,lisParticipacao))



def fun_dictAmpliacaoCh_02(id_municipio,anomes,tupla_eventos_ampliacao):
    cursor = connection.cursor()
    cursor.execute("select fm.posgraduacao,fm.suporte, \
    fm.grupo, \
    fm.id_funcao,CAST(fe.valor AS DECIMAL(0)) as salario_100H_int,sum(fe.valor/fe.valor) as participacao \
    from folhaeventos fe,folhames fm \
    where fe.cod_servidor=fm.cod_servidor and fe.id_municipio=fm.id_municipio and fe.anomes=fm.anomes \
    and fe.id_municipio=%s and fe.anomes=%s \
    and fm.ampliacao_ch='S' \
    and fm.vinculo in ('E','T','C') \
    and fe.id_evento in %s \
    and fm.salario_100H IS NULL and fm.fundeb='S' \
    group by fm.posgraduacao,fm.suporte,fm.grupo,fm.id_funcao,CAST(fe.valor AS DECIMAL(0)) order by 1,2",[id_municipio,anomes,tupla_eventos_ampliacao])
    query = dictfetchall(cursor)

    lisSalario=[]
    lisParticipacao=[]
    for qp in query:
        chave = qp['posgraduacao']+qp['suporte']+qp['grupo']+str(qp['id_funcao'])+str(qp['salario_100H_int'])
        valor = round(qp['participacao'],4)
        lisSalario.append(chave)
        lisParticipacao.append(valor)
    return dict(zip(lisSalario,lisParticipacao))


def teste(request):
    id_municipio=15
    anomes=202211
    (ls1,ls2) = lista_SomaEventos(id_municipio,anomes)
    print ('somaEventos - detalhado')
    for k in ls1:
        print (k)
    print ('somaEventos - resumido')
    for k in ls2:
        print (k)
    return  HttpResponse("<h1>Encerrado</h1>")


def lista_SomaEventos_Geral(id_municipio,anomes):
    query = fun_nova_listaEventos(id_municipio,anomes)
    lista=[]
    listao=[]
    lsregefet=[]
    lsregefet_valor=[]

    lsregtemp=[]
    lsregtemp_valor=[]

    lssupefet=[]
    lssupefet_valor=[]

    lssuptemp=[]
    lssuptemp_valor=[]
    lista_vencimento_base=[k.id_evento for k in Evento.objects.filter(id_municipio=id_municipio,evento='VENCIMENTO BASE')]
    lista_carater_salario=[k.id_evento for k in Evento.objects.filter(id_municipio=id_municipio,carater_salario='S')]
    lista_suporte=[k.id_evento for k in Evento.objects.filter(id_municipio=id_municipio,suporte='S')]
    lista_eventos_ampliacao=[ev.id_evento for ev in Evento.objects.filter(id_municipio=id_municipio,ampliacao_ch='S')]
    tipo=''
    ordenacao=0
    lista_soma=[]
    for ky in query:
        tipo='adcional'
        ordenacao=9
        if ky['id_evento'] not in lista:
            lista.append(ky['id_evento'])
        if ky['suporte']=='N':
            if ky['vinculo']=='E':
                lsregefet.append(ky['id_evento'])
                lsregefet_valor.append(ky['soma'])
            else:
                lsregtemp.append(ky['id_evento'])
                lsregtemp_valor.append(ky['soma'])
        else:
            if ky['vinculo']=='E':
                lssupefet.append(ky['id_evento'])
                lssupefet_valor.append(ky['soma'])
            else:
                lssuptemp.append(ky['id_evento'])
                lssuptemp_valor.append(ky['soma'])


    dicRegEfet=dict(zip(lsregefet,lsregefet_valor))
    dicRegTemp=dict(zip(lsregtemp,lsregtemp_valor))
    dicSupEfet=dict(zip(lssupefet,lssupefet_valor))
    dicSupTemp=dict(zip(lssuptemp,lssuptemp_valor))
    dicEventos=listagens.criarDictEventos(id_municipio)


    for kv in lista:
        if kv in lista_vencimento_base or kv in lista_eventos_ampliacao:
            tipo='salario-base'
            ordenacao=1
        elif kv in lista_carater_salario:
            tipo='carater-salario'
            ordenacao=4
        elif kv in lista_suporte:
            tipo='suporte'
            ordenacao=7
        else:
            tipo='adicional'
            ordenacao=9
        descEvento=dicEventos.get(kv,'')

        e1=Decimal(dicRegEfet.get(kv,0))
        e2=Decimal(dicRegTemp.get(kv,0))
        e3=Decimal(dicSupEfet.get(kv,0))
        e4=Decimal(dicSupTemp.get(kv,0))
        listao.append([ordenacao,tipo,descEvento,kv,e1,e2,e3,e4])

    sb1=Decimal(0)
    sb2=Decimal(0)
    sb3=Decimal(0)
    sb4=Decimal(0)
    cs1=Decimal(0)
    cs2=Decimal(0)
    cs3=Decimal(0)
    cs4=Decimal(0)
    sp1=Decimal(0)
    sp2=Decimal(0)
    sp3=Decimal(0)
    sp4=Decimal(0)
    ad1=Decimal(0)
    ad2=Decimal(0)
    ad3=Decimal(0)
    ad4=Decimal(0)


    for k in listao:
        if k[1]=='salario-base':
            sb1+=k[4]
            sb2+=k[5]
            sb3+=k[6]
            sb4+=k[7]
        elif k[1]=='carater-salario':
            cs1+=k[4]
            cs2+=k[5]
            cs3+=k[6]
            cs4+=k[7]
        elif k[1]=='suporte':
            sp1+=k[4]
            sp2+=k[5]
            sp3+=k[6]
            sp4+=k[7]
        elif k[1]=='adicional':
            ad1+=k[4]
            ad2+=k[5]
            ad3+=k[6]
            ad4+=k[7]

    tsbr=Decimal(sb1+sb2)
    tsbs=Decimal(sb3+sb4)

    tcsr=Decimal(cs1+cs2)
    tcss=Decimal(cs3+cs4)

    tspr=Decimal(sp1+sp2)
    tsps=Decimal(sp3+sp4)

    tadr=Decimal(ad1+ad2)
    tads=Decimal(ad3+ad4)

    tot1=sb1+cs1+sp1+ad1
    tot2=sb2+cs2+sp2+ad2
    tot3=tsbr+tcsr+tspr+tadr
    tot4=sb3+cs3+sp3+ad3
    tot5=sb4+cs4+sp4+ad4
    tot6=tsbs+tcss+tsps+tads

    lista_soma.append([1,'Salario Base',sb1,sb2,tsbr,sb3,sb4,tsbs])
    lista_soma.append([2,'Carater Salario',cs1,cs2,tcsr,cs3,cs4,tcss])
    lista_soma.append([3,'Suporte',sp1,sp2,tspr,sp3,sp4,tsps])
    lista_soma.append([4,'Adicional',ad1,ad2,tadr,ad3,ad4,tads])
    lista_soma.append([9,'Soma',tot1,tot2,tot3,tot4,tot5,tot6])

    return [listao,lista_soma]


def fun_Auxiliar_tabelaMatriz(lista):
    v1=lista[4]+lista[5]*2+lista[6]+lista[10]+lista[11]*2+lista[12]
    v2=lista[7]+lista[8]*2+lista[9]+lista[13]+lista[14]*2+lista[15]
    c1=v1*lista[3]
    c2=v2*lista[3]

    return [v1,c1,v2,c2]


def coletaDados_novo(id_municipio,anomes):


    eventos_ampliacao = Evento.objects.filter(id_municipio=id_municipio,ampliacao_ch='S').count()

    if eventos_ampliacao>0:
        tupla_ampliacao = tuple([k.id_evento for k in Evento.objects.filter(id_municipio=id_municipio,ampliacao_ch='S')])
    else:
        tupla_ampliacao=None

    query = fun_lista_participacao(id_municipio,anomes)
    dictAmpliacao=0
    if tupla_ampliacao:
        dictAmpliacao_01=fun_dictAmpliacaoCh_01(id_municipio,anomes,tupla_ampliacao)
        dictAmpliacao_02=fun_dictAmpliacaoCh_02(id_municipio,anomes,tupla_ampliacao)
        dictAmpliacao=1

    dictFuncoes=listagens.criarDictIdFuncoes(id_municipio)
    lista_funcao_prof=[ev.id_funcao for ev in Funcao.objects.filter(id_municipio=id_municipio,professor='S')]

    listao=[]
    lista_computados=[]

    for qq in query:

        re_ampliacao=Decimal(0)
        rt_ampliacao=Decimal(0)
        se_ampliacao=Decimal(0)
        st_ampliacao=Decimal(0)

        re_participacao_100h=Decimal(0)
        re_participacao_200h=Decimal(0)
        rt_participacao_100h=Decimal(0)
        rt_participacao_200h=Decimal(0)
        se_participacao_100h=Decimal(0)
        se_participacao_200h=Decimal(0)
        st_participacao_100h=Decimal(0)
        st_participacao_200h=Decimal(0)

        posgraduacao = qq['posgraduacao']
        id_funcao = qq['id_funcao']
        salario_int = qq['salario_100H_int']

        if [posgraduacao,id_funcao,salario_int] in lista_computados:
            continue
        lista_computados.append([posgraduacao,id_funcao,salario_int])


        funcao=dictFuncoes.get(id_funcao,'')
        if posgraduacao=='N':
            descfuncao=funcao+' ('+str(id_funcao)+')'
        else:
            descfuncao=funcao+' POS ('+str(id_funcao)+')'

        if id_funcao in lista_funcao_prof:
            tipo_funcao=21
        else:
            tipo_funcao=11

        for qqq in query:
            if qqq['posgraduacao']==posgraduacao and qqq['id_funcao']==id_funcao and qqq['salario_100H_int']==salario_int:
                suporte = qqq['suporte']
                grupo = qqq['grupo']
                salario = qqq['salario_100H']
                carga_horaria = qqq['carga_horaria']
                soma_participacao = qqq['somaParticipacao']
                string=posgraduacao+suporte+grupo+str(id_funcao)+str(salario_int) 

                if suporte=='N':
                    if grupo=='E':
                        if dictAmpliacao==1:
                            re_ampliacao=dictAmpliacao_01.get(string,Decimal(0))+dictAmpliacao_02.get(string,Decimal(0))
                        if carga_horaria==100:
                            re_participacao_100h=soma_participacao
                        else:
                            re_participacao_200h=soma_participacao
                    else:
                        if dictAmpliacao==1:
                            rt_ampliacao=dictAmpliacao_01.get(string,Decimal(0))+dictAmpliacao_02.get(string,Decimal(0))
                        if carga_horaria==100:
                            rt_participacao_100h=soma_participacao
                        else:
                            rt_participacao_200h=soma_participacao
                else:
                    if grupo=='E':
                        if dictAmpliacao==1:
                            se_ampliacao=dictAmpliacao_01.get(string,Decimal(0))+dictAmpliacao_02.get(string,Decimal(0))
                        if carga_horaria==100:
                            se_participacao_100h=soma_participacao
                        else:
                            se_participacao_200h=soma_participacao
                    else:
                        if dictAmpliacao==1:
                            st_ampliacao=dictAmpliacao_01.get(string,Decimal(0))+dictAmpliacao_02.get(string,Decimal(0))
                        if carga_horaria==100:
                            st_participacao_100h=soma_participacao
                        else:
                            st_participacao_200h=soma_participacao

        listao.append(
            [
            tipo_funcao,
            descfuncao,
            salario_int,
            salario,
            re_participacao_100h,
            re_participacao_200h,
            re_ampliacao,
            rt_participacao_100h,
            rt_participacao_200h,
            rt_ampliacao,
            se_participacao_100h,
            se_participacao_200h,
            se_ampliacao,
            st_participacao_100h,
            st_participacao_200h,
            st_ampliacao
            ]
        )

    return listao


def tabelaParticipacaoExcel(request):
    sessao(request)
    #current_user=request.user.id
    current_user=1

    titulo='Tabela Regente/Suporte/Matriz'
    municipios = Municipio.objects.filter(empresa__in=['SS','Layout','Aspec']).order_by('municipio')
    if (request.method == "POST"):
        id_municipio=int(request.POST['municipio'])
        ano=request.POST['ano']
        mes=request.POST['mes']
        tabela='MATRIZ'
        tabela=tabela.upper()
        anomes=int(str(ano)+str(int(mes)+100)[-2:])
        if tabela=='MATRIZ':
            return tabelaRegenteSuporteMatrizExcel(id_municipio,anomes,tabela,current_user)

    return render(request, 'app01/tabelaParticipacao.html',
            {
                'titulo': titulo,
                'municipios':municipios,
            }
        )


def styleRegente(fx1,fx2,ws,tabela):
    col_style = Font(name="Calibri", size=12, color="00008B", underline="none", strike=False)
    for lin  in range(fx1[0][0], fx1[0][1]+2):
        for col in range(2,9):
            ws.cell(row=lin,column=col).font = col_style
    for lin  in range(fx2[0][0], fx2[0][1]+2):
        for col in range(2,9):
            ws.cell(row=lin,column=col).font = col_style

    listaMarronFFDEAD=[]
    listaAzul6495ED=[]
    if tabela=='MATRIZ':
        max_col=6
    else:
        max_col=12
    for row_num,row_val in enumerate(ws.iter_rows(max_col=1,min_row=2,max_row=ws.max_row)):
        if ws.cell(row=row_num+2,column=1).value in ['F u n c a o','E v e n t o']:
            listaAzul6495ED.append(row_num+2)
        elif ws.cell(row=row_num+2,column=1).value in ['S o m a']:
            listaMarronFFDEAD.append(row_num+2)


    #ftAzul=Font(color='6495ED')
    #ftMarron=Font(color='FFDEAD')

    for k in listaAzul6495ED:
        for col_range in range(1, max_col):
            cell_title = ws.cell(k, col_range)
            cell_title.fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")
            cell_title.alignment = Alignment(horizontal='center')
            cell_title.font = Font(name='Calibri',size=14)

    for k in listaMarronFFDEAD:
        for col_range in range(1, max_col):
            cell_title = ws.cell(k, col_range)
            cell_title.fill = PatternFill(start_color="FFDEAD", end_color="FFDEAD", fill_type="solid")
            cell_title.alignment = Alignment(horizontal='center')
            cell_title.font = Font(name='Calibri',size=14)
#========================================================================================================================

def styleMatriz(fx1,fx2,ws):
    col_style = Font(name="Calibri", size=12, color="00008B", underline="none", strike=False)
    for fx in fx1:
        for lin  in range(fx[0], fx[1]):
            for col in range(2,5):
                ws.cell(row=lin,column=col).font = col_style
    for fx in fx2:
        for lin  in range(fx[0], fx[1]):
            for col in range(2,9):
                ws.cell(row=lin,column=col).font = col_style

    #ftAzul=Font(color='6495ED')
    #ftMarron=Font(color='FFDEAD')

    listaMarronFFDEAD=[]
    listaAzul6495ED=[]
    max_col=11
    for row_num,row_val in enumerate(ws.iter_rows(max_col=1,min_row=2,max_row=ws.max_row)):
        if ws.cell(row=row_num+2,column=1).value in ['F u n c a o']:
            listaAzul6495ED.append([row_num+2,11])
        elif ws.cell(row=row_num+2,column=1).value in ['* S o m a']:
            listaMarronFFDEAD.append([row_num+2,11])
        elif ws.cell(row=row_num+2,column=1).value in ['N i v e l']:
            listaAzul6495ED.append([row_num+2,7])
        elif ws.cell(row=row_num+2,column=1).value in ['S o m a']:
            listaMarronFFDEAD.append([row_num+2,7])


    for k in listaAzul6495ED:
        max_col=k[1]
        for col_range in range(1, max_col):
            cell_title = ws.cell(k[0], col_range)
            cell_title.fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")
            cell_title.alignment = Alignment(horizontal='center')
            cell_title.font = Font(name='Calibri',size=14)

    for k in listaMarronFFDEAD:
        max_col=k[1]
        for col_range in range(1, max_col):
            cell_title = ws.cell(k[0], col_range)
            cell_title.fill = PatternFill(start_color="FFDEAD", end_color="FFDEAD", fill_type="solid")
            cell_title.alignment = Alignment(horizontal='center')
            cell_title.font = Font(name='Calibri',size=14)


def styleMatrizSintetico(ws):
    listaMarronFFDEAD=[]
    listaAzul6495ED=[]
    col_style = Font(name="Calibri", size=12, color="00008B", underline="none", strike=False)
    for lin  in range(3, 9):
            for col in range(9,19):
                ws.cell(row=lin,column=col).font = col_style

    for row_num,row_val in enumerate(ws.iter_rows(max_col=1,min_row=2,max_row=ws.max_row)):
        if ws.cell(row=row_num+2,column=9).value in ['E v e n t o']:
            listaAzul6495ED.append([row_num+2,19])
        elif ws.cell(row=row_num+2,column=9).value in ['S o m a']:
            listaMarronFFDEAD.append([row_num+2,19])

    for k in listaAzul6495ED:
        max_col=k[1]
        for col_range in range(9, max_col):
            cell_title = ws.cell(k[0], col_range)
            cell_title.fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")
            cell_title.alignment = Alignment(horizontal='center')

    top = Side(border_style='thick',color='6495ED')

    for k in listaMarronFFDEAD:
        max_col=k[1]
        for col_range in range(9, max_col):
            cell_title = ws.cell(k[0], col_range)
            cell_title.fill = PatternFill(start_color="FFDEAD", end_color="FFDEAD", fill_type="solid")
            cell_title.alignment = Alignment(horizontal='center')

    for kl in range(3,9):
        for col_range in range(9, 19):
            cell_title = ws.cell(kl, col_range)
            cell_title.border = Border(top=top, left=top, right=top, bottom=top)



def tabelaRegenteSuporteMatrizExcel(id_municipio,anomes,tabela,current_user):
    ano=str(anomes)[0:4]
    obj=Municipio.objects.get(id_municipio=id_municipio)
    entidade=obj.entidade
    mesEano=str(anomes)[-2:]+'/'+str(anomes)[0:4]

    listao = coletaDados_novo(id_municipio,anomes)

    (listao_soma,resumo_soma)=lista_SomaEventos_Geral(id_municipio,anomes)
    listao_soma.sort()

    listao.sort()
    dados=[]
    lista=[]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Regente"
    ws2 = wb.create_sheet("Suporte")
    ws3 = wb.create_sheet("Matriz")
    wsSimulacao = wb.create_sheet("Simulacao")

    v100=Decimal(0)
    v200=Decimal(0)
    vAmpliacao=Decimal(0)
    vVencimentoBase=Decimal(0)

    f1=Decimal(0)
    f2=Decimal(0)
    f3=Decimal(0)

    nlinha=5
    fx1=[]
    fx2=[]
    fy1=[]
    fy2=[]

    for op_tabela in ['REGENTE','SUPORTE']:
        if op_tabela=='REGENTE':
            ws=ws1
        else:
            ws=ws2

        (lista_somaE,lista_somaT) = lista_Soma_Eventos(id_municipio,anomes,op_tabela)    
        inicio=0
        final=0
        relatorio=[]
        resumo=[]

        dict_relatorio = {
            'efetivo':'Tabela: '+op_tabela.upper()+' Efetivos - Referencia: '+str(anomes)[-2:]+'/'+str(anomes)[0:4],
            'temporario':'Tabela: '+op_tabela.upper()+' Temporario - Referencia: '+str(anomes)[-2:]+'/'+str(anomes)[0:4]
        }

        dict_resumo = {
            'efetivo':'Resumo dos Eventos - '+op_tabela.upper()+ ' Efetivos',
            'temporario':'Resumo dos Eventos - '+op_tabela.upper()+ ' Temporario'
        } 

        for grupo_vinculo in ['Efetivo','Temporario']:
            if grupo_vinculo=='Efetivo':
                n_linha=996
            else:
                n_linha=inicio+996
            f1=0
            f2=0
            f3=0
            dados=[]
            for k in listao:
                funcao=k[1]
                vVencimentoBase=k[3]
                if op_tabela=='REGENTE':
                    if grupo_vinculo=='Efetivo':
                        if k[4]+k[5]+k[6]==0:
                            continue
                        v100=k[4]
                        v200=k[5]
                        vAmpliacao=k[6]
                        desc_tipo='Regente-Efetivo-'+entidade+'/'+mesEano
                    else:
                        if k[7]+k[8]+k[9]==0:
                            continue
                        v100=k[7]
                        v200=k[8]
                        vAmpliacao=k[9]
                        desc_tipo='Regente-Temporario-'+entidade+'/'+mesEano
                elif op_tabela=='SUPORTE':
                    if grupo_vinculo=='Efetivo':
                        if k[10]+k[11]+k[12]==0:
                            continue
                        v100=k[10]
                        v200=k[11]
                        vAmpliacao=k[12]
                        desc_tipo='Suporte-Efetivo-'+entidade+'/'+mesEano
                    else:
                        if k[13]+k[14]+k[15]==0:
                            continue
                        v100=k[13]
                        v200=k[14]
                        vAmpliacao=k[15]
                        desc_tipo='Suporte-Temporario-'+entidade+'/'+mesEano
                dados.append([funcao,v100,v200,vAmpliacao,'','',vVencimentoBase,f3,'',desc_tipo])

            # add column headings. NB. these must be strings
            ws.append([""])
            ws.append([entidade])
            if grupo_vinculo=='Efetivo':
                ws.append([dict_relatorio['efetivo']])
                str_chave1=['* F u n c a o','* S o m a']
            else:
                ws.append([dict_relatorio['temporario']])
                str_chave1=['** F u n c a o','** S o m a']

            ws.append([str_chave1[0], "Total 100", "Total 200", "Ampliacao", "Total 100","Total","Sal 100",'Total'])
            for row in dados:
                ws.append(row)
            ws.append([str_chave1[1]])
            ws.append([''])
            ws.append([''])
            ws.append([''])

            imprimeFormulaParticipacao(ws,str_chave1[0],str_chave1[1])
            imprimeSoma(ws,str_chave1[0],str_chave1[1],['B','C','D','E','F','H'])
            if grupo_vinculo=='Efetivo':
                ws.append([dict_resumo['efetivo']])
                str_chave2=['*** E v e n t o','*** S o m a']
                lista_soma_eventos=lista_somaE
            else:
                ws.append([dict_resumo['temporario']])
                str_chave2=['**** E v e n t o','**** S o m a']
                lista_soma_eventos=lista_somaT

            ws.append([str_chave2[0],'Valor'])

            pula=1
            for k in lista_soma_eventos:
                if pula!=k[0]:
                    ws.append([''])
                    pula=k[0]

                if k[1]=='carater-salario':
                    tipo_evento=' (Caráter Salário)'
                elif k[1]=='posgraduacao':
                    tipo_evento=' (Pós)'
                elif k[1]=='suporte':
                    tipo_evento=' (Suporte)'
                else:
                    tipo_evento=''

                ws.append([k[2]+tipo_evento,k[3]])
            ws.append([str_chave2[1]])
            ws.append([''])
            ws.append([''])
            ws.append([''])

            imprimeSoma(ws,str_chave2[0],str_chave2[1],['B'])


    col_style = Font(name="Calibri", size=12, color="00008B", underline="none", strike=False)

    stylePadrao_1(ws1)
    stylePadrao_1(ws2)
    stylePadrao_2(ws1,11)
    stylePadrao_2(ws2,11)
    styleSimulacao_01(ws1)
    styleSimulacao_01(ws2)


    # inicio tabel Matriz

    primeira_parte=['Regentes e Suporte Efetivo']
    segunda_parte=['Regentes e Suporte Temporário']
    desc_tipo='Matriz - '+entidade+' - '+mesEano


    inicio=4
    final=3
    fx1=[]
    fx2=[]
    for kl in range(2):
        kk1=Decimal(0)
        kk2=Decimal(0)
        dados=[]
        if kl==0:
            titulo_regente=primeira_parte[0]
            dados.append(['Tabela Matriz'])
            dados.append([entidade+' - '+mesEano+' -Regente e Suporte Efetivos'])
            dados.append(['* N i v e l','Regente Efetivo Total 100h','Suporte Efetivo Total 100h','Vencimento Base'])
            str_chave1=['* N i v e l','* S o m a']
        else:
            titulo_regente=segunda_parte[0]
            dados.append(['Tabela Matriz'])
            dados.append([entidade+' - '+ mesEano +' - Regente e Suporte Temporarios'])
            dados.append(['** N i v e l','Regente Efetivo Total 100h','Suporte Efetivo Total 100h','Vencimento Base'])
            str_chave1=['** N i v e l','** S o m a']

        for kreg in listao:
            if kl==0:
                kk1=kreg[4]+2*kreg[5]+kreg[6]
                kk2=kreg[10]+2*kreg[11]+kreg[12]

                if kreg[4]+kreg[5]+kreg[6]+kreg[10]+kreg[11]+kreg[12]==0:
                    continue
                dados.append([kreg[1],kk1,kk2,kreg[3],'',desc_tipo])
            else:

                kk1=kreg[7]+2*kreg[8]+kreg[9]
                kk2=kreg[13]+2*kreg[14]+kreg[15]
                if kreg[7]+kreg[8]+kreg[9]+kreg[13]+kreg[14]+kreg[15]==0:
                    continue
                dados.append([kreg[1],kk1,kk2,kreg[3],'',desc_tipo])
            final+=1

        for row in dados:
            ws3.append(row)

        if kl==0:
            str_soma='* S o m a'
            fx1.append([inicio,final])
        else:
            str_soma='** S o m a'
            fx1.append([inicio,final])
        ws3.append([str_soma])
        ws3.append([''])
        ws3.append([''])
        ws3.append([''])
        imprimeSoma(ws3,str_chave1[0],str_chave1[1],['B','C'])

        inicio=final+6
        final+=5

    #------------------------------------------------------------------------------------------------
    ws3.append([''])
    ws3.append([''])
    ws3.append([''])
    ws3.append(['Tabela Matriz'])
    ws3.append([entidade+' - '+mesEano+' -Regente e Suporte'])
    ws3.append(['Resumo Geral'])
    ws3.append(['*** F u n c a o','Quantidade Efetivo','Vencto. Base','Custo Efetivo','','Quantidade Temporário','Vencto. Base','Custo Temporário'])
 
    dados=[]
    lista=[]
    final=final+4
    inicio=final+1
    for kreg in listao:
        (tot1,custo1,tot2,custo2) = fun_Auxiliar_tabelaMatriz(kreg)
        final+=1
        dados.append([kreg[1],tot1,kreg[3],custo1,'',tot2,kreg[3],custo2,'',desc_tipo])

    for row in dados:
        ws3.append(row)

    fx2.append([inicio,final])
    ws3.append(['***** S o m a'])
    str_chave2=['*** F u n c a o','***** S o m a']


    imprimeSoma(ws3,str_chave2[0],str_chave2[1],['B','D','F','G'])


    ws3.cell(row=3,column=9).value='E_v_e_n_t_o'
    ws3.cell(row=3,column=10).value='Regente Efetivo'
    ws3.cell(row=3,column=11).value='Regente Temporario'
    ws3.cell(row=3,column=12).value='Soma'
    ws3.cell(row=3,column=15).value='Suporte Efetivo'
    ws3.cell(row=3,column=16).value='Suporte Temporario'
    ws3.cell(row=3,column=17).value='Soma'
    ws3.cell(row=3,column=18).value='Total'


    nl=3
    ni=4
    nf=0
    ncol=9
    num=len(resumo_soma)
    list1=[9,10,11,12,13,14,15,16,17,18,19]
    list2=['I','J','K','L','M','N','O','P','Q','R','S']
    dict_col=dict(zip(list1,list2))
    for k in resumo_soma:
        nl+=1
        nf+=1
        if nf<num:
            ws3.cell(row=nl,column=ncol).value=k[1]
            ws3.cell(row=nl,column=ncol+1).value=k[2]
            ws3.cell(row=nl,column=ncol+2).value=k[3]
            icol=dict_col.get(ncol+1)
            fcol=dict_col.get(ncol+2)
            #ws3.cell(row=nl,column=ncol+3).value="=SUM("+icol+str(nl)+":"+fcol+str(nl)+")"
            ws3.cell(row=nl,column=ncol+3).value="={}{}+{}{}".format(icol,nl,fcol,nl)

            ws3.cell(row=nl,column=ncol+6).value=k[5]
            ws3.cell(row=nl,column=ncol+7).value=k[6]
            icol=dict_col.get(ncol+6)
            fcol=dict_col.get(ncol+7)
            #ws3.cell(row=nl,column=ncol+8).value="=SUM("+icol+str(nl)+":"+fcol+str(nl)+")"
            ws3.cell(row=nl,column=ncol+8).value="={}{}+{}{}".format(icol,nl,fcol,nl)
            icol=dict_col.get(ncol+3)
            fcol=dict_col.get(ncol+8)
            #ws3.cell(row=nl,column=ncol+9).value="="+icol+str(nl)+"+"+fcol+str(nl)
            ws3.cell(row=nl,column=ncol+9).value="={}{}+{}{}".format(icol,nl,fcol,nl)

    ws3.cell(row=nl,column=ncol).value='S_o_m_a'
    imprimeSoma2(ws3,'E_v_e_n_t_o','S_o_m_a',['J','K','L','O','P','Q','R'],9)

    stylePadrao_2(ws3,7)
    stylePadrao_3(ws3,19)
    stylePadrao_1(ws3)

    formatGeral(ws1)
    formatGeral(ws2)
    formatGeral(ws3)

    #=======================================================================================================
    #=======================================================================================================
    #  tabela simulacao
    #=======================================================================================================
    #=======================================================================================================
    wsSimulacao['A3'].value='Efetivos'
    wsSimulacao.merge_cells('A3:H3')
    wsSimulacao['B4'].value='Situacao Atual'
    wsSimulacao.merge_cells('B4:D4')
    wsSimulacao['E4'].value='Situacao Proposta'
    wsSimulacao.merge_cells('E4:H4')

    wsSimulacao['A5'].value='* N i v e l'
    wsSimulacao['B5'].value='Quantidade'
    wsSimulacao['C5'].value='Vencto. Base'
    wsSimulacao['D5'].value='Custo'
    wsSimulacao['E5'].value='Quantidade'
    wsSimulacao['F5'].value='Venc.Base'
    wsSimulacao['G5'].value='Custo'
    wsSimulacao['H5'].value='Fator'

    faixa=faixaDeMatrizParaSimulacao(ws3,'*** F u n c a o','***** S o m a')
    inicio=faixa[0]
    final=faixa[1]
    rr=5
    for r in range(inicio,final+1):
        rr+=1
        const=1.08
        #print ('r: '+str(r)+'   -  rr:'+str(rr))
        wsSimulacao['A{}'.format(rr)].value='=MATRIZ!A{}'.format(r)
        wsSimulacao['B{}'.format(rr)].value='=MATRIZ!B{}'.format(r)
        wsSimulacao['C{}'.format(rr)].value='=MATRIZ!C{}'.format(r)
        wsSimulacao['D{}'.format(rr)].value='=B{}*C{}'.format(rr,rr)
        wsSimulacao['E{}'.format(rr)].value='=B{}'.format(rr)
        wsSimulacao['F{}'.format(rr)].value='=C{}*{}'.format(rr,const)
        wsSimulacao['G{}'.format(rr)].value='=F{}*E{}'.format(rr,rr)
        wsSimulacao['H{}'.format(rr)].value='=F{}/C{}'.format(rr,rr)
    rr+=1        
    wsSimulacao['A{}'.format(rr)].value='* S o m a'

    faixa=faixaDeMatrizParaSimulacao(wsSimulacao,'* N i v e l','* S o m a')
    inicio1=faixa[0]
    final1=faixa[1]
    wsSimulacao['B{}'.format(rr)].value='=SUM(B{}:B{})'.format(inicio1,final1)
    wsSimulacao['D{}'.format(rr)].value='=SUM(D{}:D{})'.format(inicio1,final1)
    wsSimulacao['E{}'.format(rr)].value='=SUM(E{}:E{})'.format(inicio1,final1)    
    wsSimulacao['G{}'.format(rr)].value='=SUM(G{}:G{})'.format(inicio1,final1)    

    #---------------------------------------------------------------------

    rr=rr+5
    wsSimulacao['A{}'.format(rr)].value='Temporarios'
    wsSimulacao.merge_cells('A{}:H{}'.format(rr,rr))
    rr+=1
    wsSimulacao['B{}'.format(rr)].value='Situacao Atual'
    wsSimulacao.merge_cells('B{}:D{}'.format(rr,rr))
    wsSimulacao['E{}'.format(rr)].value='Situacao Proposta'
    wsSimulacao.merge_cells('E{}:H{}'.format(rr,rr))
    rr+=1

    wsSimulacao['A{}'.format(rr)].value='** N i v e l'
    wsSimulacao['B{}'.format(rr)].value='Quantidade'
    wsSimulacao['C{}'.format(rr)].value='Vencto. Base'
    wsSimulacao['D{}'.format(rr)].value='Custo'
    wsSimulacao['E{}'.format(rr)].value='Quantidade'
    wsSimulacao['F{}'.format(rr)].value='Venc.Base'
    wsSimulacao['G{}'.format(rr)].value='Custo'
    wsSimulacao['H{}'.format(rr)].value='Fator'


    for r in range(inicio,final+1):
        rr+=1
        const=1.08
        wsSimulacao['A{}'.format(rr)].value='=MATRIZ!A{}'.format(r)
        wsSimulacao['B{}'.format(rr)].value='=MATRIZ!F{}'.format(r)
        wsSimulacao['C{}'.format(rr)].value='=MATRIZ!G{}'.format(r)
        wsSimulacao['D{}'.format(rr)].value='=B{}*C{}'.format(rr,rr)
        wsSimulacao['E{}'.format(rr)].value='=B{}'.format(rr)
        wsSimulacao['F{}'.format(rr)].value='=C{}*{}'.format(rr,const)
        wsSimulacao['G{}'.format(rr)].value='=F{}*E{}'.format(rr,rr)
        wsSimulacao['H{}'.format(rr)].value='=F{}/C{}'.format(rr,rr)
    rr+=1
    wsSimulacao['A{}'.format(rr)].value='** S o m a'
    faixa=faixaDeMatrizParaSimulacao(wsSimulacao,'** N i v e l','** S o m a')
    inicio=faixa[0]
    final=faixa[1]

    wsSimulacao['B{}'.format(rr)].value='=SUM(B{}:B{})'.format(inicio,final)
    wsSimulacao['D{}'.format(rr)].value='=SUM(D{}:D{})'.format(inicio,final)
    wsSimulacao['E{}'.format(rr)].value='=SUM(E{}:E{})'.format(inicio,final)    
    wsSimulacao['G{}'.format(rr)].value='=SUM(G{}:G{})'.format(inicio,final)    


    wsSimulacao.append([''])
    wsSimulacao.append([''])
    wsSimulacao.append(['Componentes da Folha Analisada'])
    wsSimulacao.append(['','E f e t i v o','','T e m p o r a r i o',''])
    wsSimulacao.append(['.Descricao','Atual','Proposta','Atual','Proposta'])
    for k in listao_soma:
        wsSimulacao.append([k[2],k[4]+k[6],'',k[5]+k[7],''])
    wsSimulacao.append(['Sub-total'])

    ###imprimeSoma(wsSimulacao,'.Descricao','.Subtotal',['B','C','D','E'])    
    styleResumoSimulacao(wsSimulacao,'Efetivos','* N i v e l',1,1,1,6,True,'87CEEB',True,'center')    
    styleResumoSimulacao(wsSimulacao,'Temporarios','** N i v e l',1,1,1,6,True,'87CEEB',True,'center')    

    styleResumoSimulacao(wsSimulacao,'Componentes da Folha Analisada','.Descricao',1,1,1,6,True,'87CEEB',True,'center')    
    styleResumoSimulacao(wsSimulacao,'Sub-total','',1,1,1,6,True,'FFEFD5',False,'center')    

    funcaoMerge(wsSimulacao,'Componentes da Folha Analisada',1,'A','E',1,5)
    funcaoMerge(wsSimulacao,'E f e t i v o',2,'B','C',2,3)
    funcaoMerge(wsSimulacao,'T e m p o r a r i o',4,'D','E',4,5)

    stylePadrao_2(wsSimulacao,9)
    stylePadrao_1(wsSimulacao)
    formatGeral(wsSimulacao)
    imprimeSoma2(wsSimulacao,'.Descricao','Sub-total',['B','C','D','E'],1)
    styleSimulacao_01(wsSimulacao)



    str_data_hora= f_funcoes.fun_datahora()
    caminho='/home/civitas/documentos/'
    #caminho='c:/projetos/'
    nome_do_arquivo=entidade+'_matriz_'+str(anomes)+'_'+str_data_hora+'.xlsx'


    #if not os.objects.filter(id_municipio=id_municipio,anomes=anomes,nome_do_arquivo=nome_do_arquivo).exists():
    wb.save(caminho+nome_do_arquivo)
    Documento.objects.create(id_municipio=id_municipio,ano=ano,anomes=anomes,nome_do_arquivo=nome_do_arquivo,tipo='matriz',id_user=current_user)
    obj=Documento.objects.filter(id_municipio=id_municipio,ano=ano,anomes=anomes,nome_do_arquivo=nome_do_arquivo).last()
    if obj:
        id_arquivo=obj.id_documento
    else:
        id_arquivo=0
        return HttpResponse("<h1>Arquivo salvo com sucesso!</h1>")

    return redirect ('app01:abrirRegenteSuporteMatriz', id_arquivo=id_arquivo)




def stylePadrao_1(sheet):
    for rows in sheet.iter_cols(min_col=1, max_col=20, min_row=2, max_row=None):
        for cell in rows:
            cell.font = Font(size=12)


def stylePadrao_2(ws,max_col):
    col_style = Font(name="Calibri", size=12, color="00008B", underline="none", strike=False)
    for row_num,row_val in enumerate(ws.iter_rows(max_col=1,min_row=2,max_row=ws.max_row)):
        if ws.cell(row=row_num+2,column=1).value in ['* F u n c a o','*** E v e n t o','** F u n c a o','**** E v e n t o','* N i v e l','** N i v e l']:
            for col in range(1,max_col):
                if ws.cell(row=row_num+2,column=col).value!='':
                    ws.cell(row=row_num+2,column=col).fill = PatternFill(start_color="ADD8EC", end_color="ADD8EC", fill_type="solid")
                    ws.cell(row=row_num+2,column=col).alignment = Alignment(horizontal='center')
        elif ws.cell(row=row_num+2,column=1).value in ['* S o m a','*** S o m a','** S o m a','**** S o m a','* S o m a','** S o m a']:
            for col in range(1,max_col):
                if ws.cell(row=row_num+2,column=col).value!='':
                    ws.cell(row=row_num+2,column=col).fill = PatternFill(start_color="FFDEAD", end_color="FFDEAD", fill_type="solid")
                #ws.cell(row=row_num+2,column=col).alignment = Alignment(horizontal='center')
        elif ws.cell(row=row_num+2,column=1).value in ['*** F u n c a o']:
            for col in range(1,max_col+4):
                if ws.cell(row=row_num+2,column=col).value!='':
                    ws.cell(row=row_num+2,column=col).fill = PatternFill(start_color="ADD8EC", end_color="ADD8EC", fill_type="solid")
                    ws.cell(row=row_num+2,column=col).alignment = Alignment(horizontal='center')
        elif ws.cell(row=row_num+2,column=1).value in ['***** S o m a']:
            for col in range(1,max_col+4):
                if ws.cell(row=row_num+2,column=col).value!='':
                    ws.cell(row=row_num+2,column=col).fill = PatternFill(start_color="FFDEAD", end_color="FFDEAD", fill_type="solid")


def faixaDeMatrizParaSimulacao(ws,str_inicio,str_final):
    row_inicio=1
    row_final=1
    for row_num,row_val in enumerate(ws.iter_rows(max_col=1,min_row=2,max_row=ws.max_row)):
        if ws.cell(row=row_num+2,column=1).value==str_inicio:
            row_inicio=row_num+2
        elif ws.cell(row=row_num+2,column=1).value==str_final:
            row_final=row_num+2
    return [row_inicio+1,row_final-1]



def stylePadrao_3(ws,max_col):
    col_style = Font(name="Calibri", size=12, color="00008B", underline="none", strike=False)
    for row_num,row_val in enumerate(ws.iter_rows(max_col=1,min_row=2,max_row=ws.max_row)):
        if ws.cell(row=row_num+2,column=9).value in ['E_v_e_n_t_o']:
            for col in range(9,max_col):
                ws.cell(row=row_num+2,column=col).fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                ws.cell(row=row_num+2,column=col).alignment = Alignment(horizontal='center')
        elif ws.cell(row=row_num+2,column=9).value in ['S_o_m_a']:
            for col in range(9,max_col):
                ws.cell(row=row_num+2,column=col).fill = PatternFill(start_color="FFDEAD", end_color="FFDEAD", fill_type="solid")

#========================================================================================================================

def styleMatriz_01(fx1,fx2,ws,col1,col2):
    font_style = Font(name="Calibri", size=14, color="00008B", underline="none", strike=False)
    alignment_style = Alignment(horizontal='center')
    for fx in fx1:
        for lin  in range(fx[0], fx[1]):
            for col in range(col1,col2):
                ws.cell(row=lin,column=col).font = font_style
                ws.cell(row=lin,column=col).alignment = alignment_style


def styleSimulacao_01(ws):
    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column+1, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
    col=ws.min_column
    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=60)
    ws.column_dimensions = dim_holder


    listaMarronFFDEAD=[]
    listaAzul6495ED=[]
    max_col=11
    for row_num,row_val in enumerate(ws.iter_rows(max_col=1,min_row=2,max_row=ws.max_row)):
        if ws.cell(row=row_num+2,column=1).value in ['F u n c a o']:
            listaAzul6495ED.append([row_num+2,11])
        elif ws.cell(row=row_num+2,column=1).value in ['* S o m a']:
            listaMarronFFDEAD.append([row_num+2,10])
        elif ws.cell(row=row_num+2,column=1).value in ['N i v e l']:
            listaAzul6495ED.append([row_num+2,7])
        elif ws.cell(row=row_num+2,column=1).value in ['S o m a']:
            listaMarronFFDEAD.append([row_num+2,7])


    for k in listaAzul6495ED:
        max_col=k[1]
        for col_range in range(1, max_col):
            cell_title = ws.cell(k[0], col_range)
            cell_title.fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")
            cell_title.alignment = Alignment(horizontal='center')
            cell_title.font = Font(name='Calibri',size=12)

    for k in listaMarronFFDEAD:
        max_col=k[1]
        for col_range in range(1, max_col):
            cell_title = ws.cell(k[0], col_range)
            cell_title.fill = PatternFill(start_color="FFDEAD", end_color="FFDEAD", fill_type="solid")
            #cell_title.alignment = Alignment(horizontal='center')
            cell_title.font = Font(name='Calibri',size=12)


def fun_range(wsheet):
    lista1=[]
    lista2=[]
    lista3=[]
    lista4=[]
    lista=[]
    k=0
    for row_num,row_val in enumerate(wsheet.iter_rows(max_col=1,min_row=2,max_row=wsheet.max_row)):
        if wsheet.cell(row=row_num+2,column=1).value=='*/':
            k+=1
            if k==1 or k==2:
                lista1.append(row_num+2)
            elif k==3 or k==4:
                lista2.append(row_num+2)
            elif k==5 or k==6:
                lista3.append(row_num+2)
            elif k==7 or k==8:
                lista4.append(row_num+2)
    lista.append(lista1)
    lista.append(lista2)
    lista.append(lista3)
    lista.append(lista4)
    return lista

def formatCelulas(ws,row_inicio,row_fim,max_col):
    for k in range(row_inicio,row_fim+1):
        for col_range in range(1, max_col):
            cell_title = ws.cell(k, col_range)
            cell_title.font = Font(name='Calibri',size=12)

def imprimeSoma(wsheet,chv1,chv2,list_col):
    inicio=1
    final=1
    for row_num,row_val in enumerate(wsheet.iter_rows(max_col=1,min_row=2,max_row=wsheet.max_row)):
        if wsheet.cell(row=row_num+2,column=1).value==chv1:
            inicio=row_num+2
        elif wsheet.cell(row=row_num+2,column=1).value==chv2:
            final=row_num+2
            break

    lista=['','A','B','C','D','E','F','G','H','I','J']

    for letra in list_col:
        numCol = lista.index(letra)
        #wsheet.cell(row=final,column=numCol).value="=SUM("+letra+str(inicio+1)+":"+letra+str(final-1)+")"
        wsheet.cell(row=final,column=numCol).value="=SUM({}{}:{}{})".format(letra,inicio+1,letra,final-1)


def imprimeSoma2(wsheet,chv1,chv2,list_col,n_col):
    inicio=1
    final=1
    for row_num,row_val in enumerate(wsheet.iter_rows(max_col=1,min_row=2,max_row=wsheet.max_row)):
        if wsheet.cell(row=row_num+2,column=n_col).value==chv1:
            inicio=row_num+2
        elif wsheet.cell(row=row_num+2,column=n_col).value==chv2:
            final=row_num+2
            break

    lista=['','A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']

    for letra in list_col:
        numCol = lista.index(letra)
        #wsheet.cell(row=final,column=numCol).value="=SUM("+letra+str(inicio+1)+":"+letra+str(final-1)+")"
        wsheet.cell(row=final,column=numCol).value="=SUM({}{}:{}{})".format(letra,inicio+1,letra,final-1)

def imprimeFormulaParticipacao(wsheet,chv1,chv2):
    inicio=1
    final=1
    for row_num,row_val in enumerate(wsheet.iter_rows(max_col=1,min_row=2,max_row=wsheet.max_row)):
        if wsheet.cell(row=row_num+2,column=1).value==chv1:
            inicio=row_num+2
        elif wsheet.cell(row=row_num+2,column=1).value==chv2:
            final=row_num+2
            break

    for lr in range(inicio+1,final):
        wsheet.cell(row=lr,column=5).value="=B{}+C{}+D{}".format(lr,lr,lr)
        wsheet.cell(row=lr,column=6).value="=B{}+2*C{}+D{}".format(lr,lr,lr)
        wsheet.cell(row=lr,column=8).value="=F{}*G{}".format(lr,lr)

def formatGeral(ws):
    for letter in ['A']:
        max_width = 0
        for row_number in range(1,ws.max_row+1):
            if ws[f'{letter}{row_number}'].value is not None:
                if len(ws[f'{letter}{row_number}'].value) > max_width:
                    max_width = len(ws[f'{letter}{row_number}'].value)
        ws.column_dimensions[letter].width = max_width +1

    for letter in ['B','C','D','E','F','G','H']:
        ws.column_dimensions[letter].width = 20

    for letter in ['I','J','K','L','O','P','Q','R']:
        ws.column_dimensions[letter].width = 20


    for letter in ['B','C','D','E','F','G','H']:
        for row_number in range(1,ws.max_row+1):
            _cell = ws[f'{letter}{row_number}']
            _cell.number_format = '#,##0.00'

    for letter in ['I','J','K','L','O','P','Q','R']:
        for row_number in range(1,ws.max_row+1):
            _cell = ws[f'{letter}{row_number}']
            _cell.number_format = '#,##0.00'


def funcaoMerge(ws,str_chave,col_procura,col_inicio,col_final,n_col1,n_col2):
    inicio=10
    for row_num,row_val in enumerate(ws.iter_rows(max_col=1,min_row=2,max_row=ws.max_row)):
        if ws.cell(row=row_num+2,column=col_procura).value==str_chave:
            inicio=row_num+2
    #ws.merge_cells('A{}:E{}'.format(inicio,inicio))
    ws.merge_cells('{}{}:{}{}'.format(col_inicio,inicio,col_final,inicio))

    for i in range(n_col1,n_col2+1):
        ws.cell(row=inicio,column=i).alignment = Alignment(horizontal='center')


def styleResumoSimulacao(ws,str_chave1,str_chave2,n_col1,n_col2,n_col3,n_col4,p_fill,p_fillcolor,p_alignment,p_alignpositon):
    top = Side(style='thin',color='cb0e00')
    inicio=1
    final=1
    for row_num,row_val in enumerate(ws.iter_rows(max_col=1,min_row=2,max_row=ws.max_row)):
        if ws.cell(row=row_num+2,column=n_col1).value==str_chave1:
            inicio=row_num+2
            final=row_num+2
        elif str_chave2!='':
            if ws.cell(row=row_num+2,column=n_col2).value==str_chave2:
                final=row_num+2
                break
        if final>1 and str_chave2=='':
            break
    #azul='87CEEB
    for kl in range(inicio,final+1):
        for col in range(n_col3,n_col4):
            cell_title = ws.cell(kl, col)
            cell_title.border = Border(top=top, left=top, right=top, bottom=top)
            if p_fill:
                cell_title.fill = PatternFill(fill_type='solid',start_color=p_fillcolor,end_color=p_fillcolor)
            if p_alignment:
                cell_title.alignment = Alignment(horizontal=p_alignpositon)





